"""Import local product materials and script sheets into the app database."""
from __future__ import annotations

import argparse
import json
import re
import shutil
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any


PRODUCT_CATEGORY_MAP = {
    "袋装刀叉": "烘焙配件",
    "盒装刀叉": "烘焙配件",
    "3元盒装刀叉": "烘焙配件",
    "翻糖膏": "烘焙装饰",
    "彩色翻糖片": "烘焙装饰",
    "色粉盘": "烘焙调色",
    "拉线膏": "烘焙装饰",
    "手绘拉线膏": "烘焙装饰",
    "豆沙奶油霜": "烘焙装饰",
    "果蔬色素": "烘焙调色",
    "高浓果蔬色素": "烘焙调色",
    "水性色素": "烘焙调色",
    "水状色素": "烘焙调色",
    "油性色素": "烘焙调色",
    "水溶色粉": "烘焙调色",
    "油溶色粉": "烘焙调色",
    "竹炭粉": "烘焙调色",
    "果蔬粉": "烘焙调色",
    "红曲粉": "烘焙调色",
    "布蕾粉": "烘焙夹心",
    "奶冻粉": "烘焙夹心",
    "慕斯粉": "烘焙夹心",
    "巧克力糖": "烘焙装饰",
    "糖珠": "烘焙装饰",
    "肉松": "烘焙调味",
    "香草荚": "烘焙调味",
    "香草精": "烘焙调味",
    "红丝绒香精": "烘焙调味",
    "杏仁粉": "烘焙调味",
    "杏仁片": "烘焙调味",
    "抹茶粉": "烘焙调味",
    "可可粉": "烘焙调味",
    "斑斓粉": "烘焙调味",
    "开心果酱": "烘焙调味",
    "0卡糖粉": "烘焙调味",
}

STANDARD_SERIES_CATEGORY_MAP = {
    "调味系列": "烘焙调味",
    "夹心系列": "烘焙夹心",
    "调色系列": "烘焙调色",
    "装饰系列": "烘焙装饰",
    "工具系列": "烘焙配件",
    "刀叉配件系列": "烘焙配件",
    "包装系列": "烘焙配件",
}

PRODUCT_CARD_SKIP_SHEETS = {"常见问题", "解决方案"}

EXCLUDED_PRODUCT_NAMES = {"调味奶酱", "调味花酱", "巧克力酱"}

PRODUCT_CARD_SHEET_NAME_OVERRIDES = {
    "果泥": "夹心果泥",
    "果馅（多肉）": "多肉果酱",
    "芋泥": "夹心芋泥",
    "薄荷糖浆": "调味糖浆",
    "脆皮酱": "巧克力脆皮酱",
    "红丝绒液": "红丝绒",
    "白色翻糖": "白色翻糖膏",
    "彩色翻糖": "彩色翻糖膏",
    "翻糖片": "翻糖压片",
    "慕斯粉": "慕斯粉（液）",
    "拉线膏": "彩色拉线膏",
    "Q 弹奶冻粉": "Q弹奶冻粉",
}

PRODUCT_CARD_POINT_LABELS = {
    "产品名称",
    "一句话卖点",
    "核心卖点",
    "产品亮点",
    "产品价值",
    "解决痛点",
    "应用场景",
    "使用场景",
    "用途",
    "用途简述",
    "使用方法",
    "规格",
    "保质期",
    "储存",
    "储存方式",
    "保存方式",
    "口味",
    "用户人群",
    "用户画像",
    "适用人群",
    "配料表",
}


SCRIPT_SOURCE = "资料/脚本生成.xlsx"


REQUESTED_VIDEO_TYPES = [
    "机制类",
    "痛点类",
    "需求类",
    "认知类",
    "达人分享类",
    "制作方便",
    "成本低",
    "对比类",
    "情绪类",
    "场景类",
]


SCRIPT_TYPE_MAP = {
    "机制": "机制类",
    "机制类": "机制类",
    "展会机制类": "机制类",
    "AI生成": "机制类",
    "痛点": "痛点类",
    "痛点类": "痛点类",
    "痛点激发": "痛点类",
    "需求": "需求类",
    "需求类": "需求类",
    "真实需求": "需求类",
    "认知": "认知类",
    "认知类": "认知类",
    "专家口播": "认知类",
    "分享类": "达人分享类",
    "达人分享": "达人分享类",
    "达人分享类": "达人分享类",
    "爆款翻拍": "达人分享类",
    "开箱体验": "达人分享类",
    "制作简单": "制作方便",
    "制作方便": "制作方便",
    "省成本": "成本低",
    "成本低": "成本低",
    "限时优惠": "成本低",
    "对比": "对比类",
    "对比类": "对比类",
    "测评对比": "对比类",
    "创意": "情绪类",
    "情感": "情绪类",
    "情绪": "情绪类",
    "情绪类": "情绪类",
    "情感共鸣": "情绪类",
    "调动情绪": "情绪类",
    "情景": "场景类",
    "场景": "场景类",
    "场景类": "场景类",
    "使用场景": "场景类",
    "使用场景展示": "场景类",
    "一包多用": "场景类",
    "纯产品展示": "场景类",
}


def _script_template(
    video_type: str,
    description: str,
    segments: list[dict[str, str]],
    hooks: list[str],
    ctas: list[str],
    example: str,
) -> dict[str, Any]:
    return {
        "name": f"{video_type}模板",
        "video_type": video_type,
        "description": description,
        "duration_range": "30-60s",
        "structure": {"segments": segments},
        "hook_templates": hooks,
        "cta_templates": ctas,
        "example_script": example,
    }


REQUESTED_SCRIPT_TEMPLATES = [
    _script_template(
        "机制类",
        "把产品机制、优惠机制或使用机制讲清楚，让用户知道为什么现在值得买。",
        [
            {"name": "机制钩子", "duration": "0-5s", "goal": "用规则或利益点引发好奇"},
            {"name": "机制拆解", "duration": "5-25s", "goal": "解释省心、省钱或好用的原因"},
            {"name": "产品承接", "duration": "25-45s", "goal": "把机制落到具体产品卖点"},
            {"name": "行动提示", "duration": "45-60s", "goal": "给出下单理由"},
        ],
        ["这个东西不是便宜，是它的机制真的划算。", "很多人没看懂这个产品好用在哪里，我给你拆一下。"],
        ["需要的直接点左下角。", "现在拍更划算，别等用到的时候才找。"],
        "【机制钩子】这个烘焙工具不是普通便宜，是用一次就知道省在哪。\n【机制拆解】它把切分、分装和摆盘合在一起，新手也不用反复返工。\n【产品承接】家用做甜品、打包出摊都能用，稳定又省时间。\n【行动提示】需要的点左下角，先备一套不亏。",
    ),
    _script_template(
        "痛点类",
        "先还原用户在烘焙制作、出品或售卖中的麻烦，再用产品解决。",
        [
            {"name": "痛点场景", "duration": "0-8s", "goal": "说出具体麻烦"},
            {"name": "情绪放大", "duration": "8-18s", "goal": "让用户觉得被理解"},
            {"name": "解决方案", "duration": "18-45s", "goal": "展示产品如何解决"},
            {"name": "转化收口", "duration": "45-60s", "goal": "引导购买"},
        ],
        ["做烘焙最烦的不是步骤多，是这里总翻车。", "你是不是也遇到过这种情况？"],
        ["别再硬扛了，直接用这个。", "链接放左下角，做烘焙的可以备上。"],
        "【痛点场景】做蛋糕最怕最后装饰毁在细节上。\n【情绪放大】颜色不稳、出品不整齐，前面时间全白花。\n【解决方案】换成这个产品，操作更顺手，效果也更稳定。\n【转化收口】想少翻车，左下角先备起来。",
    ),
    _script_template(
        "需求类",
        "围绕真实使用需求展开，强调用户为什么需要这个产品。",
        [
            {"name": "需求点名", "duration": "0-6s", "goal": "明确人群和需求"},
            {"name": "使用理由", "duration": "6-25s", "goal": "说明为什么需要"},
            {"name": "卖点匹配", "duration": "25-48s", "goal": "匹配产品功能"},
            {"name": "购买建议", "duration": "48-60s", "goal": "给出行动建议"},
        ],
        ["如果你经常做甜品，这个需求一定会遇到。", "不是专业师傅也能用上的烘焙小东西。"],
        ["常做烘焙的可以备一个。", "需要稳定出品的，点左下角。"],
        "【需求点名】经常做甜品的人，最需要的是稳定和省事。\n【使用理由】临时找工具、临时调状态，很容易影响出品。\n【卖点匹配】这个产品就是为了减少准备成本，让流程更顺。\n【购买建议】常做烘焙的可以直接备上。",
    ),
    _script_template(
        "认知类",
        "用知识科普或行业经验建立信任，再自然推荐产品。",
        [
            {"name": "认知反差", "duration": "0-8s", "goal": "打破常见误区"},
            {"name": "知识解释", "duration": "8-30s", "goal": "讲清原理"},
            {"name": "产品验证", "duration": "30-48s", "goal": "用产品证明判断"},
            {"name": "专业建议", "duration": "48-60s", "goal": "给出购买建议"},
        ],
        ["很多新手做不好，不是手法问题，是材料没选对。", "烘焙里这个细节，很多人一直忽略了。"],
        ["按这个思路选，不容易踩坑。", "想省心的直接看左下角这款。"],
        "【认知反差】蛋糕不好看，很多时候不是装饰手法差。\n【知识解释】材料稳定度、颜色表现和操作窗口都会影响成品。\n【产品验证】这个产品上手更稳，适合新手和批量制作。\n【专业建议】按这个标准选，踩坑会少很多。",
    ),
    _script_template(
        "达人分享类",
        "用达人/店主/烘焙爱好者的真实分享口吻种草。",
        [
            {"name": "身份开场", "duration": "0-5s", "goal": "建立真实分享感"},
            {"name": "使用体验", "duration": "5-30s", "goal": "讲亲身使用变化"},
            {"name": "细节展示", "duration": "30-48s", "goal": "补足可信度"},
            {"name": "推荐收尾", "duration": "48-60s", "goal": "自然引导购买"},
        ],
        ["我最近做烘焙一直在用这个。", "这个不是广告感那种，是真的顺手。"],
        ["我自己会回购，链接放左下角。", "想试的可以先拍一份。"],
        "【身份开场】我最近做甜品一直在用这个小东西。\n【使用体验】之前要反复调整，现在出品速度快很多。\n【细节展示】细节更稳，摆盘也更干净。\n【推荐收尾】我自己会回购，需要的点左下角。",
    ),
    _script_template(
        "制作方便",
        "突出省步骤、易上手、少翻车，适合新手和批量制作。",
        [
            {"name": "省事钩子", "duration": "0-6s", "goal": "强调少步骤"},
            {"name": "操作演示", "duration": "6-32s", "goal": "展示简单流程"},
            {"name": "成品结果", "duration": "32-48s", "goal": "展示效果"},
            {"name": "下单理由", "duration": "48-60s", "goal": "引导购买"},
        ],
        ["想把烘焙步骤变简单，这个真的可以。", "新手也能做得像样，关键是少折腾。"],
        ["怕麻烦的直接备这个。", "想做得更顺手，左下角看看。"],
        "【省事钩子】想让烘焙少一步麻烦，这个很适合。\n【操作演示】拿出来直接用，不用反复调、不用一直返工。\n【成品结果】成品看起来干净，效率也高。\n【下单理由】怕麻烦的可以直接备上。",
    ),
    _script_template(
        "成本低",
        "从性价比、损耗、批量使用成本切入，强调低成本高效果。",
        [
            {"name": "成本钩子", "duration": "0-6s", "goal": "点出省钱"},
            {"name": "成本对账", "duration": "6-25s", "goal": "说明省在哪里"},
            {"name": "效果证明", "duration": "25-48s", "goal": "证明便宜不等于凑合"},
            {"name": "促单收口", "duration": "48-60s", "goal": "引导下单"},
        ],
        ["做烘焙想省成本，别只看单价。", "这个东西便宜，但省下来的不止一点点。"],
        ["想控制成本的，左下角可以看。", "出摊和家用都适合备。"],
        "【成本钩子】做烘焙想省成本，真的别只盯着单价。\n【成本对账】少返工、少浪费、出品更稳定，才是真省。\n【效果证明】这个产品价格不高，但效果够用。\n【促单收口】想控制成本的可以备一份。",
    ),
    _script_template(
        "对比类",
        "通过前后、同类或错误用法对比，强化产品差异。",
        [
            {"name": "对比开场", "duration": "0-6s", "goal": "提出对比问题"},
            {"name": "同屏对照", "duration": "6-32s", "goal": "展示差异"},
            {"name": "原因解释", "duration": "32-48s", "goal": "解释差异来源"},
            {"name": "结论转化", "duration": "48-60s", "goal": "给出推荐"},
        ],
        ["同样做烘焙，为什么别人出品更稳？", "这两个效果放一起，差别很明显。"],
        ["想要右边这种效果，直接看左下角。", "对比完你就知道怎么选了。"],
        "【对比开场】同样做甜品，为什么别人出品更整齐？\n【同屏对照】普通做法容易乱，用这个细节会更干净。\n【原因解释】关键是稳定、省步骤，还减少损耗。\n【结论转化】想要这种效果，左下角看这款。",
    ),
    _script_template(
        "情绪类",
        "调动开心、惊喜、焦虑缓解等情绪，让用户被场景打动。",
        [
            {"name": "情绪引入", "duration": "0-8s", "goal": "制造情绪"},
            {"name": "场景共鸣", "duration": "8-25s", "goal": "让用户代入"},
            {"name": "产品释放", "duration": "25-48s", "goal": "用产品缓解或放大情绪"},
            {"name": "情绪收尾", "duration": "48-60s", "goal": "形成购买冲动"},
        ],
        ["做好一个甜品的瞬间，真的很有成就感。", "每次出品翻车，最崩溃的就是最后一步。"],
        ["想要这种稳定感，左下角看看。", "让做甜品变轻松一点，先备上。"],
        "【情绪引入】甜品做好那一刻，真的会让人开心。\n【场景共鸣】但最怕忙了半天，最后细节翻车。\n【产品释放】这个产品能让最后一步更稳，成品也更好看。\n【情绪收尾】想要这种成就感，可以备一个。",
    ),
    _script_template(
        "场景类",
        "把产品放进家庭烘焙、门店出品、节日礼盒等具体场景。",
        [
            {"name": "场景建立", "duration": "0-8s", "goal": "明确使用环境"},
            {"name": "动作展示", "duration": "8-32s", "goal": "展示怎么用"},
            {"name": "成品呈现", "duration": "32-48s", "goal": "展示场景结果"},
            {"name": "场景转化", "duration": "48-60s", "goal": "引导同场景用户购买"},
        ],
        ["家里做甜品，最适合备这种小工具。", "门店出品想快一点，这个场景一定用得上。"],
        ["同样场景需要的，左下角看看。", "家用、出摊都能备。"],
        "【场景建立】家里做甜品或者门店小批量出品，都很适合用这个。\n【动作展示】拿出来直接配合制作流程，用起来顺手。\n【成品呈现】成品更整齐，打包也更体面。\n【场景转化】同样场景需要的，可以点左下角。",
    ),
]


MANUAL_ALIASES = {
    "袋装刀叉": ["袋装刀叉", "刀叉"],
    "盒装刀叉": ["盒装刀叉", "刀叉"],
    "翻糖膏": ["翻糖膏", "防潮翻糖"],
    "彩色翻糖片": ["彩色翻糖片"],
    "色粉盘": ["色粉盘", "色粉", "油溶色粉", "水溶色粉"],
    "拉线膏": ["拉线膏", "手绘膏"],
    "豆沙奶油霜": ["豆沙奶油霜"],
    "果蔬色素": ["果蔬色素"],
    "高浓果蔬色素": ["高浓果蔬色素"],
    "水性色素": ["胶状色素", "水性色素"],
    "水状色素": ["水状色素"],
    "油性色素": ["油性色素"],
    "水溶色粉": ["水溶色粉"],
    "油溶色粉": ["油溶色粉"],
    "竹炭粉": ["竹炭粉"],
    "果蔬粉": ["果蔬粉"],
    "红曲粉": ["红曲粉", "红曲米粉"],
    "布蕾粉": ["布蕾粉"],
    "奶冻粉": ["奶冻粉"],
    "慕斯粉": ["慕斯粉"],
    "香草荚": ["香草荚"],
    "香草精": ["香草精"],
    "红丝绒香精": ["红丝绒香精", "红丝绒"],
    "杏仁粉": ["杏仁粉", "扁桃仁粉"],
    "杏仁片": ["杏仁片", "扁桃仁片"],
    "抹茶粉": ["抹茶粉", "宇治抹茶"],
    "可可粉": ["可可粉"],
    "斑斓粉": ["斑斓粉", "斑斓叶粉"],
    "开心果酱": ["开心果酱"],
    "0卡糖粉": ["0 卡糖粉", "0卡糖粉", "零卡糖"],
}


MANUAL_PRODUCT_NAME_MAP = {
    "胶状色素": "水性色素",
    "竹炭粉": "竹炭粉",
    "杏仁粉": "杏仁粉",
    "0卡糖粉": "0卡糖粉",
}


PRICE_RULES: dict[str, dict[str, Any]] = {
    "袋装刀叉": {
        "knife_price": "bag",
        "products": ["0.7元款刀叉"],
        "product_price_field": "unit_tag_price",
    },
    "盒装刀叉": {
        "knife_price": "box",
        "products": ["2.5元盒装", "2元盒装"],
        "product_price_field": "unit_tag_price",
    },
    "水性色素": {"products": ["胶状色素-小"]},
    "水状色素": {"products": ["水状色素"]},
    "果蔬色素": {"products": ["果蔬色素-小"]},
    "高浓果蔬色素": {"products": ["果蔬色素-大"]},
    "油性色素": {"products": ["油性色素-小"]},
    "水溶色粉": {"products": ["水溶色粉"], "exclude_specs": ["竹炭粉"]},
    "油溶色粉": {"products": ["油溶色粉"]},
    "竹炭粉": {"products": ["竹炭粉"]},
    "果蔬粉": {"products": ["果蔬粉"], "exclude_specs": ["可可", "抹茶", "斑斓", "红曲"]},
    "红曲粉": {"products": ["果蔬粉"], "specs": ["红曲"]},
    "香草荚": {"products": ["香草荚"]},
    "香草精": {"products": ["香草精"]},
    "红丝绒香精": {"products": ["红丝绒"]},
    "杏仁粉": {"products": ["杏仁粉"], "exclude_specs": ["套餐", "糖粉", "可可粉", "抹茶粉", "草莓粉"]},
    "杏仁片": {"products": ["杏仁片"]},
    "抹茶粉": {"products": ["宇治抹茶"]},
    "可可粉": {"products": ["果蔬粉"], "specs": ["可可粉"]},
    "斑斓粉": {"products": ["果蔬粉"], "specs": ["斑斓"]},
    "开心果酱": {"products": ["开心果酱"]},
    "0卡糖粉": {"products": ["零卡糖"]},
    "调味果酱": {"products": ["调味果酱"]},
    "多肉果酱": {"products": ["多肉果酱"]},
    "茶酱": {"products": ["茶酱"]},
    "黑芝麻酱": {"products": ["黑芝麻酱"]},
    "焦糖酱": {"products": ["焦糖酱"]},
    "栗子泥": {"products": ["栗子泥"]},
    "调味糖浆": {"products": ["薄荷糖浆"]},
    "宇治抹茶": {"products": ["宇治抹茶"]},
    "茶粉": {"products": ["茶粉"]},
    "零卡糖": {"products": ["零卡糖"]},
    "海藻糖": {"products": ["海藻糖"]},
    "翻糖膏": {"products": ["白色翻糖膏"]},
    "白色翻糖膏": {"products": ["白色翻糖膏"]},
    "彩色翻糖膏": {"products": ["彩色翻糖膏"]},
    "翻糖压片": {"products": ["彩色翻糖片"]},
    "彩色翻糖片": {"products": ["彩色翻糖片"]},
    "拉线膏": {"products": ["彩色拉线膏"]},
    "彩色拉线膏": {"products": ["彩色拉线膏"]},
    "手绘膏": {"products": ["手绘膏"]},
    "豆沙奶油霜": {"products": ["豆沙奶油霜"]},
    "糖珠": {"products": ["糖珠（小规格）", "糖珠（500g）"]},
    "肉松": {"products": ["肉松"]},
    "巧克力脆皮酱": {"products": ["巧克力脆皮酱"]},
    "巧克力砖": {"products": ["巧克力砖"]},
    "色粉盘": {"products": ["油溶色粉"], "specs": ["套餐"]},
    "色素笔": {"products": ["色素笔"]},
    "胶状色素-小": {"products": ["胶状色素-小"]},
    "胶状色素-大": {"products": ["胶状色素-大"]},
    "油性色素-小": {"products": ["油性色素-小"]},
    "油性色素-大": {"products": ["油性色素-大"]},
    "水状色素-小": {"products": ["水状色素"]},
    "水状色素-大": {"products": ["水状色素"]},
    "布蕾粉": {"products": ["布蕾粉"], "specs": ["200g"]},
    "奶冻粉": {"products": ["奶冻粉"], "specs": ["100g"]},
    "Q弹奶冻粉": {"products": ["奶冻粉"], "specs": ["Q弹奶冻粉200g"]},
    "晶冻粉": {"products": ["奶冻粉"], "specs": ["晶冻粉200g"]},
    "慕斯粉": {"products": ["慕斯粉"], "specs": ["200g"]},
    "慕斯粉（液）": {"products": ["慕斯粉"], "specs": ["200g"]},
    "夹心脆": {"products": ["巧克力夹心脆", "果蔬夹心脆"]},
    "夹心珠": {"products": ["跳跳糖夹心珠"]},
    "夹心芋泥": {"products": ["芋泥"]},
    "黄油薄脆": {"products": ["黄油薄脆"]},
    "巧克力脆珠": {"products": ["巧克力脆珠"]},
    "巧克力脆馅": {"products": ["巧克力脆馅"]},
    "开心果碎": {"products": ["开心果碎"]},
    "1.1浆纸盘": {"products": ["1元纸浆款刀叉"], "product_price_field": "unit_price"},
    "2元盒装": {
        "knife_price": "box",
        "products": ["2元盒装"],
        "product_price_field": "unit_tag_price",
    },
    "2.5元盒装": {
        "knife_price": "box",
        "products": ["2.5元盒装"],
        "product_price_field": "unit_tag_price",
    },
    "5元盒装刀叉": {"products": ["5元云朵盒装刀叉"]},
    "刀叉": {
        "knife_price": "bag",
        "products": ["0.7元款刀叉"],
        "product_price_field": "unit_tag_price",
    },
    "丝带": {"products": ["丝带"]},
    "裱花袋": {"products": ["裱花袋"]},
    "工具模具": {"products": ["模具"]},
}


@dataclass(frozen=True)
class MaterialPaths:
    materials_dir: Path
    product_knowledge_md: Path
    product_manual_md: Path
    scripts_xlsx: Path
    price_system_xlsx: Path | None = None
    knife_price_xlsx: Path | None = None
    product_2026_dir: Path | None = None
    product_card_xlsx: Path | None = None


@dataclass
class ProductInput:
    name: str
    category: str
    price: float
    original_price: float | None
    brand: str
    description: str
    selling_points: list[dict[str, Any]]
    section_text: str


@dataclass
class ScriptInput:
    product_name: str
    sheet_name: str
    source_row: int
    code: str
    raw_type: str
    video_type: str
    title: str
    category: str
    script_content: str
    is_high_conversion: bool
    performance_data: dict[str, Any]
    tags: str


@dataclass
class ImportSummary:
    products_parsed: int = 0
    scripts_parsed: int = 0
    product_created: int = 0
    product_updated: int = 0
    selling_points_written: int = 0
    material_files_written: int = 0
    templates_created: int = 0
    templates_updated: int = 0
    scripts_created: int = 0
    scripts_updated: int = 0
    product_indexed: int = 0
    script_indexed: int = 0
    warnings: list[str] = field(default_factory=list)

    def as_dict(self) -> dict[str, Any]:
        return {
            "products_parsed": self.products_parsed,
            "scripts_parsed": self.scripts_parsed,
            "product_created": self.product_created,
            "product_updated": self.product_updated,
            "selling_points_written": self.selling_points_written,
            "material_files_written": self.material_files_written,
            "templates_created": self.templates_created,
            "templates_updated": self.templates_updated,
            "scripts_created": self.scripts_created,
            "scripts_updated": self.scripts_updated,
            "product_indexed": self.product_indexed,
            "script_indexed": self.script_indexed,
            "warnings": self.warnings,
        }


def get_material_paths(root: Path | str | None = None) -> MaterialPaths:
    root_path = Path(root) if root is not None else Path(__file__).resolve().parent
    materials_dir = root_path / "资料"
    if not materials_dir.exists():
        candidates = [p for p in root_path.iterdir() if p.is_dir() and p.name == "资料"]
        if not candidates:
            raise FileNotFoundError(f"找不到资料目录: {materials_dir}")
        materials_dir = candidates[0]

    md_files = _material_file_candidates(materials_dir.glob("*.md"))
    xlsx_files = _material_file_candidates(materials_dir.glob("*.xlsx"))
    product_knowledge = _single_match(md_files, "知识库")
    product_manual = _optional_match(md_files, "产品手册") or _single_match(md_files, "产品资料")
    scripts_xlsx = _single_match(xlsx_files, "脚本")
    price_system_xlsx = _optional_match(xlsx_files, "价格体系")
    knife_price_xlsx = _optional_match(xlsx_files, "刀叉价格")
    product_2026_dir = materials_dir / "2026产品知识库"
    if not product_2026_dir.exists():
        product_2026_dir = None
    product_card_xlsx = None
    if product_2026_dir is not None:
        product_card_xlsx = _optional_match(
            _material_file_candidates(product_2026_dir.glob("*.xlsx")),
            "产品手卡",
        )
    return MaterialPaths(
        materials_dir,
        product_knowledge,
        product_manual,
        scripts_xlsx,
        price_system_xlsx,
        knife_price_xlsx,
        product_2026_dir,
        product_card_xlsx,
    )


def parse_product_knowledge(path: Path | str) -> list[ProductInput]:
    text = _read_text(Path(path))
    index = _parse_product_index(text)
    sections = _split_product_sections(text)
    products: list[ProductInput] = []

    for name, section_text in sections:
        basic_info = _parse_basic_info(section_text)
        selling_points = _parse_selling_points(section_text)
        category = PRODUCT_CATEGORY_MAP.get(name, "烘焙配件")
        price_text = basic_info.get("价格定位") or basic_info.get("售价") or ""
        price = _parse_price(price_text)
        index_info = index.get(name, {})
        description = _build_product_description(basic_info, index_info, section_text)
        products.append(ProductInput(
            name=name,
            category=category,
            price=price,
            original_price=None,
            brand="法采",
            description=description,
            selling_points=selling_points,
            section_text=section_text.strip(),
        ))

    return products


def parse_product_manual(path: Path | str) -> list[ProductInput]:
    text = _read_text(Path(path))
    products: list[ProductInput] = []
    for section_number, raw_name, section_text in _split_manual_product_sections(text):
        name = _canonical_manual_product_name(raw_name)
        basic_info = _parse_manual_basic_info(section_text)
        selling_points = _parse_manual_selling_points(section_text, basic_info)
        description = _build_manual_description(basic_info, section_text)
        category = PRODUCT_CATEGORY_MAP.get(name, _manual_category(section_number))
        products.append(ProductInput(
            name=name,
            category=category,
            price=_parse_price(section_text),
            original_price=None,
            brand="法采",
            description=description,
            selling_points=selling_points,
            section_text=section_text.strip(),
        ))
    return products


def parse_2026_product_knowledge(path: Path | str | None) -> list[ProductInput]:
    """Parse the 2026 product knowledge pack into app product records."""
    if path is None:
        return []
    knowledge_dir = Path(path)
    if not knowledge_dir.exists():
        return []

    naming_path = knowledge_dir / "05_产品命名主数据与旧称对照.md"
    overview_path = knowledge_dir / "02_核心产品卖点速览.md"
    faq_path = knowledge_dir / "04_产品常见问题精选.md"
    index_path = knowledge_dir / "00_产品知识总索引.md"
    solutions_path = knowledge_dir / "01_五大门店解决方案.md"
    product_card_path = _optional_match(
        _material_file_candidates(knowledge_dir.glob("*.xlsx")),
        "产品手卡",
    )

    standards = _parse_2026_standard_products(naming_path) if naming_path.exists() else {}
    aliases = _parse_2026_aliases(naming_path) if naming_path.exists() else {}
    overview_sections = _parse_2026_markdown_sections(overview_path) if overview_path.exists() else {}
    faq_sections = _parse_2026_markdown_sections(faq_path) if faq_path.exists() else {}
    solution_points = _parse_2026_solution_points(index_path, solutions_path)

    products: list[ProductInput] = []
    for name, metadata in standards.items():
        if name in EXCLUDED_PRODUCT_NAMES:
            continue
        products.append(_build_2026_product_input(
            name=name,
            metadata=metadata,
            aliases=aliases,
            overview_sections=overview_sections,
            faq_sections=faq_sections,
            solution_points=solution_points,
            card_section_text="",
            card_points=[],
            card_price=0.0,
        ))

    if product_card_path is not None and product_card_path.exists():
        products.extend(_parse_2026_product_card_workbook(
            product_card_path,
            standards=standards,
            aliases=aliases,
            overview_sections=overview_sections,
            faq_sections=faq_sections,
            solution_points=solution_points,
        ))

    return merge_product_inputs([
        product for product in products
        if product.name not in EXCLUDED_PRODUCT_NAMES
    ])


def merge_product_inputs(products: list[ProductInput]) -> list[ProductInput]:
    merged: dict[str, ProductInput] = {}
    for product in products:
        if product.name in EXCLUDED_PRODUCT_NAMES:
            continue
        existing = merged.get(product.name)
        if existing is None:
            merged[product.name] = product
            continue

        if not existing.price and product.price:
            existing.price = product.price
        if not existing.original_price and product.original_price:
            existing.original_price = product.original_price
        if product.description and product.description not in existing.description:
            existing.description = (
                existing.description.rstrip()
                + "\n\n产品手册补充：\n"
                + product.description.strip()
            ).strip()
        if product.section_text and product.section_text not in existing.section_text:
            existing.section_text = (
                existing.section_text.rstrip()
                + "\n\n---\n\n## 产品手册补充\n\n"
                + product.section_text.strip()
            )

        seen_points = {
            (_clean_markdown(point["point_type"]), _clean_markdown(point["content"]))
            for point in existing.selling_points
        }
        for point in product.selling_points:
            key = (_clean_markdown(point["point_type"]), _clean_markdown(point["content"]))
            if key in seen_points:
                continue
            existing.selling_points.append({
                "point_type": point["point_type"],
                "content": point["content"],
                "priority": len(existing.selling_points) + 1,
            })
            seen_points.add(key)
    return [_normalize_product_points(product) for product in merged.values()]


def apply_product_prices(
    products: list[ProductInput],
    price_path: Path | None,
    knife_price_path: Path | None = None,
) -> None:
    price_rows = _parse_price_system_rows(price_path) if price_path and price_path.exists() else []
    knife_price_rows_by_type: dict[str, list[dict[str, Any]]] = {}
    if knife_price_path and knife_price_path.exists():
        knife_price_rows_by_type = {
            "bag": _parse_knife_price_rows(knife_price_path, "bag"),
            "box": _parse_knife_price_rows(knife_price_path, "box"),
        }
    if not price_rows and not knife_price_rows_by_type:
        return

    for product in products:
        rule = PRICE_RULES.get(product.name)
        if not rule:
            continue
        override = rule.get("override")
        if override:
            product.price = float(override)
            product.original_price = None
            continue

        candidates: list[float] = []
        rows = (
            knife_price_rows_by_type.get(rule.get("knife_price"), [])
            if rule.get("knife_price") else price_rows
        )
        for row in rows:
            if not _row_matches_price_rule(row, rule):
                continue
            price = _product_price_from_price_row(row, rule)
            if isinstance(price, (int, float)) and price > 0:
                candidates.append(float(price))
        if candidates:
            product.price = _round_price(min(candidates))
            product.original_price = None


def parse_excel_scripts(path: Path | str) -> list[ScriptInput]:
    import openpyxl

    workbook = openpyxl.load_workbook(Path(path), read_only=True, data_only=True)
    scripts: list[ScriptInput] = []
    try:
        for worksheet in workbook.worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                continue
            header = [_cell_text(value) for value in rows[0]]
            columns = _find_script_columns(header)
            if columns["script"] is None:
                continue

            for row_index, row in enumerate(rows[1:], start=2):
                script_text = _get_cell(row, columns["script"])
                if not script_text:
                    continue

                row_product = _get_cell(row, columns["product"])
                product_name = normalize_excel_product_name(worksheet.title, row_product)
                raw_type = _get_cell(row, columns["type"])
                if not raw_type and worksheet.title == "展会机制类":
                    raw_type = "机制"
                video_type = map_script_type(raw_type)
                code = _get_cell(row, columns["id"]) or f"第{row_index}行"
                high_value = _get_cell(row, columns["high"])
                title_type = raw_type or video_type
                title = f"{code} [{title_type}] {product_name} R{row_index}"
                notes = _get_cell(row, columns["notes"])
                case_video = _get_cell(row, columns["case"])
                category = f"法采-{product_name}"
                tags = ",".join(part for part in ["法采", product_name, raw_type] if part)

                scripts.append(ScriptInput(
                    product_name=product_name,
                    sheet_name=worksheet.title,
                    source_row=row_index,
                    code=code,
                    raw_type=raw_type,
                    video_type=video_type,
                    title=title,
                    category=category,
                    script_content=script_text,
                    is_high_conversion=is_high_conversion_marker(high_value),
                    performance_data={
                        "source": SCRIPT_SOURCE,
                        "sheet": worksheet.title,
                        "row": row_index,
                        "product": product_name,
                        "编号": code,
                        "原始类型": raw_type,
                        "高成交标记": high_value,
                        "注意事项": notes,
                        "案例视频": case_video,
                    },
                    tags=tags,
                ))
    finally:
        workbook.close()
    return scripts


def normalize_excel_product_name(sheet_name: str, row_product: str | None = None) -> str:
    raw = (row_product or "").strip() or sheet_name.strip()
    replacements = {
        "刀叉（袋装）": "袋装刀叉",
        "刀叉(袋装)": "袋装刀叉",
        "刀叉（盒装）": "盒装刀叉",
        "刀叉(盒装)": "盒装刀叉",
        "展会机制类": "展会机制",
        "胶状大色素": "水性色素",
    }
    if raw in replacements:
        return replacements[raw]
    normalized = re.sub(r"（.*?）|\(.*?\)", "", raw).strip()
    return replacements.get(normalized, normalized)


def map_script_type(raw_type: Any) -> str:
    key = _cell_text(raw_type)
    return SCRIPT_TYPE_MAP.get(key, "机制类")


def is_high_conversion_marker(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value > 0
    text = str(value).strip()
    if not text:
        return False
    if text in {"否", "无", "0", "false", "False", "已拍"}:
        return False
    return "高成交" in text or text in {"是", "yes", "YES", "Y", "y", "1"}


def build_script_dedupe_key(script: ScriptInput) -> tuple[str, str]:
    return script.category, script.title


def build_script_source_key(script: ScriptInput) -> tuple[str, str, int]:
    return SCRIPT_SOURCE, script.sheet_name, script.source_row


def needs_index_reset(
    expected_products: int,
    expected_scripts: int,
    product_indexed: int,
    script_indexed: int,
) -> bool:
    product_failed = expected_products > 0 and product_indexed == 0
    script_failed = expected_scripts > 0 and script_indexed == 0
    return product_failed or script_failed


def import_materials(root: Path | str | None = None, dry_run: bool = False) -> ImportSummary:
    root_path = Path(root) if root is not None else Path(__file__).resolve().parent
    paths = get_material_paths(root_path)
    products = merge_product_inputs(
        parse_product_knowledge(paths.product_knowledge_md)
        + parse_product_manual(paths.product_manual_md)
        + parse_2026_product_knowledge(paths.product_2026_dir)
    )
    apply_product_prices(products, paths.price_system_xlsx, paths.knife_price_xlsx)
    scripts = parse_excel_scripts(paths.scripts_xlsx)
    summary = ImportSummary(products_parsed=len(products), scripts_parsed=len(scripts))

    if dry_run:
        return summary

    from database import SessionLocal, init_db
    from models import Product, ScriptTemplate, SellingPoint, ViralScript

    init_db()
    db = SessionLocal()
    try:
        manual_text = _read_text(paths.product_manual_md)
        material_dir = root_path / "data" / "product_files"
        material_dir.mkdir(parents=True, exist_ok=True)

        for product_input in products:
            material_path = _write_product_material_file(material_dir, product_input, manual_text)
            product = db.query(Product).filter(Product.name == product_input.name).first()
            if product is None:
                product = Product(name=product_input.name)
                db.add(product)
                summary.product_created += 1
            else:
                summary.product_updated += 1

            product.category = product_input.category
            product.price = product_input.price
            product.original_price = product_input.original_price
            product.commission_rate = 0.0
            product.brand = product_input.brand
            product.description = product_input.description
            product.info_file = str(material_path)
            product.status = "active"
            db.flush()

            db.query(SellingPoint).filter(SellingPoint.product_id == product.id).delete()
            for point in product_input.selling_points:
                db.add(SellingPoint(product_id=product.id, **point))
                summary.selling_points_written += 1
            summary.material_files_written += 1

        for template_data in REQUESTED_SCRIPT_TEMPLATES:
            template = db.query(ScriptTemplate).filter(
                ScriptTemplate.name == template_data["name"]
            ).first()
            if template is None:
                db.add(ScriptTemplate(**template_data))
                summary.templates_created += 1
            else:
                for key, value in template_data.items():
                    setattr(template, key, value)
                summary.templates_updated += 1

        db.query(ScriptTemplate).filter(
            ~ScriptTemplate.video_type.in_(REQUESTED_VIDEO_TYPES)
        ).delete(synchronize_session=False)

        existing_material_scripts = _index_existing_material_scripts(db, ViralScript)
        for script_input in scripts:
            category, title = build_script_dedupe_key(script_input)
            source_key = build_script_source_key(script_input)
            script = existing_material_scripts.get(source_key)
            if script is None:
                script = db.query(ViralScript).filter(
                    ViralScript.category == category,
                    ViralScript.title == title,
                ).first()
            if script is None:
                script = ViralScript(category=category, title=title)
                db.add(script)
                summary.scripts_created += 1
            else:
                summary.scripts_updated += 1

            script.category = category
            script.title = title
            script.video_type = script_input.video_type
            script.script_content = script_input.script_content
            script.performance_data = script_input.performance_data
            script.tags = script_input.tags
            script.is_high_conversion = 1 if script_input.is_high_conversion else 0
            existing_material_scripts[source_key] = script

        db.commit()

        _rebuild_vector_indexes(db, root_path, summary, len(products), len(scripts))

    except Exception:
        db.rollback()
        raise
    finally:
        db.close()

    return summary


def main() -> int:
    parser = argparse.ArgumentParser(description="导入资料目录中的产品和脚本数据")
    parser.add_argument("--root", type=Path, default=Path(__file__).resolve().parent)
    parser.add_argument("--dry-run", action="store_true", help="只解析并输出统计，不写数据库")
    args = parser.parse_args()

    summary = import_materials(args.root, dry_run=args.dry_run)
    print(json.dumps(summary.as_dict(), ensure_ascii=False, indent=2))
    return 0


def _single_match(paths: list[Path], needle: str) -> Path:
    matches = [path for path in paths if needle in path.name]
    if len(matches) != 1:
        raise FileNotFoundError(f"期望找到 1 个包含 {needle!r} 的文件，实际找到 {len(matches)} 个")
    return matches[0]


def _optional_match(paths: list[Path], needle: str) -> Path | None:
    matches = [path for path in paths if needle in path.name]
    if not matches:
        return None
    if len(matches) > 1:
        matches.sort(key=lambda path: path.stat().st_mtime, reverse=True)
    return matches[0]


def _material_file_candidates(paths: Any) -> list[Path]:
    return [
        path for path in paths
        if not path.name.startswith("~$")
    ]


def _rebuild_vector_indexes(
    db,
    root_path: Path,
    summary: ImportSummary,
    expected_products: int,
    expected_scripts: int,
) -> None:
    reset_done = _reset_chroma_persist_dir(root_path)
    if not reset_done:
        summary.warnings.append("Chroma 目录被占用，跳过目录重置并改为原地更新索引")
    summary.product_indexed, summary.script_indexed = _run_vector_indexers(db, summary)
    if not needs_index_reset(
        expected_products,
        expected_scripts,
        summary.product_indexed,
        summary.script_indexed,
    ):
        db.commit()
        return

    summary.warnings.append("索引重建返回 0，请检查 Chroma/embedding 运行环境")
    db.commit()


def _run_vector_indexers(db, summary: ImportSummary) -> tuple[int, int]:
    product_indexed = 0
    script_indexed = 0
    try:
        from vector_store.product_store import ProductVectorStore
        product_indexed = ProductVectorStore().index_all_products(db)
    except Exception as exc:  # pragma: no cover - depends on local model/runtime
        summary.warnings.append(f"产品索引重建失败: {exc}")

    try:
        from vector_store.script_store import ScriptVectorStore
        script_indexed = ScriptVectorStore().index_all_scripts(db)
    except Exception as exc:  # pragma: no cover - depends on local model/runtime
        summary.warnings.append(f"脚本索引重建失败: {exc}")
    return product_indexed, script_indexed


def _index_existing_material_scripts(db, viral_model) -> dict[tuple[str, str, int], Any]:
    scripts_by_source: dict[tuple[str, str, int], Any] = {}
    duplicates: list[Any] = []
    for script in db.query(viral_model).order_by(viral_model.id).all():
        source_key = _performance_source_key(script.performance_data)
        if source_key is None:
            continue
        if source_key in scripts_by_source:
            duplicates.append(script)
            continue
        scripts_by_source[source_key] = script

    for script in duplicates:
        db.delete(script)
    if duplicates:
        db.flush()
    return scripts_by_source


def _performance_source_key(performance_data: Any) -> tuple[str, str, int] | None:
    if isinstance(performance_data, str):
        try:
            performance_data = json.loads(performance_data)
        except json.JSONDecodeError:
            return None
    if not isinstance(performance_data, dict):
        return None

    source = performance_data.get("source")
    sheet = performance_data.get("sheet")
    row = performance_data.get("row")
    if source != SCRIPT_SOURCE or not sheet or row is None:
        return None
    try:
        row_number = int(row)
    except (TypeError, ValueError):
        return None
    return source, str(sheet), row_number


def _reset_chroma_persist_dir(root_path: Path) -> bool:
    from config import CHROMA_PERSIST_DIR
    import vector_store

    persist_dir = Path(CHROMA_PERSIST_DIR)
    if not persist_dir.is_absolute():
        persist_dir = root_path / persist_dir
    persist_dir = persist_dir.resolve()
    root_resolved = root_path.resolve()
    if persist_dir != root_resolved and root_resolved not in persist_dir.parents:
        raise RuntimeError(f"拒绝删除工作区外的 Chroma 目录: {persist_dir}")

    vector_store._store = None
    if persist_dir.exists():
        try:
            shutil.rmtree(persist_dir)
        except PermissionError:
            return False
    return True


def _read_text(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="replace")


def _split_product_sections(text: str) -> list[tuple[str, str]]:
    pattern = re.compile(r"^## 产品名称[:：](.+?)\s*$", re.MULTILINE)
    matches = list(pattern.finditer(text))
    sections: list[tuple[str, str]] = []
    for index, match in enumerate(matches):
        start = match.start()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        name = _clean_markdown(match.group(1))
        sections.append((name, text[start:end].strip()))
    return sections


def _split_manual_product_sections(text: str) -> list[tuple[int, str, str]]:
    pattern = re.compile(r"^第\s*(\d+)\s*节[:：]?\s*(.+?)\s*$", re.MULTILINE)
    matches = list(pattern.finditer(text))
    sections: list[tuple[int, str, str]] = []
    for index, match in enumerate(matches):
        start = match.start()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        sections.append((int(match.group(1)), _clean_markdown(match.group(2)), text[start:end].strip()))
    return sections


def _canonical_manual_product_name(raw_name: str) -> str:
    name = re.sub(r"[（(].*?[）)]", "", _clean_markdown(raw_name)).replace(" ", "")
    return MANUAL_PRODUCT_NAME_MAP.get(name, name)


def _manual_category(section_number: int) -> str:
    if section_number <= 10:
        return "烘焙调色"
    if section_number <= 20:
        return "烘焙调味"
    return "烘焙装饰"


def _parse_manual_basic_info(section_text: str) -> dict[str, str]:
    labels = [
        "顾客群体",
        "产品名称",
        "产品规格",
        "产品定义",
        "有效期",
        "保质期",
        "储存说明",
        "储存条件",
        "产品状态",
        "原料",
        "适用范围",
    ]
    info: dict[str, str] = {}
    for raw_line in section_text.splitlines():
        line = _clean_manual_line(raw_line)
        if not line:
            continue
        for label in labels:
            index = line.find(label)
            if index < 0:
                continue
            value = line[index + len(label):].strip(" ：:-")
            if not value or value == label:
                continue
            value = re.sub(r"^(基本|信息|规格及)\s*", "", value).strip(" ：:-")
            if value and label not in info:
                info[label] = value[:220]
            break
    return info


def _parse_manual_selling_points(
    section_text: str, basic_info: dict[str, str]
) -> list[dict[str, Any]]:
    points: list[dict[str, Any]] = []
    for key, label in [("顾客群体", "适用人群"), ("产品规格", "规格"), ("适用范围", "场景")]:
        value = basic_info.get(key)
        if value:
            _append_point(points, label, value)

    keyword_types = [
        ("优势", "优势"),
        ("优点", "优势"),
        ("特质", "特质"),
        ("特点", "特质"),
        ("产品状态", "状态"),
        ("使用", "使用"),
        ("对比", "对比"),
        ("利益", "利益点"),
        ("赠品", "赠品"),
    ]
    for raw_line in section_text.splitlines():
        line = _clean_manual_line(raw_line)
        if not _is_manual_point_candidate(line):
            continue
        for keyword, point_type in keyword_types:
            if keyword not in line:
                continue
            content = _strip_manual_point_prefix(line, keyword)
            _append_point(points, point_type, content)
            break
        if len(points) >= 6:
            break

    if len(points) < 3:
        for raw_line in section_text.splitlines():
            line = _clean_manual_line(raw_line)
            if not _is_manual_point_candidate(line):
                continue
            if not re.match(r"^[①②③④⑤⑥⑦⑧⑨⑩\d]+[、.]", line):
                continue
            _append_point(points, "核心卖点", line)
            if len(points) >= 4:
                break

    return points or [{
        "point_type": "资料",
        "content": _clean_markdown(section_text)[:160],
        "priority": 1,
    }]


def _build_manual_description(basic_info: dict[str, str], section_text: str) -> str:
    parts = [f"{key}：{value}" for key, value in basic_info.items() if value]
    if parts:
        return "\n".join(parts)
    return _clean_markdown(re.sub(r"^## 第 \d+ 页$", "", section_text, flags=re.MULTILINE))[:1000]


def _clean_manual_line(raw_line: str) -> str:
    line = _clean_markdown(raw_line)
    line = re.sub(r"^## 第 \d+ 页$", "", line)
    line = re.sub(r"^第\s*\d+\s*节[:：]?", "", line).strip()
    line = re.sub(r"^(基本|信息|规格及)\s+", "", line)
    return line.strip()


def _is_manual_point_candidate(line: str) -> bool:
    if not line or len(line) < 8 or len(line) > 260:
        return False
    if line.startswith("#") or line.startswith("第") or line.startswith("【"):
        return False
    if "？" in line or "?" in line or "客户常见问题" in line:
        return False
    return True


def _strip_manual_point_prefix(line: str, keyword: str) -> str:
    line = re.sub(r"^[①②③④⑤⑥⑦⑧⑨⑩\d]+[、.]\s*", "", line)
    index = line.find(keyword)
    if 0 <= index <= 8:
        line = line[index + len(keyword):]
    return line.strip(" ：:-")


def _append_point(points: list[dict[str, Any]], point_type: str, content: str) -> None:
    content = _clean_markdown(content)
    if not content:
        return
    key = (_clean_markdown(point_type), content)
    existing = {
        (_clean_markdown(point["point_type"]), _clean_markdown(point["content"]))
        for point in points
    }
    if key in existing:
        return
    points.append({
        "point_type": point_type,
        "content": content[:240],
        "priority": len(points) + 1,
    })


def _normalize_product_points(product: ProductInput) -> ProductInput:
    for index, point in enumerate(product.selling_points, start=1):
        point["priority"] = index
    return product


def _parse_price_system_rows(path: Path) -> list[dict[str, Any]]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = next(
            (sheet for sheet in workbook.worksheets if "3月10" in sheet.title),
            workbook.worksheets[2],
        )
        header_rows = list(worksheet.iter_rows(min_row=2, max_row=3, values_only=True))
        activity_columns = _price_activity_columns(header_rows)
        rows: list[dict[str, Any]] = []
        category = ""
        product = ""
        for index, row in enumerate(worksheet.iter_rows(min_row=4, values_only=True), start=4):
            if row[0]:
                category = _cell_text(row[0])
            if row[1]:
                product = _cell_text(row[1])
            if not product:
                continue
            activity_prices = _parse_price_activity_values(row, activity_columns)
            daily_activity = _find_price_activity(activity_prices, "淘宝-调整")
            rows.append({
                "row": index,
                "category": category,
                "product": product,
                "spec": _cell_text(row[2]) if len(row) > 2 else "",
                "tag_price": (
                    daily_activity.get("tag_price")
                    if daily_activity else row[9] if len(row) > 9 else None
                ),
                "daily_price": (
                    daily_activity.get("final_price") or daily_activity.get("activity_price")
                    if daily_activity else row[12] if len(row) > 12 else None
                ),
                "unit_price": daily_activity.get("unit_price") if daily_activity else None,
                "count": (
                    daily_activity.get("count")
                    if daily_activity else row[16] if len(row) > 16 else None
                ),
                "activity_prices": activity_prices,
            })
        return rows
    finally:
        workbook.close()


def _parse_knife_price_rows(path: Path, price_type: str) -> list[dict[str, Any]]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = _knife_price_worksheet(workbook.worksheets, price_type)
        if worksheet is None:
            return []

        header_rows = list(worksheet.iter_rows(min_row=1, max_row=2, values_only=True))
        activity_columns = _price_activity_columns(header_rows)
        rows: list[dict[str, Any]] = []
        category = ""
        product = ""
        daily_prefix = "现在价格" if price_type == "bag" else "线上-日常涨价"

        for index, row in enumerate(worksheet.iter_rows(min_row=3, values_only=True), start=3):
            first = _cell_text(row[0]) if len(row) > 0 else ""
            second = _cell_text(row[1]) if len(row) > 1 else ""
            spec = _cell_text(row[2]) if len(row) > 2 else ""
            if first in {"系列", "分销价格"} or second == "款式":
                break
            if first:
                category = first
            if second:
                product = second
            if not product or not spec:
                continue

            activity_prices = _parse_price_activity_values(row, activity_columns)
            daily_activity = _find_price_activity(activity_prices, daily_prefix)
            if daily_activity is None:
                daily_activity = _find_price_activity(activity_prices, "线上-现有")
            count = row[3] if len(row) > 3 else None
            tag_price = daily_activity.get("tag_price") if daily_activity else None
            daily_price = (
                daily_activity.get("final_price") or daily_activity.get("activity_price")
                if daily_activity else None
            )

            rows.append({
                "row": index,
                "category": category,
                "product": product,
                "spec": spec,
                "tag_price": tag_price,
                "daily_price": daily_price,
                "count": count,
                "activity_prices": activity_prices,
                "unit_tag_price": _divide_numeric(tag_price, count),
                "unit_daily_price": (
                    daily_activity.get("unit_price") if daily_activity else None
                ),
            })
        return rows
    finally:
        workbook.close()


def _knife_price_worksheet(worksheets: list[Any], price_type: str) -> Any | None:
    if price_type == "bag":
        return next((sheet for sheet in worksheets if "袋装价格表" in sheet.title), None)
    if price_type == "box":
        return next(
            (sheet for sheet in worksheets if "盒装" in sheet.title and "最终涨价后" in sheet.title),
            next((sheet for sheet in worksheets if "盒装价格表" in sheet.title), None),
        )
    return None


def _price_activity_columns(header_rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    if len(header_rows) < 2:
        return []
    group_row, label_row = header_rows
    columns: list[dict[str, Any]] = []
    current_group = ""
    for index, (group_value, label_value) in enumerate(zip(group_row, label_row)):
        group = _cell_text(group_value)
        label = _cell_text(label_value)
        if group:
            current_group = group
        if not current_group or not label:
            continue
        if current_group in {"品类", "产品", "规格"}:
            continue
        if label in {"调整前 到手价-吊牌价", "调整后 和调整前差异", "是否报名活动"}:
            continue
        columns.append({
            "index": index,
            "mechanism": current_group,
            "label": label,
            "field": _price_activity_field(label),
        })
    return columns


def _price_activity_field(label: str) -> str:
    if label == "吊牌价":
        return "tag_price"
    if label == "商品折扣":
        return "discount"
    if "商品优惠券" in label:
        return "coupon"
    if label in {"85折价", "活动价", "折后价"}:
        return "activity_price"
    if label in {"到手价", "券后到手价", "现售价"}:
        return "final_price"
    if label in {"券后到手价/瓶", "单套价格"}:
        return "unit_price"
    if label == "单品活动":
        return "single_activity"
    if label in {"瓶数", "套数"}:
        return "count"
    return label


def _parse_price_activity_values(
    row: tuple[Any, ...],
    activity_columns: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    by_mechanism: dict[str, dict[str, Any]] = {}
    for column in activity_columns:
        index = column["index"]
        value = row[index] if len(row) > index else None
        if value is None or _cell_text(value) == "":
            continue
        activity = by_mechanism.setdefault(
            column["mechanism"],
            {"mechanism": column["mechanism"]},
        )
        activity[column["field"]] = value
    return [
        activity for activity in by_mechanism.values()
        if any(key in activity for key in ("tag_price", "activity_price", "final_price"))
    ]


def _find_price_activity(
    activity_prices: list[dict[str, Any]],
    mechanism_prefix: str,
) -> dict[str, Any] | None:
    return next(
        (
            activity for activity in activity_prices
            if activity.get("mechanism", "").startswith(mechanism_prefix)
        ),
        None,
    )


def _row_matches_price_rule(row: dict[str, Any], rule: dict[str, Any]) -> bool:
    product = row.get("product", "")
    spec = row.get("spec", "")
    full_text = f"{product} {spec}"
    products = rule.get("products") or []
    if products and not any(alias == product or alias in product for alias in products):
        return False
    specs = rule.get("specs") or []
    if specs and not any(keyword in full_text for keyword in specs):
        return False
    exclude_specs = rule.get("exclude_specs") or []
    if exclude_specs and any(keyword in full_text for keyword in exclude_specs):
        return False
    return True


def _product_price_from_price_row(row: dict[str, Any], rule: dict[str, Any]) -> Any:
    field = rule.get("product_price_field")
    if field:
        return row.get(str(field))
    return row.get("tag_price")


def _divide_numeric(numerator: Any, denominator: Any) -> float | None:
    if not isinstance(numerator, (int, float)):
        return None
    if not isinstance(denominator, (int, float)) or denominator <= 0:
        return None
    return float(numerator) / float(denominator)


def _round_price(value: float) -> float:
    rounded = round(value, 2)
    if abs(rounded - round(rounded)) < 0.005:
        return float(round(rounded))
    return rounded


def _parse_product_index(text: str) -> dict[str, dict[str, str]]:
    start = text.find("## 产品索引表")
    if start < 0:
        return {}
    end = text.find("---", start)
    table_text = text[start:end if end > start else len(text)]
    index: dict[str, dict[str, str]] = {}
    for line in table_text.splitlines():
        if not line.startswith("|") or "---" in line or "产品名称" in line:
            continue
        cells = [_clean_markdown(cell) for cell in line.strip("|").split("|")]
        if len(cells) >= 4 and cells[1]:
            index[cells[1]] = {"产品定位": cells[2], "适用人群": cells[3]}
    return index


def _parse_basic_info(section_text: str) -> dict[str, str]:
    block = _extract_heading_block(section_text, "### 基本信息")
    info: dict[str, str] = {}
    for line in block.splitlines():
        line = line.strip()
        match = re.match(r"^-\s*\*\*(.+?)\*\*[:：]\s*(.+)$", line)
        if match:
            info[_clean_markdown(match.group(1))] = _clean_markdown(match.group(2))
    return info


def _parse_selling_points(section_text: str) -> list[dict[str, Any]]:
    block = _extract_heading_block(section_text, "### 核心卖点")
    points: list[dict[str, Any]] = []
    for line in block.splitlines():
        line = line.strip()
        match = re.match(r"^\d+[.、]\s*(?:\*\*(.+?)\*\*[:：]?)?\s*(.+)$", line)
        if not match:
            continue
        label = _clean_markdown(match.group(1) or "核心卖点")
        content = _clean_markdown(match.group(2))
        if not content:
            continue
        points.append({
            "point_type": label,
            "content": content,
            "priority": len(points) + 1,
        })
    return points


def _parse_2026_standard_products(path: Path) -> dict[str, dict[str, str]]:
    text = _read_text(path)
    products: dict[str, dict[str, str]] = {}
    in_table = False
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if line.startswith("## 标准产品主数据"):
            in_table = True
            continue
        if in_table and line.startswith("## 已确认旧称对照"):
            break
        if not in_table or not line.startswith("|") or "---" in line or "标准产品名称" in line:
            continue
        cells = [_clean_markdown(cell) for cell in line.strip("|").split("|")]
        if len(cells) < 5 or not cells[2]:
            continue
        series, product_type, name, grade, lifecycle = cells[:5]
        products[name] = {
            "series": series,
            "product_type": product_type,
            "grade": grade,
            "lifecycle": lifecycle,
            "category": STANDARD_SERIES_CATEGORY_MAP.get(series, _infer_2026_category(name, series)),
        }
    return products


def _parse_2026_aliases(path: Path) -> dict[str, str]:
    text = _read_text(path)
    aliases: dict[str, str] = {}
    mode = ""
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if line.startswith("## 已确认旧称对照"):
            mode = "aliases"
            continue
        if line.startswith("## SKU 与品类统称"):
            mode = "sku"
            continue
        if line.startswith("## ") and mode:
            mode = ""
        if not mode or not line.startswith("|") or "---" in line:
            continue
        cells = [_clean_markdown(cell) for cell in line.strip("|").split("|")]
        if mode == "aliases":
            if len(cells) < 2 or cells[0] == "旧资料名称":
                continue
            standard = cells[1]
            for alias in _split_2026_alias_names(cells[0]):
                aliases[alias] = standard
        elif mode == "sku":
            if len(cells) < 3 or cells[0] == "名称":
                continue
            parent = cells[2].split("/")[0].strip()
            if parent and parent != "色素":
                aliases[cells[0]] = parent
    return aliases


def _split_2026_alias_names(value: str) -> list[str]:
    cleaned = _clean_markdown(value)
    names: list[str] = []
    for part in re.split(r"[、,，/]+", cleaned):
        part = part.strip()
        if not part:
            continue
        names.append(part)
        without_parens = re.sub(r"[（(].*?[）)]", "", part).strip()
        if without_parens and without_parens != part:
            names.append(without_parens)
    return names


def _parse_2026_markdown_sections(path: Path) -> dict[str, str]:
    text = _read_text(path)
    matches = list(re.finditer(r"^##\s+(.+?)\s*$", text, flags=re.MULTILINE))
    sections: dict[str, str] = {}
    for index, match in enumerate(matches):
        heading = _clean_markdown(match.group(1))
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        sections[heading] = text[start:end].strip()
    return sections


def _parse_2026_solution_points(*paths: Path) -> list[dict[str, str]]:
    points: list[dict[str, str]] = []
    for path in paths:
        if not path.exists():
            continue
        current_heading = ""
        for raw_line in _read_text(path).splitlines():
            line = raw_line.strip()
            if line.startswith("### ") or re.match(r"^##\s+\d+[.、]", line):
                current_heading = _clean_markdown(line.lstrip("# "))
                continue
            if not line.startswith("- "):
                continue
            content = _clean_markdown(line[2:])
            if not content:
                continue
            points.append({
                "source": path.name,
                "content": f"{current_heading}：{content}" if current_heading else content,
            })
    return points


def _build_2026_product_input(
    name: str,
    metadata: dict[str, str] | None,
    aliases: dict[str, str],
    overview_sections: dict[str, str],
    faq_sections: dict[str, str],
    solution_points: list[dict[str, str]],
    card_section_text: str,
    card_points: list[dict[str, Any]],
    card_price: float,
) -> ProductInput:
    metadata = metadata or {}
    category = metadata.get("category") or _infer_2026_category(name, metadata.get("series", ""))
    known_aliases = _aliases_for_2026_product(name, aliases)
    points: list[dict[str, Any]] = []
    section_blocks = [
        f"## 产品名称：{name}",
        "",
        "### 2026 标准命名",
        "",
        f"- 系列：{metadata.get('series', '未标注')}",
        f"- 品类：{metadata.get('product_type', category)}",
        f"- 产品等级：{metadata.get('grade', '未标注')}",
        f"- 新/老品：{metadata.get('lifecycle', '未标注')}",
    ]
    if known_aliases:
        section_blocks.append(f"- 旧称/别名：{'、'.join(known_aliases)}")

    overview = _find_2026_section(overview_sections, name, aliases)
    if overview:
        overview_points = _parse_2026_bullet_points(overview)
        points.extend(overview_points)
        section_blocks.extend(["", "### 核心产品卖点速览", "", overview])

    faq = _find_2026_section(faq_sections, name, aliases)
    if faq:
        faq_points = _parse_2026_bullet_points(faq)
        points.extend(faq_points)
        section_blocks.extend(["", "### 常见问题精选", "", faq])

    for solution in _matching_2026_solution_points(name, aliases, solution_points):
        _append_point(points, "门店方案", solution["content"])

    for point in card_points:
        _append_point(points, point.get("point_type", "产品手卡"), point.get("content", ""))
    if card_section_text:
        section_blocks.extend(["", card_section_text])

    if not points:
        _append_point(points, "标准命名", _build_2026_metadata_sentence(name, metadata, known_aliases))

    description = _build_2026_description(name, metadata, known_aliases, points)
    return ProductInput(
        name=name,
        category=category,
        price=card_price,
        original_price=None,
        brand="法采",
        description=description,
        selling_points=points,
        section_text="\n".join(section_blocks).strip(),
    )


def _parse_2026_product_card_workbook(
    path: Path,
    standards: dict[str, dict[str, str]],
    aliases: dict[str, str],
    overview_sections: dict[str, str],
    faq_sections: dict[str, str],
    solution_points: list[dict[str, str]],
) -> list[ProductInput]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    products: list[ProductInput] = []
    try:
        for worksheet in workbook.worksheets:
            raw_name = _clean_markdown(worksheet.title)
            if not raw_name or raw_name in PRODUCT_CARD_SKIP_SHEETS:
                continue
            rows = _extract_2026_card_rows(worksheet)
            if not rows:
                continue
            name = _canonical_2026_card_product_name(raw_name, standards, aliases)
            products.append(_build_2026_product_input(
                name=name,
                metadata=standards.get(name),
                aliases=aliases,
                overview_sections=overview_sections,
                faq_sections=faq_sections,
                solution_points=solution_points,
                card_section_text=_format_2026_card_section(raw_name, rows),
                card_points=_extract_2026_card_points(rows),
                card_price=_extract_2026_card_price(rows),
            ))
    finally:
        workbook.close()
    return products


def _extract_2026_card_rows(worksheet) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in worksheet.iter_rows(values_only=True):
        values = [_cell_text(value) for value in row]
        values = [value for value in values if value]
        if not values:
            continue
        rows.append(values[:12])
        if len(rows) >= 120:
            break
    return rows


def _format_2026_card_section(sheet_name: str, rows: list[list[str]]) -> str:
    lines = [
        "### 2026 产品手卡",
        "",
        f"- 工作表：{sheet_name}",
        "",
    ]
    for values in rows[:80]:
        content = " | ".join(value[:220] for value in values)
        if content:
            lines.append(f"- {content}")
    return "\n".join(lines).strip()


def _extract_2026_card_points(rows: list[list[str]]) -> list[dict[str, Any]]:
    points: list[dict[str, Any]] = []
    label_keys = {_label_key(label) for label in PRODUCT_CARD_POINT_LABELS}
    for values in rows:
        for index, value in enumerate(values[:-1]):
            label = _clean_markdown(value)
            if _label_key(label) not in label_keys:
                continue
            content = _join_2026_card_content(values[index + 1:])
            _append_point(points, label, content)
        if len(points) >= 10:
            break
    if points:
        return points

    for values in rows[:8]:
        content = _join_2026_card_content(values)
        if len(content) >= 8:
            _append_point(points, "产品手卡", content)
        if len(points) >= 4:
            break
    return points


def _extract_2026_card_price(rows: list[list[str]]) -> float:
    for values in rows:
        for index, value in enumerate(values[:-1]):
            label = _label_key(value)
            if not _is_2026_price_label(label):
                continue
            content = _join_2026_card_content(values[index + 1:])
            price = _first_2026_price_number(content)
            if price:
                return _round_price(price)
    return 0.0


def _is_2026_price_label(label: str) -> bool:
    if "折" in label:
        return False
    return label in {"售价", "价格", "吊牌价", "日常售价", "标价"} or "到手价" in label


def _first_2026_price_number(content: str) -> float:
    for match in re.finditer(r"\d+(?:\.\d+)?", content):
        start, end = match.span()
        around = content[max(0, start - 2): min(len(content), end + 3)]
        if any(unit in around for unit in ("折", "g", "G", "kg", "KG", "个", "件", "包", "套", "人份", "个月", "年")):
            if "元" not in around:
                continue
        price = float(match.group(0))
        if 0 < price < 10000:
            return price
    return 0.0


def _join_2026_card_content(values: list[str]) -> str:
    content = "；".join(_clean_markdown(value) for value in values if _clean_markdown(value))
    content = re.sub(r"\s+", " ", content).strip(" ；")
    return content[:500]


def _canonical_2026_card_product_name(
    raw_name: str,
    standards: dict[str, dict[str, str]],
    aliases: dict[str, str],
) -> str:
    if raw_name in PRODUCT_CARD_SHEET_NAME_OVERRIDES:
        return PRODUCT_CARD_SHEET_NAME_OVERRIDES[raw_name]
    if raw_name in standards:
        return raw_name
    if raw_name in aliases and aliases[raw_name] in standards:
        return aliases[raw_name]
    base_name = re.sub(r"[（(].*?[）)]", "", raw_name).strip()
    if base_name in PRODUCT_CARD_SHEET_NAME_OVERRIDES:
        return PRODUCT_CARD_SHEET_NAME_OVERRIDES[base_name]
    if base_name in standards:
        return base_name
    if base_name in aliases and aliases[base_name] in standards:
        return aliases[base_name]
    return raw_name


def _parse_2026_bullet_points(block: str) -> list[dict[str, Any]]:
    points: list[dict[str, Any]] = []
    for raw_line in block.splitlines():
        line = raw_line.strip()
        if not line.startswith("- "):
            continue
        body = _clean_markdown(line[2:])
        if not body:
            continue
        match = re.match(r"^(.{2,24}?)[：:]\s*(.+)$", body)
        if match:
            _append_point(points, match.group(1), match.group(2))
        else:
            _append_point(points, "产品知识", body)
    return points


def _find_2026_section(sections: dict[str, str], name: str, aliases: dict[str, str]) -> str:
    candidate_keys = _candidate_2026_keys(name, aliases)
    for heading, block in sections.items():
        if _name_key_2026(heading) in candidate_keys:
            return block
    return ""


def _matching_2026_solution_points(
    name: str,
    aliases: dict[str, str],
    solution_points: list[dict[str, str]],
) -> list[dict[str, str]]:
    candidate_keys = _candidate_2026_keys(name, aliases)
    matched: list[dict[str, str]] = []
    for point in solution_points:
        content_key = _name_key_2026(point["content"])
        if any(key and key in content_key for key in candidate_keys):
            matched.append(point)
    return matched[:4]


def _aliases_for_2026_product(name: str, aliases: dict[str, str]) -> list[str]:
    found = [alias for alias, standard in aliases.items() if standard == name and alias != name]
    return sorted(set(found))


def _candidate_2026_keys(name: str, aliases: dict[str, str]) -> set[str]:
    candidates = {name}
    if name in aliases:
        candidates.add(aliases[name])
    candidates.update(_aliases_for_2026_product(name, aliases))
    candidates.update(PRODUCT_CARD_SHEET_NAME_OVERRIDES.get(alias, alias) for alias in list(candidates))
    return {_name_key_2026(candidate) for candidate in candidates if candidate}


def _name_key_2026(value: str) -> str:
    return re.sub(r"[\s（）()·\-_/、，,：:]+", "", _clean_markdown(value)).lower()


def _label_key(value: str) -> str:
    return re.sub(r"[\s（）()：:]+", "", _clean_markdown(value)).lower()


def _build_2026_metadata_sentence(
    name: str,
    metadata: dict[str, str],
    aliases: list[str],
) -> str:
    parts = [
        f"{name}属于{metadata.get('series', '产品线')}",
        f"品类为{metadata.get('product_type', '未标注')}",
    ]
    if metadata.get("grade"):
        parts.append(f"产品等级{metadata['grade']}")
    if metadata.get("lifecycle"):
        parts.append(metadata["lifecycle"])
    if aliases:
        parts.append(f"旧称/别名：{'、'.join(aliases)}")
    return "，".join(parts) + "。"


def _build_2026_description(
    name: str,
    metadata: dict[str, str],
    aliases: list[str],
    points: list[dict[str, Any]],
) -> str:
    parts = [_build_2026_metadata_sentence(name, metadata, aliases)]
    for point in points[:3]:
        content = point.get("content", "")
        if content:
            parts.append(f"{point.get('point_type', '卖点')}：{content}")
    return "\n".join(parts)


def _infer_2026_category(name: str, series: str = "") -> str:
    if series in STANDARD_SERIES_CATEGORY_MAP:
        return STANDARD_SERIES_CATEGORY_MAP[series]
    if any(keyword in name for keyword in ("夹心", "奶冻", "布蕾", "慕斯", "晶冻", "芋泥", "栗子泥")):
        return "烘焙夹心"
    if any(keyword in name for keyword in ("果酱", "茶酱", "开心果", "焦糖", "糖浆", "香草", "杏仁", "抹茶", "可可", "斑斓", "零卡", "黑芝麻", "巧克力酱")):
        return "烘焙调味"
    if any(keyword in name for keyword in ("色素", "色粉", "果蔬粉", "竹炭", "红丝绒")):
        return "烘焙调色"
    if any(keyword in name for keyword in ("翻糖", "拉线", "手绘", "糖珠", "脆皮", "肉松")):
        return "烘焙装饰"
    if any(keyword in name for keyword in ("刀叉", "盒装", "纸盘", "包装", "蛋糕盒", "丝带", "保温袋", "裱花袋", "模具")):
        return "烘焙配件"
    return "烘焙配件"


def _extract_heading_block(text: str, heading: str) -> str:
    start = text.find(heading)
    if start < 0:
        return ""
    next_heading = re.search(r"^#{2,3}\s+", text[start + len(heading):], flags=re.MULTILINE)
    if not next_heading:
        return text[start + len(heading):]
    end = start + len(heading) + next_heading.start()
    return text[start + len(heading):end]


def _build_product_description(
    basic_info: dict[str, str], index_info: dict[str, str], section_text: str
) -> str:
    parts: list[str] = []
    if index_info:
        parts.extend(f"{key}：{value}" for key, value in index_info.items() if value)
    parts.extend(f"{key}：{value}" for key, value in basic_info.items() if value)
    if not parts:
        content = re.sub(r"^#{2,3}.+$", "", section_text, flags=re.MULTILINE)
        return _clean_markdown(content)[:1000]
    return "\n".join(parts)


def _parse_price(text: str) -> float:
    if not text:
        return 0.0
    mao = re.search(r"(\d+(?:\.\d+)?)\s*毛", text)
    if mao:
        return round(float(mao.group(1)) / 10, 2)
    yuan = re.search(r"(\d+(?:\.\d+)?)\s*(?:元|块)", text)
    if yuan:
        return float(yuan.group(1))
    return 0.0


def _find_script_columns(header: list[str]) -> dict[str, int | None]:
    return {
        "id": _find_col(header, lambda value: "编号" in value or value == "展会时间"),
        "high": _find_col(header, lambda value: "高成交" in value),
        "type": _find_col(header, lambda value: "类型" in value),
        "script": _find_col(header, lambda value: "视频脚本" in value),
        "product": _find_col(header, lambda value: "展会产品" in value),
        "notes": _find_col(header, lambda value: "注意事项" in value),
        "case": _find_col(header, lambda value: "案例视频" in value or "参考视频" in value),
    }


def _find_col(header: list[str], predicate) -> int | None:
    for index, value in enumerate(header):
        if predicate(value):
            return index
    return None


def _get_cell(row: tuple[Any, ...], column: int | None) -> str:
    if column is None or column >= len(row):
        return ""
    return _cell_text(row[column])


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    return re.sub(r"\n{3,}", "\n\n", text)


def _clean_markdown(text: str) -> str:
    text = text.replace("\\-", "-").replace("<br>", " ")
    text = re.sub(r"\*\*", "", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip(" ：:\t")


def _write_product_material_file(material_dir: Path, product: ProductInput, manual_text: str) -> Path:
    excerpts = _extract_manual_excerpts(manual_text, product.name)
    content = [
        f"# {product.name} 资料",
        "",
        "## 产品知识库",
        "",
        product.section_text,
    ]
    if excerpts:
        content.extend(["", "## 产品手册补充", "", excerpts])
    path = material_dir / f"product_material_{_safe_filename(product.name)}.md"
    path.write_text("\n".join(content).strip() + "\n", encoding="utf-8")
    return path


def _extract_manual_excerpts(manual_text: str, product_name: str) -> str:
    aliases = MANUAL_ALIASES.get(product_name, [product_name])
    pages = re.split(r"(?=^## 第 \d+ 页)", manual_text, flags=re.MULTILINE)
    matched: list[str] = []
    for page in pages:
        if any(alias in page for alias in aliases):
            matched.append(page.strip())
    return "\n\n".join(matched)[:12000]


def _safe_filename(value: str) -> str:
    return re.sub(r'[<>:"/\\|?*\s]+', "_", value).strip("_")


if __name__ == "__main__":
    raise SystemExit(main())
