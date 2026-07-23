"""脚本模板 & 爆款脚本 API"""
import asyncio
import hashlib
import json
import mimetypes
import os
import re
import threading
import uuid
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import List, Optional

from fastapi import APIRouter, Depends, File, Form, Header, HTTPException, Query, Request, UploadFile
from fastapi.responses import FileResponse
from sqlalchemy import or_
from sqlalchemy.orm import Session

from config import MAX_UPLOAD_SIZE
from database import SessionLocal, get_db
from routers.jobs import owner_key_for_request
from models import (
    QianchuanImportBatch,
    QianchuanMaterialPerformance,
    QianchuanScriptBinding,
    ScriptTemplate,
    ViralScript,
)
from schemas import (
    ApiResponse,
    ScriptTemplateCreate,
    ScriptTemplateOut,
    ViralScriptCreate,
    ViralScriptOut,
    ViralScriptPageOut,
)
from services import task_queue
from services.ai_service import ai_service
from services.product_markdown_importer import normalize_product_name
from services.qianchuan_importer import parse_qianchuan_workbook
from services.upload_limits import read_upload_bytes

router = APIRouter()

LOCAL_TXT_SCRIPT_SOURCE_DIR = os.getenv(
    "LOCAL_TXT_SCRIPT_SOURCE_DIR",
    r"\\192.168.0.118\法采共享盘2026\视频素材\源素材",
)
LOCAL_TXT_SCAN_SESSION_FACTORY = SessionLocal
LOCAL_TXT_SCAN_ERROR_LIMIT = 50
MAX_TXT_BATCH_FILES = int(os.getenv("MAX_TXT_BATCH_FILES", "50"))
QIANCHUAN_AUTO_BIND_SCORE = 12
QIANCHUAN_STRUCTURED_AUTO_BIND_SCORE = 24
QIANCHUAN_STRUCTURED_AUTO_BIND_MARGIN = 4
QIANCHUAN_AUTO_HIGH_FLAG = "qianchuan_auto_high_conversion"
BASE_DIR = Path(__file__).resolve().parents[1]
SCRIPT_WORKBOOK_IMPORT_MAX_SIZE = int(os.getenv(
    "SCRIPT_WORKBOOK_IMPORT_MAX_SIZE",
    str(max(MAX_UPLOAD_SIZE, 100 * 1024 * 1024)),
))
WORKBOOK_IMPORT_SESSION_FACTORY = SessionLocal
WORKBOOK_IMPORT_ERROR_LIMIT = 50
WORKBOOK_IMPORT_UPLOAD_DIR = BASE_DIR / "data" / "workbook_imports"
VIRAL_SCRIPT_IMAGE_DIR = Path(os.getenv(
    "VIRAL_SCRIPT_IMAGE_DIR",
    str(BASE_DIR / "data" / "viral_script_images"),
))

_local_txt_scan_lock = threading.RLock()
_local_txt_scan_state = {
    "is_running": False,
    "source_dir": LOCAL_TXT_SCRIPT_SOURCE_DIR,
    "recursive": True,
    "total": 0,
    "processed": 0,
    "success": 0,
    "skipped": 0,
    "error_count": 0,
    "errors": [],
    "ids": [],
    "started_at": None,
    "finished_at": None,
    "message": "待扫描",
    "job_id": None,
}

_workbook_import_lock = threading.RLock()
_workbook_import_state = {
    "is_running": False,
    "filename": "",
    "total": 0,
    "processed": 0,
    "created": 0,
    "updated": 0,
    "skipped": 0,
    "image_count": 0,
    "index_error_count": 0,
    "error_count": 0,
    "errors": [],
    "ids": [],
    "started_at": None,
    "finished_at": None,
    "message": "待导入",
    "job_id": None,
}

_qianchuan_auto_match_lock = threading.RLock()
_qianchuan_auto_match_state = {
    "is_running": False,
    "started_at": None,
    "finished_at": None,
    "last_result": None,
    "message": "待匹配",
}


def format_script(text: str) -> str:
    """将带时间戳/标记的脚本格式化为纯段落形式"""
    # 去掉时间戳如 (00:00-00:03)、(0-3s) 等
    text = re.sub(r'[（(]\d{1,2}:\d{2}[-~]\d{1,2}:\d{2}[）)]', '', text)
    text = re.sub(r'[（(]\d{1,2}[-~]\d{1,2}[sS秒][）)]', '', text)
    # 去掉所有形如【xxx-s】或【xxx秒】的段落标记（如【开场钩子-0-3s】、【痛点-3-10s】）
    text = re.sub(r'【[^】]*?\d+[-~]\d+[sS秒][^】]*?】', '', text)
    # 去掉纯时间标记如 00:00-00:03 或 00:00
    text = re.sub(r'\d{1,2}:\d{2}\s*[-~]\s*\d{1,2}:\d{2}', '', text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    # 去掉行首时间戳如 "00:00  文本" 或 "00:00"
    text = re.sub(r'^\d{1,2}:\d{2}\s*', '', text, flags=re.MULTILINE)
    # 去掉行内空格后，合并所有行为一整段话
    text = re.sub(r'\n\s*', '', text)
    return text.strip()


async def analyze_script_ai(text: str, user_type: str = "") -> dict:
    """AI分析脚本：识别视频类型、结构、跑量点"""
    if not ai_service.is_available:
        return {}

    prompt = f"""分析以下带货短视频脚本，输出JSON：

{text[:3000]}

提取：
1. video_type：视频类型（从以下选一个最匹配的：机制类/痛点类/需求类/认知类/达人分享类/制作方便/成本低/对比类/情绪类/场景类）
2. structure：脚本结构（如"价格钩子→痛点→卖点展示→促销叠加→强CTA"）
3. viral_points：跑量点分析（30字以内，如"低价冲击+年货节紧迫感+体感对比"）
4. tags：标签（逗号分隔，如"刀叉盘,年货节,价格锚点"）

只输出JSON：{{"video_type":"...","structure":"...","viral_points":"...","tags":"..."}}"""

    try:
        result = await ai_service.chat([
            {"role": "system", "content": "你是短视频脚本分析专家，只输出JSON"},
            {"role": "user", "content": prompt},
        ], temperature=0.2, interface_key="viral_script_analyze")
        start = result.find('{')
        end = result.rfind('}') + 1
        if start >= 0 and end > start:
            return json.loads(result[start:end])
    except:
        pass
    return {}


def _decode_txt(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gbk", "gb2312", "utf-16"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("无法识别文件编码，请使用 UTF-8、GBK 或 UTF-16 文本")


def _title_from_filename(filename: str | None) -> str:
    name = (filename or "").replace("\\", "/").rsplit("/", 1)[-1].strip()
    if not name:
        return "导入脚本"
    if "." in name:
        name = name.rsplit(".", 1)[0]
    return name.strip() or "导入脚本"


def _reset_local_txt_scan_state() -> None:
    with _local_txt_scan_lock:
        _local_txt_scan_state.update({
            "is_running": False,
            "source_dir": LOCAL_TXT_SCRIPT_SOURCE_DIR,
            "recursive": True,
            "total": 0,
            "processed": 0,
            "success": 0,
            "skipped": 0,
            "error_count": 0,
            "errors": [],
            "ids": [],
            "started_at": None,
            "finished_at": None,
            "message": "待扫描",
            "job_id": None,
        })


def _local_txt_scan_snapshot() -> dict:
    with _local_txt_scan_lock:
        snapshot = dict(_local_txt_scan_state)
        snapshot["errors"] = list(_local_txt_scan_state["errors"])
        snapshot["ids"] = list(_local_txt_scan_state["ids"])
    from services.job_runs import latest_job
    snapshot["job_run"] = latest_job("local_script_scan")
    return snapshot


def _update_local_txt_scan_state(**changes) -> None:
    with _local_txt_scan_lock:
        _local_txt_scan_state.update(changes)
        snapshot = dict(_local_txt_scan_state)
    if snapshot.get("job_id"):
        from services.job_runs import update_job
        update_job(
            snapshot["job_id"],
            current=snapshot.get("processed", 0),
            total=snapshot.get("total", 0),
            message=snapshot.get("message", ""),
        )


def _append_local_txt_scan_error(message: str) -> None:
    with _local_txt_scan_lock:
        _local_txt_scan_state["error_count"] += 1
        if len(_local_txt_scan_state["errors"]) < LOCAL_TXT_SCAN_ERROR_LIMIT:
            _local_txt_scan_state["errors"].append(message)


def _normalized_file_path(path: str | os.PathLike[str]) -> str:
    return os.path.normcase(os.path.abspath(os.path.normpath(str(path))))


def _script_content_hash(script_text: str) -> str:
    return hashlib.sha256((script_text or "").encode("utf-8")).hexdigest()


def _relative_path_for_display(path: Path, source_dir: Path) -> str:
    try:
        relative_path = os.path.relpath(str(path), str(source_dir))
    except ValueError:
        relative_path = path.name
    return relative_path.replace("\\", "/")


def _title_from_relative_path(relative_path: str) -> str:
    stem = os.path.splitext(relative_path.replace("\\", "/"))[0]
    title = " / ".join(part for part in stem.split("/") if part).strip()
    if not title:
        title = "导入脚本"
    if len(title) > 300:
        title = "..." + title[-297:]
    return title


def _combine_tags(base_tags: str, ai_tags: str) -> str:
    parts = [part.strip() for part in (base_tags or "").split(",") if part.strip()]
    parts.extend(part.strip() for part in (ai_tags or "").split(",") if part.strip())
    return ",".join(dict.fromkeys(parts))


def _reset_workbook_import_state() -> None:
    with _workbook_import_lock:
        _workbook_import_state.update({
            "is_running": False,
            "filename": "",
            "total": 0,
            "processed": 0,
            "created": 0,
            "updated": 0,
            "skipped": 0,
            "image_count": 0,
            "index_error_count": 0,
            "error_count": 0,
            "errors": [],
            "ids": [],
            "started_at": None,
            "finished_at": None,
            "message": "待导入",
            "job_id": None,
        })


def _workbook_import_snapshot() -> dict:
    with _workbook_import_lock:
        snapshot = dict(_workbook_import_state)
        snapshot["errors"] = list(_workbook_import_state["errors"])
        snapshot["ids"] = list(_workbook_import_state["ids"])
    from services.job_runs import latest_job
    snapshot["job_run"] = latest_job("workbook_import")
    return snapshot


def _update_workbook_import_state(**changes) -> None:
    with _workbook_import_lock:
        _workbook_import_state.update(changes)
        snapshot = dict(_workbook_import_state)
    if snapshot.get("job_id"):
        from services.job_runs import update_job
        update_job(
            snapshot["job_id"],
            current=snapshot.get("processed", 0),
            total=snapshot.get("total", 0),
            message=snapshot.get("message", ""),
        )


def _append_workbook_import_error(message: str) -> None:
    with _workbook_import_lock:
        _workbook_import_state["error_count"] += 1
        if len(_workbook_import_state["errors"]) < WORKBOOK_IMPORT_ERROR_LIMIT:
            _workbook_import_state["errors"].append(message)


def _normalize_workbook_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _workbook_sheet_key(sheet_name: str) -> str:
    return re.sub(r"\s+", "", sheet_name or "")


def _category_from_workbook_sheet(sheet_name: str) -> str:
    key = _workbook_sheet_key(sheet_name)
    category_map = {
        "刀叉（袋装）": "烘焙配件",
        "刀叉(袋装)": "烘焙配件",
        "刀叉（盒装）": "烘焙配件",
        "刀叉(盒装)": "烘焙配件",
        "水性色素": "烘焙调色",
        "水状色素": "烘焙调色",
        "果蔬色素": "烘焙调色",
        "果蔬粉": "烘焙调色",
        "慕斯粉": "烘焙夹心",
        "奶冻粉": "烘焙夹心",
        "布蕾粉": "烘焙夹心",
        "翻糖": "烘焙装饰",
        "彩色翻糖膏": "烘焙装饰",
        "手绘拉线膏": "烘焙装饰",
        "巧克力糖": "烘焙装饰",
        "糖珠": "烘焙装饰",
        "肉松": "烘焙装饰",
    }
    if key in category_map:
        return category_map[key]
    return _infer_local_txt_category_by_keywords(key) or "烘焙配件"


def _video_type_from_workbook(raw_type: str) -> str:
    value = _normalize_workbook_cell(raw_type)
    if not value:
        return "机制类"
    video_type_map = {
        "机制": "机制类",
        "机制类": "机制类",
        "痛点": "痛点类",
        "痛点类": "痛点类",
        "需求": "需求类",
        "需求类": "需求类",
        "认知": "认知类",
        "认知类": "认知类",
        "对比": "对比类",
        "对比类": "对比类",
        "场景": "场景类",
        "场景类": "场景类",
        "爆款翻拍": "达人分享类",
        "达人分享": "达人分享类",
        "达人分享类": "达人分享类",
        "创意": "场景类",
        "AI生成": "机制类",
        "AI": "机制类",
    }
    return video_type_map.get(value, value if value.endswith("类") else "机制类")


def _truthy_high_conversion(value) -> bool:
    text = _normalize_workbook_cell(value).lower()
    if not text:
        return False
    return text not in {"0", "false", "no", "n", "否", "不是", "无", "nan", "none"}


def _workbook_title(sheet_name: str, script_code: str, raw_type: str) -> str:
    parts = [
        _normalize_workbook_cell(sheet_name),
        _normalize_workbook_cell(script_code),
        _normalize_workbook_cell(raw_type),
    ]
    title = " / ".join(part for part in parts if part).strip() or "Excel脚本"
    if len(title) > 300:
        title = "..." + title[-297:]
    return title


def _workbook_tags(sheet_name: str, raw_type: str, has_images: bool) -> str:
    tags = ["Excel脚本", _normalize_workbook_cell(sheet_name), _normalize_workbook_cell(raw_type)]
    if has_images:
        tags.append("有蛋糕图")
    return ",".join(dict.fromkeys(tag for tag in tags if tag))


def _workbook_image_extension(image_bytes: bytes, image_format: str | None) -> str:
    ext = (image_format or "").lower().strip(".")
    if ext in {"jpeg", "jpg"}:
        return "jpg"
    if ext in {"png", "gif", "bmp", "webp"}:
        return ext
    if image_bytes.startswith(b"\x89PNG"):
        return "png"
    if image_bytes.startswith(b"\xff\xd8"):
        return "jpg"
    return "bin"


def _image_anchor_position(image) -> tuple[int, int]:
    anchor = getattr(image, "anchor", None)
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return 0, 0
    return int(marker.row) + 1, int(marker.col) + 1


def _collect_workbook_images(workbook) -> dict[str, dict[int, list[dict]]]:
    image_map: dict[str, dict[int, list[dict]]] = {}
    for ws in workbook.worksheets:
        row_map: dict[int, list[dict]] = {}
        for index, image in enumerate(getattr(ws, "_images", []) or [], start=1):
            row_number, col_number = _image_anchor_position(image)
            if row_number <= 0:
                continue
            try:
                image_bytes = image._data()
            except Exception:
                continue
            row_map.setdefault(row_number, []).append({
                "bytes": image_bytes,
                "extension": _workbook_image_extension(image_bytes, getattr(image, "format", None)),
                "column": col_number,
                "index": index,
            })
        image_map[ws.title] = row_map
    return image_map


def _header_index_map(row_values: list[str]) -> dict[str, int]:
    headers: dict[str, int] = {}
    for index, value in enumerate(row_values):
        text = _normalize_workbook_cell(value)
        if text:
            headers[text] = index
    return headers


def _find_workbook_column(headers: dict[str, int], *keywords: str) -> int | None:
    for header, index in headers.items():
        if all(keyword in header for keyword in keywords):
            return index
    return None


def _workbook_script_columns(row_values: list[str]) -> dict[str, int | None]:
    headers = _header_index_map(row_values)
    return {
        "code": _find_workbook_column(headers, "编号"),
        "high": _find_workbook_column(headers, "高成交"),
        "type": _find_workbook_column(headers, "类型"),
        "script": _find_workbook_column(headers, "视频脚本") or _find_workbook_column(headers, "脚本"),
        "image": _find_workbook_column(headers, "蛋糕") or _find_workbook_column(headers, "参考图"),
    }


def _extract_workbook_rows(workbook_path: Path, workbook_sha256: str, filename: str) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path, data_only=True)
    image_map = _collect_workbook_images(workbook)
    rows: list[dict] = []
    for ws in workbook.worksheets:
        header_values = [_normalize_workbook_cell(cell.value) for cell in ws[1]]
        columns = _workbook_script_columns(header_values)
        if columns["script"] is None:
            continue
        for row in ws.iter_rows(min_row=2):
            values = [cell.value for cell in row]

            def cell_at(index: int | None) -> str:
                if index is None or index >= len(values):
                    return ""
                return _normalize_workbook_cell(values[index])

            raw_script = cell_at(columns["script"])
            script_text = format_script(raw_script)
            if len(script_text) < 20:
                continue
            row_number = row[0].row
            script_code = cell_at(columns["code"]) or str(row_number)
            raw_type = cell_at(columns["type"])
            row_images = list(image_map.get(ws.title, {}).get(row_number, []))
            content_sha256 = _script_content_hash(script_text)
            rows.append({
                "filename": filename,
                "workbook_sha256": workbook_sha256,
                "sheet_name": ws.title,
                "row_number": row_number,
                "script_code": script_code,
                "raw_type": raw_type,
                "category": _category_from_workbook_sheet(ws.title),
                "video_type": _video_type_from_workbook(raw_type),
                "title": _workbook_title(ws.title, script_code, raw_type),
                "script_content": script_text,
                "content_sha256": content_sha256,
                "is_high_conversion": 1 if _truthy_high_conversion(cell_at(columns["high"])) else 0,
                "images": row_images,
            })
    return rows


def _existing_cake_images(script: ViralScript) -> list[dict]:
    data = script.performance_data if isinstance(script.performance_data, dict) else {}
    images = data.get("cake_images")
    return list(images) if isinstance(images, list) else []


def _find_existing_workbook_script(db: Session, row_info: dict):
    for script in db.query(ViralScript).all():
        data = script.performance_data if isinstance(script.performance_data, dict) else {}
        if (
            data.get("workbook_sha256") == row_info["workbook_sha256"]
            and data.get("sheet_name") == row_info["sheet_name"]
            and str(data.get("row_number")) == str(row_info["row_number"])
            and str(data.get("script_code")) == str(row_info["script_code"])
        ):
            return script
        if data.get("content_sha256") == row_info["content_sha256"]:
            return script
        if _script_content_hash(script.script_content or "") == row_info["content_sha256"]:
            return script
    return None


def _save_workbook_cake_images(script_id: int, row_info: dict) -> list[dict]:
    if not row_info.get("images"):
        return []
    base_dir = Path(VIRAL_SCRIPT_IMAGE_DIR)
    workbook_dir = row_info["workbook_sha256"][:16]
    target_dir = base_dir / workbook_dir / str(script_id)
    target_dir.mkdir(parents=True, exist_ok=True)
    saved: list[dict] = []
    for image in row_info["images"]:
        extension = image.get("extension") or "bin"
        filename = f"row{row_info['row_number']}_{uuid.uuid4().hex[:10]}.{extension}"
        target = target_dir / filename
        image_bytes = image["bytes"]
        target.write_bytes(image_bytes)
        relative_path = f"{workbook_dir}/{script_id}/{filename}"
        media_type = mimetypes.guess_type(filename)[0] or "application/octet-stream"
        saved.append({
            "filename": filename,
            "relative_path": relative_path,
            "url": f"/api/templates/viral/{script_id}/cake-images/{filename}",
            "content_type": media_type,
            "file_size": len(image_bytes),
            "sheet_name": row_info["sheet_name"],
            "row_number": row_info["row_number"],
            "column": image.get("column"),
        })
    return saved


def _workbook_performance_data(row_info: dict, existing_data: dict | None, cake_images: list[dict]) -> dict:
    data = dict(existing_data or {})
    existing_images = data.get("cake_images") if isinstance(data.get("cake_images"), list) else []
    image_keys = {item.get("relative_path") or item.get("filename") for item in existing_images if isinstance(item, dict)}
    merged_images = list(existing_images)
    for image in cake_images:
        key = image.get("relative_path") or image.get("filename")
        if key not in image_keys:
            merged_images.append(image)
            image_keys.add(key)
    data.update({
        "source": "Excel脚本表导入",
        "source_file": row_info["filename"],
        "workbook_sha256": row_info["workbook_sha256"],
        "sheet_name": row_info["sheet_name"],
        "row_number": row_info["row_number"],
        "script_code": row_info["script_code"],
        "original_type": row_info["raw_type"],
        "content_sha256": row_info["content_sha256"],
    })
    if merged_images:
        data["cake_images"] = merged_images
    return data


def _sync_workbook_viral_index(viral: ViralScript, db: Session) -> None:
    try:
        indexed = _sync_viral_index(viral, db)
        if indexed is False:
            with _workbook_import_lock:
                _workbook_import_state["index_error_count"] += 1
    except Exception as exc:
        with _workbook_import_lock:
            _workbook_import_state["index_error_count"] += 1
        _append_workbook_import_error(f"{viral.title}: 索引同步失败：{exc}")


def _import_workbook_scripts(workbook_path: Path, filename: str, workbook_sha256: str, db: Session) -> None:
    """Compatibility import seam retained for existing routes and tests."""
    from services.workbook_import import import_workbook_scripts

    return import_workbook_scripts(workbook_path, filename, workbook_sha256, db)


def _run_workbook_import(workbook_path: str, filename: str, workbook_sha256: str, job_id: int | None = None) -> None:
    db = WORKBOOK_IMPORT_SESSION_FACTORY()
    terminal_status = "succeeded"
    terminal_error = ""
    try:
        _import_workbook_scripts(Path(workbook_path), filename, workbook_sha256, db)
        snapshot = _workbook_import_snapshot()
        _update_workbook_import_state(
            message=(
                f"导入完成：新增 {snapshot['created']}，补充 {snapshot['updated']}，"
                f"跳过 {snapshot['skipped']}，错误 {snapshot['error_count']}"
            )
        )
    except Exception as exc:
        terminal_status = "failed"
        terminal_error = str(exc)
        _append_workbook_import_error(str(exc))
        _update_workbook_import_state(message=f"导入失败：{exc}")
    finally:
        db.close()
        try:
            Path(workbook_path).unlink()
        except FileNotFoundError:
            pass
        _update_workbook_import_state(is_running=False, finished_at=datetime.now().replace(microsecond=0).isoformat())
        if job_id:
            from services.job_runs import finish_job
            snapshot = _workbook_import_snapshot()
            finish_job(
                job_id,
                status=terminal_status,
                message=snapshot.get("message", ""),
                details={key: snapshot.get(key) for key in ("total", "processed", "created", "updated", "skipped", "error_count", "index_error_count")},
                error_summary=terminal_error,
            )


def _run_workbook_import_task(payload: dict) -> None:
    _run_workbook_import(
        str(payload["workbook_path"]),
        str(payload["filename"]),
        str(payload["workbook_sha256"]),
        payload.get("job_id"),
    )


from services.task_queue import register_task_handler

register_task_handler("workbook_import", _run_workbook_import_task)


def _resolve_viral_script_category(db: Session, category: str = "", product_name: str = "") -> str:
    auto_category = category or ""
    if product_name and not auto_category:
        from models import Product
        matched = db.query(Product).filter(Product.name == product_name).first()
        if matched:
            auto_category = matched.category
    return auto_category or "烘焙配件"


def _local_txt_path_tokens(relative_path: str) -> list[str]:
    stem = os.path.splitext(relative_path.replace("\\", "/"))[0]
    return [part.strip() for part in stem.split("/") if part.strip()]


def _infer_local_txt_category_by_keywords(value: str) -> str:
    category_keywords = [
        ("烘焙夹心", ("夹心", "奶冻", "布蕾", "慕斯", "晶冻", "芋泥", "栗子泥")),
        ("烘焙调味", ("果酱", "茶酱", "开心果", "焦糖", "糖浆", "香草", "杏仁", "抹茶", "可可", "斑斓", "零卡", "黑芝麻", "巧克力酱")),
        ("烘焙调色", ("色素", "色粉", "果蔬粉", "竹炭", "红丝绒")),
        ("烘焙装饰", ("翻糖", "拉线", "手绘", "糖珠", "脆皮", "肉松")),
        ("烘焙配件", ("刀叉", "盒装", "纸盘", "包装", "蛋糕盒", "丝带", "保温袋", "裱花袋", "模具")),
    ]
    for category, keywords in category_keywords:
        if any(keyword in value for keyword in keywords):
            return category
    return ""


def _resolve_local_txt_category_from_path(db: Session, relative_path: str) -> str:
    tokens = _local_txt_path_tokens(relative_path)
    if not tokens:
        return ""

    token_keys = [(token, normalize_product_name(token)) for token in tokens if normalize_product_name(token)]
    if not token_keys:
        return ""

    from models import Product

    products = db.query(Product).filter(Product.status == "active").all()
    for token, token_key in token_keys[:3]:
        matches = [product for product in products if normalize_product_name(product.name) == token_key]
        if len(matches) == 1 and matches[0].category:
            return matches[0].category

    candidates: list[tuple[int, int, str]] = []
    for product in products:
        product_key = normalize_product_name(product.name)
        if len(product_key) < 2 or not product.category:
            continue
        for token, token_key in token_keys[:4]:
            if product_key in token_key or token_key in product_key:
                candidates.append((len(product_key), product.id, product.category))
                break
    if candidates:
        candidates.sort(key=lambda item: item[0], reverse=True)
        top_score = candidates[0][0]
        top_categories = {category for score, _product_id, category in candidates if score == top_score}
        if len(top_categories) == 1:
            return next(iter(top_categories))

    return _infer_local_txt_category_by_keywords("".join(tokens[:4]))


def _resolve_local_txt_script_category(
    db: Session,
    *,
    category: str,
    product_name: str,
    relative_path: str,
    path_category: str = "",
) -> str:
    path_category = path_category or _resolve_local_txt_category_from_path(db, relative_path)
    if path_category:
        return path_category
    return _resolve_viral_script_category(db, category=category, product_name=product_name)


def _find_existing_local_txt_script(db: Session, local_path: str, content_sha256: str):
    normalized_path = _normalized_file_path(local_path)
    for script in db.query(ViralScript).all():
        data = script.performance_data if isinstance(script.performance_data, dict) else {}
        existing_path = data.get("local_path")
        if existing_path and _normalized_file_path(existing_path) == normalized_path:
            return script
        if data.get("content_sha256") == content_sha256:
            return script
        if _script_content_hash(script.script_content or "") == content_sha256:
            return script
    return None


async def _analyze_local_txt_script(script_text: str, video_type: str) -> dict:
    try:
        return await analyze_script_ai(script_text, video_type)
    except Exception:
        return {}


def _discover_local_txt_files(source_dir: Path) -> list[Path]:
    files: list[Path] = []
    for current_root, _dir_names, file_names in os.walk(source_dir):
        file_names.sort()
        for name in file_names:
            if name.lower().endswith(".txt"):
                files.append(Path(current_root) / name)
    files.sort(key=lambda item: _relative_path_for_display(item, source_dir).lower())
    return files


async def _scan_local_txt_scripts(
    *,
    source_dir: Path,
    category: str,
    video_type: str,
    tags: str,
    product_name: str,
    db: Session,
) -> None:
    """Compatibility import seam retained for existing routes and tests."""
    from services.local_script_scan import scan_local_txt_scripts

    return await scan_local_txt_scripts(
        source_dir=source_dir,
        category=category,
        video_type=video_type,
        tags=tags,
        product_name=product_name,
        db=db,
    )


def _run_local_txt_scan(source_dir: str, category: str, video_type: str, tags: str, product_name: str, job_id: int | None = None) -> None:
    db = LOCAL_TXT_SCAN_SESSION_FACTORY()
    terminal_status = "succeeded"
    terminal_error = ""
    try:
        asyncio.run(_scan_local_txt_scripts(
            source_dir=Path(source_dir),
            category=category,
            video_type=video_type,
            tags=tags,
            product_name=product_name,
            db=db,
        ))
        snapshot = _local_txt_scan_snapshot()
        _update_local_txt_scan_state(
            message=(
                f"扫描完成：新增 {snapshot['success']}，跳过 {snapshot['skipped']}，"
                f"错误 {snapshot['error_count']}"
            )
        )
    except Exception as exc:
        terminal_status = "failed"
        terminal_error = str(exc)
        _append_local_txt_scan_error(str(exc))
        _update_local_txt_scan_state(message=f"扫描失败：{exc}")
    finally:
        db.close()
        _update_local_txt_scan_state(is_running=False, finished_at=datetime.now().replace(microsecond=0).isoformat())
        if job_id:
            from services.job_runs import finish_job
            snapshot = _local_txt_scan_snapshot()
            finish_job(
                job_id,
                status=terminal_status,
                message=snapshot.get("message", ""),
                details={key: snapshot.get(key) for key in ("total", "processed", "success", "skipped", "error_count")},
                error_summary=terminal_error,
            )


def _run_local_txt_scan_task(payload: dict) -> None:
    _run_local_txt_scan(
        str(payload["source_dir"]),
        str(payload.get("category") or ""),
        str(payload.get("video_type") or ""),
        str(payload.get("tags") or "本地txt"),
        str(payload.get("product_name") or ""),
        payload.get("job_id"),
    )


task_queue.register_task_handler("local_script_scan", _run_local_txt_scan_task)


# ========== 脚本模板管理 ==========
def list_templates(
    video_type: Optional[str] = Query(None, description="视频类型筛选"),
    db: Session = Depends(get_db),
):
    """获取模板列表"""
    query = db.query(ScriptTemplate)
    if video_type:
        query = query.filter(ScriptTemplate.video_type == video_type)
    return query.order_by(ScriptTemplate.id).all()


def list_video_types(db: Session = Depends(get_db)):
    """获取所有视频类型"""
    types = db.query(ScriptTemplate.video_type).distinct().order_by(
        ScriptTemplate.video_type
    ).all()
    return [t[0] for t in types]


def get_template(template_id: int, db: Session = Depends(get_db)):
    """获取模板详情"""
    template = db.query(ScriptTemplate).filter(
        ScriptTemplate.id == template_id
    ).first()
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")
    return template


def create_template(data: ScriptTemplateCreate, db: Session = Depends(get_db)):
    """创建脚本模板"""
    template = ScriptTemplate(**data.model_dump())
    db.add(template)
    db.commit()
    db.refresh(template)
    return template


def delete_template(template_id: int, db: Session = Depends(get_db)):
    """删除脚本模板"""
    template = db.query(ScriptTemplate).filter(
        ScriptTemplate.id == template_id
    ).first()
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")
    db.delete(template)
    db.commit()
    return ApiResponse(message="模板已删除")


# ========== 爆款脚本库 ==========
def list_viral_scripts(
    category: Optional[str] = Query(None, description="品类筛选"),
    video_type: Optional[str] = Query(None, description="视频类型筛选"),
    q: Optional[str] = None,
    page: int = 1,
    per_page: int = 24,
    sort: str = "desc",
    high_only: bool = Query(False, description="仅返回高成交脚本"),
    db: Session = Depends(get_db),
):
    """获取爆款脚本列表"""
    page = max(1, int(page or 1))
    per_page = max(1, min(int(per_page or 24), 200))
    query = db.query(ViralScript)
    if category:
        query = query.filter(ViralScript.category == category)
    if video_type:
        query = query.filter(ViralScript.video_type == video_type)
    if q:
        needle = f"%{q.strip()}%"
        if needle != "%%":
            query = query.filter(or_(
                ViralScript.title.ilike(needle),
                ViralScript.script_content.ilike(needle),
                ViralScript.tags.ilike(needle),
                ViralScript.category.ilike(needle),
                ViralScript.video_type.ilike(needle),
            ))
    if high_only:
        query = query.filter(ViralScript.is_high_conversion == 1)
    total = query.count()
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = min(page, total_pages)
    order_by = (ViralScript.created_at.asc(), ViralScript.id.asc()) if sort == "oldest" else (ViralScript.created_at.desc(), ViralScript.id.desc())
    scripts = query.order_by(*order_by).offset((page - 1) * per_page).limit(per_page).all()
    return ViralScriptPageOut(
        items=[ViralScriptOut.model_validate(script) for script in scripts],
        total=total,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
    )


# ========== RAG: 语义搜索 + 索引管理 ==========

def semantic_search_scripts(
    q: str = Query(..., description="自然语言搜索查询"),
    limit: int = Query(10, ge=1, le=50),
    video_type: Optional[str] = Query(None),
    high_only: bool = Query(False),
    db: Session = Depends(get_db),
):
    """语义搜索爆款脚本/参考脚本"""
    try:
        from vector_store.script_store import ScriptVectorStore
        svs = ScriptVectorStore()
        results = svs.search(q, limit=limit, video_type=video_type)
        if not results:
            return []
        viral_ids = [r["db_id"] for r in results if r.get("source") == "viral"]
        ref_ids = [r["db_id"] for r in results if r.get("source") == "reference"]
        out = []
        if viral_ids:
            virals = db.query(ViralScript).filter(ViralScript.id.in_(viral_ids)).all()
            vmap = {v.id: v for v in virals}
            for r in results:
                if r.get("source") == "viral" and r["db_id"] in vmap:
                    s = vmap[r["db_id"]]
                    out.append({
                        "id": s.id, "title": s.title, "category": s.category,
                        "video_type": s.video_type, "tags": s.tags,
                        "script_content": (s.script_content or "")[:300],
                        "performance_data": s.performance_data,
                        "is_high_conversion": bool(s.is_high_conversion),
                        "source": "viral", "distance": r.get("distance", 0),
                        "created_at": str(s.created_at) if s.created_at else None,
                    })
        if ref_ids:
            refs = db.query(ReferenceScript).filter(ReferenceScript.id.in_(ref_ids)).all()
            rmap = {r.id: r for r in refs}
            for r in results:
                if r.get("source") == "reference" and r["db_id"] in rmap:
                    s = rmap[r["db_id"]]
                    out.append({
                        "id": s.id, "title": s.title,
                        "video_type": s.video_type, "tags": s.tags,
                        "script_content": (s.script_content or "")[:300],
                        "is_high_conversion": bool(s.is_high_conversion),
                        "source": "reference", "distance": r.get("distance", 0),
                        "created_at": str(s.created_at) if s.created_at else None,
                    })
        if high_only:
            out = [s for s in out if s.get("is_high_conversion")]
        out.sort(key=lambda x: x.get("distance", 1))
        return out[:limit]
    except Exception as e:
        raise HTTPException(
            status_code=503,
            detail=f"脚本语义检索失败，请检查 ARK_API_KEY、ARK_BASE_URL、EMBEDDING_MODEL_NAME 和火山方舟 endpoint 权限: {e}",
        )


def reindex_scripts(db: Session = Depends(get_db)):
    """批量重建所有脚本的向量索引"""
    try:
        from vector_store import VectorStoreError
        from vector_store.script_store import ScriptVectorStore
        svs = ScriptVectorStore()
        svs.store.reset_script_collection()
        count = svs.index_all_scripts(db)
        return ApiResponse(message=f"已重建 {count} 个脚本的向量索引")
    except VectorStoreError as e:
        raise HTTPException(
            status_code=503,
            detail=f"索引未重建：向量集合重置或写入失败，已停止以避免混用旧索引；请检查 ARK_API_KEY、ARK_BASE_URL、EMBEDDING_MODEL_NAME 和火山方舟 endpoint 权限: {e}",
        )
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"索引重建失败，请检查 ARK_API_KEY、ARK_BASE_URL、EMBEDDING_MODEL_NAME 和火山方舟 endpoint 权限: {e}",
        )


def start_local_txt_scan(
    request: Request = None,
    category: str = Form(""),
    video_type: str = Form(""),
    tags: str = Form(""),
    product_name: str = Form(""),
):
    """Start a background scan of the configured local TXT script directory."""
    source_dir = LOCAL_TXT_SCRIPT_SOURCE_DIR
    if not os.path.exists(source_dir):
        raise HTTPException(status_code=404, detail=f"路径不可访问：{source_dir}")

    with _local_txt_scan_lock:
        if _local_txt_scan_state["is_running"]:
            return ApiResponse(message="本地脚本扫描正在运行", data=_local_txt_scan_snapshot())
        from services.job_runs import start_job
        job_id = start_job(
            "local_script_scan",
            message="本地脚本扫描启动中",
            details={"source_dir": source_dir},
            owner_key=owner_key_for_request(request, request.headers.get("X-Facai-Client-Id")) if request is not None and request.headers.get("X-Facai-Client-Id") else None,
            origin_path="/app/import",
        )
        _local_txt_scan_state.update({
            "is_running": True,
            "source_dir": source_dir,
            "recursive": True,
            "total": 0,
            "processed": 0,
            "success": 0,
            "skipped": 0,
            "error_count": 0,
            "errors": [],
            "ids": [],
            "started_at": datetime.now().replace(microsecond=0).isoformat(),
            "finished_at": None,
            "message": "扫描启动中",
            "job_id": job_id,
        })

    task_payload = {
        "source_dir": source_dir,
        "category": category,
        "video_type": video_type,
        "tags": tags or "本地txt",
        "product_name": product_name,
        "job_id": job_id,
    }
    if task_queue.task_worker_status()["alive"]:
        task_queue.enqueue_task("local_script_scan", task_payload, max_attempts=3, job_run_id=job_id)
    else:
        # Router-only test/dev apps do not run the main lifespan worker.
        threading.Thread(
            target=_run_local_txt_scan_task,
            args=(task_payload,),
            name="facai-local-txt-scan-fallback",
            daemon=True,
        ).start()
    return ApiResponse(message="本地脚本扫描已启动", data=_local_txt_scan_snapshot())


def local_txt_scan_status():
    """Return the latest local TXT script scan status."""
    return ApiResponse(message="ok", data=_local_txt_scan_snapshot())


def _round_metric(value: float, digits: int = 4) -> float:
    return round(float(value or 0), digits)


def _qianchuan_material_to_dict(material: QianchuanMaterialPerformance) -> dict:
    return {
        "id": material.id,
        "material_id": material.material_id,
        "material_name": material.material_name,
        "material_evaluation": material.material_evaluation or "",
        "material_duration": material.material_duration or "",
        "material_created_time": material.material_created_time or "",
        "material_source": material.material_source or "",
        "tags": material.tags or "",
        "amount_field": material.amount_field or "",
        "transaction_amount": _round_metric(material.transaction_amount, 2),
        "order_count": int(material.order_count or 0),
        "user_pay_amount": _round_metric(material.user_pay_amount, 2),
        "roi": _round_metric(material.roi, 4),
        "impressions": int(material.impressions or 0),
        "ctr": _round_metric(material.ctr, 4),
        "spend": _round_metric(material.spend, 2),
        "clicks": int(material.clicks or 0),
        "cvr": _round_metric(material.cvr, 4),
        "play_3s_rate": _round_metric(material.play_3s_rate, 4),
        "play_10s_rate": _round_metric(material.play_10s_rate, 4),
        "avg_watch_seconds": _round_metric(material.avg_watch_seconds, 2),
        "completion_rate": _round_metric(material.completion_rate, 4),
        "plan_count": int(material.plan_count or 0),
        "product_count": int(material.product_count or 0),
    }


def _summarize_qianchuan_materials(materials: list[QianchuanMaterialPerformance]) -> dict:
    amount = sum(float(item.transaction_amount or 0) for item in materials)
    spend = sum(float(item.spend or 0) for item in materials)
    impressions = sum(int(item.impressions or 0) for item in materials)
    clicks = sum(int(item.clicks or 0) for item in materials)
    orders = sum(int(item.order_count or 0) for item in materials)
    amount_fields = sorted({item.amount_field for item in materials if item.amount_field})
    return {
        "material_count": len({item.material_id for item in materials}),
        "row_count": len(materials),
        "amount_field": amount_fields[0] if len(amount_fields) == 1 else " / ".join(amount_fields),
        "transaction_amount": _round_metric(amount, 2),
        "spend": _round_metric(spend, 2),
        "roi": _round_metric(amount / spend if spend > 0 else 0, 4),
        "average_order_value": _round_metric(amount / orders if orders > 0 else 0, 2),
        "impressions": impressions,
        "clicks": clicks,
        "ctr": _round_metric(clicks / impressions if impressions > 0 else 0, 4),
        "order_count": orders,
        "cvr": _round_metric(orders / clicks if clicks > 0 else 0, 4),
        "play_3s_rate": _round_metric(_average_metric(materials, "play_3s_rate"), 4),
        "play_10s_rate": _round_metric(_average_metric(materials, "play_10s_rate"), 4),
        "avg_watch_seconds": _round_metric(_average_metric(materials, "avg_watch_seconds"), 2),
        "completion_rate": _round_metric(_average_metric(materials, "completion_rate"), 4),
    }


def _average_metric(materials: list[QianchuanMaterialPerformance], field: str) -> float:
    values = [float(getattr(item, field) or 0) for item in materials if getattr(item, field) not in (None, 0)]
    return sum(values) / len(values) if values else 0.0


def _qianchuan_binding_to_dict(
    binding: QianchuanScriptBinding,
    materials: list[QianchuanMaterialPerformance],
) -> dict:
    latest = materials[-1] if materials else None
    data = _summarize_qianchuan_materials(materials)
    data.update({
        "id": binding.id,
        "material_id": binding.material_id,
        "material_name": binding.material_name or (latest.material_name if latest else ""),
        "latest": _qianchuan_material_to_dict(latest) if latest else None,
    })
    return data


def _qianchuan_tokens(value: str) -> set[str]:
    text = (value or "").lower().replace(".mp4", "")
    text = re.sub(r"20(\d{2})", r"\1", text)
    parts = re.split(r"[\s/\\_\-+()（）【】\[\]·.,，。:：]+", text)
    blocked = {
        "脚本", "脚本参考", "文案", "法采", "烘焙", "品质", "需求", "机制", "痛点", "价格",
        "烘焙调味", "烘焙配件", "烘焙调色", "烘焙造型", "调味", "配件",
    }
    return {part for part in parts if len(part) >= 2 and part not in blocked and not part.isdigit()}


QIANCHUAN_PRODUCT_ALIAS_GROUPS = [
    {"茶酱", "调味茶酱", "调味果酱"},
    {"翻糖", "翻糖膏"},
    {"果蔬粉", "果蔬色素"},
]
QIANCHUAN_GENERIC_PRODUCT_TERMS = {
    "烘焙调味", "烘焙配件", "烘焙调色", "烘焙造型", "调味", "配件", "素材", "产品",
}
QIANCHUAN_VIDEO_TYPE_TERMS = ("需求", "机制", "品质", "痛点", "价格", "场景")

WORKBOOK_QIANCHUAN_YEAR = 2026
WORKBOOK_SCRIPT_CODE_RE = re.compile(r"(?<!\d)(\d{1,2})\s*-\s*(\d{1,2})\s*-\s*(\d{1,2})(?!\d)")
WORKBOOK_QIANCHUAN_PRODUCT_ALIASES = {
    "刀叉（袋装）": {"袋装刀叉", "刀叉"},
    "刀叉(袋装)": {"袋装刀叉", "刀叉"},
    "刀叉（盒装）": {"盒装刀叉", "刀叉"},
    "刀叉(盒装)": {"盒装刀叉", "刀叉"},
    "翻糖": {"翻糖", "翻糖膏"},
    "彩色翻糖膏": {"彩色翻糖膏", "翻糖膏", "翻糖"},
    "水性色素": {"水性色素", "水状色素", "浅柔色素", "色素"},
    "水状色素": {"水状色素", "水性色素", "浅柔色素", "色素"},
    "果蔬色素": {"果蔬色素", "果蔬粉", "色素"},
    "果蔬粉": {"果蔬粉", "果蔬色素"},
    "慕斯粉": {"慕斯粉"},
    "奶冻粉": {"奶冻粉"},
    "布蕾粉": {"布蕾粉"},
    "手绘拉线膏": {"手绘拉线膏", "拉线膏"},
    "巧克力糖": {"巧克力糖"},
    "糖珠": {"糖珠"},
    "肉松": {"肉松"},
}
WORKBOOK_QIANCHUAN_TYPE_ALIASES = {
    "需求": {"需求"},
    "需求类": {"需求"},
    "机制": {"机制"},
    "机制类": {"机制"},
    "痛点": {"痛点"},
    "痛点类": {"痛点"},
    "价格": {"价格"},
    "价格类": {"价格"},
    "对比": {"对比"},
    "对比类": {"对比"},
    "认知": {"认知"},
    "认知类": {"认知"},
    "场景": {"场景", "情景"},
    "场景类": {"场景", "情景"},
    "情景": {"场景", "情景"},
    "情景演绎": {"场景", "情景"},
    "爆款翻拍": {"爆款翻拍", "爆款", "翻拍", "达人分享"},
    "达人分享": {"爆款翻拍", "爆款", "翻拍", "达人分享"},
    "达人分享类": {"爆款翻拍", "爆款", "翻拍", "达人分享"},
    "创意": {"创意", "场景", "情景"},
    "AI生成": {"AI", "AIGC", "机制"},
    "AI创意": {"AI", "AIGC", "创意", "场景"},
}


def _compact_qianchuan_text(value: str) -> str:
    return re.sub(r"\s+", "", (value or "").lower().replace(".mp4", ""))


def _qianchuan_script_text(script: ViralScript) -> str:
    return _compact_qianchuan_text(" ".join([
        script.title or "",
        script.category or "",
        script.video_type or "",
        (script.script_content or "")[:500],
        script.tags or "",
    ]))


def _extract_qianchuan_dates(value: str) -> set[str]:
    text = (value or "").lower()
    dates: set[str] = set()
    for year, month, day in re.findall(r"(?:20)?(\d{2})[.\-/年](\d{1,2})[.\-/月](\d{1,2})", text):
        dates.add(f"{int(year):02d}.{int(month)}.{int(day)}")
        dates.add(f"{int(month)}.{int(day)}")
    for month, day in re.findall(r"(?<!\d)(\d{1,2})[.\-/月](\d{1,2})(?!\d)", text):
        dates.add(f"{int(month)}.{int(day)}")
    return dates


def _workbook_script_code_target_date(script_code: str, date_offset_days: int = 1) -> str | None:
    match = WORKBOOK_SCRIPT_CODE_RE.search(str(script_code or ""))
    if not match:
        return None
    month, day, _sequence = (int(part) for part in match.groups())
    try:
        target = date(WORKBOOK_QIANCHUAN_YEAR, month, day) + timedelta(days=int(date_offset_days or 0))
    except ValueError:
        return None
    return f"{target.year % 100:02d}.{target.month}.{target.day}"


def _workbook_qianchuan_aliases_for_sheet(sheet_name: str) -> set[str]:
    key = _workbook_sheet_key(sheet_name)
    return {
        _compact_qianchuan_text(alias)
        for alias in WORKBOOK_QIANCHUAN_PRODUCT_ALIASES.get(key, set())
        if alias
    }


def _workbook_qianchuan_type_aliases(raw_type: str, video_type: str = "") -> set[str]:
    values = [
        _normalize_workbook_cell(raw_type),
        _normalize_workbook_cell(video_type).replace("类", ""),
    ]
    for value in values:
        if not value:
            continue
        for key, aliases in WORKBOOK_QIANCHUAN_TYPE_ALIASES.items():
            if key in value:
                return {_compact_qianchuan_text(alias) for alias in aliases if alias}
    return set()


def _is_workbook_script(script: ViralScript, workbook_sha256: str = "") -> bool:
    data = script.performance_data if isinstance(script.performance_data, dict) else {}
    if not data.get("workbook_sha256"):
        return False
    return not workbook_sha256 or data.get("workbook_sha256") == workbook_sha256


def _workbook_qianchuan_script_payload(script: ViralScript, date_offset_days: int) -> dict | None:
    data = script.performance_data if isinstance(script.performance_data, dict) else {}
    target_date = _workbook_script_code_target_date(data.get("script_code", ""), date_offset_days)
    if not target_date:
        return None
    return {
        "script_id": script.id,
        "script_title": script.title,
        "script_code": data.get("script_code", ""),
        "target_date": target_date,
        "sheet_name": data.get("sheet_name", ""),
        "original_type": data.get("original_type", ""),
        "category": script.category,
        "video_type": script.video_type,
    }


def _workbook_qianchuan_candidate_item(
    script: ViralScript,
    material: QianchuanMaterialPerformance,
    *,
    target_date: str,
    matched_aliases: set[str],
    type_aliases: set[str],
) -> dict:
    item = _qianchuan_material_to_dict(material)
    item.update({
        "script_id": script.id,
        "script_title": script.title,
        "script_code": (script.performance_data or {}).get("script_code", "") if isinstance(script.performance_data, dict) else "",
        "target_date": target_date,
        "matched_aliases": sorted(matched_aliases, key=lambda value: (-len(value), value)),
        "matched_types": sorted(type_aliases, key=lambda value: (-len(value), value)),
        "reasons": ["次日日期一致", "产品明确", "类型一致"],
    })
    return item


def _workbook_qianchuan_candidates_for_script(
    script: ViralScript,
    materials: list[QianchuanMaterialPerformance],
    *,
    date_offset_days: int,
) -> tuple[list[dict], dict | None, str]:
    payload = _workbook_qianchuan_script_payload(script, date_offset_days)
    if not payload:
        return [], None, "编号无法解析"

    data = script.performance_data if isinstance(script.performance_data, dict) else {}
    product_aliases = _workbook_qianchuan_aliases_for_sheet(data.get("sheet_name", ""))
    if not product_aliases:
        return [], payload, "产品别名缺失"

    type_aliases = _workbook_qianchuan_type_aliases(data.get("original_type", ""), script.video_type or "")
    if not type_aliases:
        return [], payload, "类型无法规范化"

    date_hits = [
        material for material in materials
        if payload["target_date"] in _extract_qianchuan_dates(material.material_name or "")
    ]
    product_hits: list[tuple[QianchuanMaterialPerformance, set[str]]] = []
    for material in date_hits:
        material_text = _compact_qianchuan_text(material.material_name or "")
        matched_aliases = {alias for alias in product_aliases if alias and alias in material_text}
        if matched_aliases:
            product_hits.append((material, matched_aliases))
    if not product_hits:
        return [], payload, "无次日产品候选"

    candidates = []
    for material, matched_aliases in product_hits:
        material_text = _compact_qianchuan_text(material.material_name or "")
        matched_types = {alias for alias in type_aliases if alias and alias in material_text}
        if matched_types:
            candidates.append(_workbook_qianchuan_candidate_item(
                script,
                material,
                target_date=payload["target_date"],
                matched_aliases=matched_aliases,
                type_aliases=matched_types,
            ))
    if not candidates:
        return [], payload, "无次日产品类型候选"

    candidates.sort(key=lambda item: (float(item.get("transaction_amount") or 0), str(item.get("material_id") or "")), reverse=True)
    return candidates, payload, ""


def _plan_workbook_qianchuan_rematch(
    db: Session,
    *,
    workbook_sha256: str = "",
    date_offset_days: int = 1,
) -> dict:
    scripts = [
        script for script in db.query(ViralScript).order_by(ViralScript.id.asc()).all()
        if _is_workbook_script(script, workbook_sha256)
    ]
    script_ids = {script.id for script in scripts}
    cleared_bindings = db.query(QianchuanScriptBinding).filter(
        QianchuanScriptBinding.script_id.in_(script_ids)
    ).count() if script_ids else 0
    materials = _latest_qianchuan_materials(db)
    bound_non_target = {
        material_id for (material_id,) in db.query(QianchuanScriptBinding.material_id).filter(
            ~QianchuanScriptBinding.script_id.in_(script_ids)
        ).all()
    } if script_ids else {
        material_id for (material_id,) in db.query(QianchuanScriptBinding.material_id).all()
    }

    provisional_matches: list[dict] = []
    review_items: list[dict] = []
    conflicts: list[dict] = []
    no_candidate_count = 0
    for script in scripts:
        candidates, payload, reason = _workbook_qianchuan_candidates_for_script(
            script,
            materials,
            date_offset_days=date_offset_days,
        )
        payload = payload or {
            "script_id": script.id,
            "script_title": script.title,
            "script_code": "",
            "target_date": None,
            "sheet_name": "",
            "original_type": "",
        }
        if not candidates:
            no_candidate_count += 1
            review_items.append({**payload, "reason": reason, "candidates": []})
            continue
        if len(candidates) > 1:
            review_items.append({**payload, "reason": "多个候选，需人工确认", "candidates": candidates[:10]})
            continue
        candidate = candidates[0]
        if candidate["material_id"] in bound_non_target:
            conflicts.append({**payload, "reason": "素材已绑定非 Excel 脚本", "candidates": candidates})
            continue
        provisional_matches.append(candidate)

    by_material: dict[str, list[dict]] = {}
    for item in provisional_matches:
        by_material.setdefault(str(item.get("material_id") or ""), []).append(item)

    matches = []
    for material_id, items in by_material.items():
        if len(items) == 1:
            matches.append(items[0])
        else:
            conflicts.append({
                "material_id": material_id,
                "material_name": items[0].get("material_name") or "",
                "reason": "多个脚本竞争同一素材",
                "candidates": items,
            })

    matches.sort(key=lambda item: int(item.get("script_id") or 0))
    return {
        "total_scripts": len(scripts),
        "processed": len(scripts),
        "material_count": len(materials),
        "date_offset_days": date_offset_days,
        "workbook_sha256": workbook_sha256,
        "target_script_ids": sorted(script_ids),
        "cleared_bindings": cleared_bindings,
        "auto_bindings": len(matches),
        "planned": len(matches),
        "created_bindings": 0,
        "review_count": len(review_items),
        "conflict_count": len(conflicts),
        "no_candidate_count": no_candidate_count,
        "matches": matches,
        "review_items": review_items,
        "conflicts": conflicts,
        "candidates_preview": (conflicts + review_items)[:20],
    }


def _apply_workbook_qianchuan_rematch(db: Session, plan: dict) -> tuple[int, int]:
    script_ids = {
        int(script_id)
        for script_id in plan.get("target_script_ids", [])
        if int(script_id or 0)
    }
    if not script_ids:
        return 0, 0

    old_bindings = db.query(QianchuanScriptBinding).filter(
        QianchuanScriptBinding.script_id.in_(script_ids)
    ).all()
    cleared = len(old_bindings)
    for binding in old_bindings:
        db.delete(binding)
    db.flush()

    created = 0
    for item in plan.get("matches", []):
        script_id = int(item.get("script_id") or 0)
        material_id = str(item.get("material_id") or "").strip()
        if not script_id or not material_id:
            continue
        db.add(QianchuanScriptBinding(
            script_id=script_id,
            material_id=material_id,
            material_name=item.get("material_name") or "",
        ))
        created += 1
    db.commit()

    _sync_qianchuan_high_conversion_for_scripts(script_ids, db)
    return cleared, created


def _qianchuan_product_matches(script: ViralScript, material_name: str) -> tuple[set[str], set[str]]:
    script_text = _qianchuan_script_text(script)
    material_text = _compact_qianchuan_text(material_name)
    matched_aliases: set[str] = set()
    canonical_matches: set[str] = set()
    for group in QIANCHUAN_PRODUCT_ALIAS_GROUPS:
        script_aliases = {term for term in group if term in script_text}
        material_aliases = {term for term in group if term in material_text}
        if script_aliases and material_aliases:
            matched_aliases.update(script_aliases | material_aliases)
            canonical_matches.add(sorted(group, key=len, reverse=True)[0])

    script_tokens = _qianchuan_tokens(" ".join([
        script.title or "",
        (script.script_content or "")[:260],
        script.tags or "",
    ]))
    material_tokens = _qianchuan_tokens(material_name)
    explicit_tokens = {
        token for token in script_tokens & material_tokens
        if token not in QIANCHUAN_GENERIC_PRODUCT_TERMS and not any(char.isdigit() for char in token)
    }
    matched_aliases.update(explicit_tokens)
    canonical_matches.update(explicit_tokens)
    return matched_aliases, canonical_matches


def _structured_qianchuan_candidate(
    script: ViralScript,
    material: QianchuanMaterialPerformance,
) -> dict | None:
    matched_aliases, product_matches = _qianchuan_product_matches(script, material.material_name or "")
    if not product_matches:
        return None

    script_text = _compact_qianchuan_text(" ".join([
        script.title or "",
        script.video_type or "",
        script.category or "",
    ]))
    material_text = _compact_qianchuan_text(material.material_name or "")
    score = 0
    reasons = []
    if matched_aliases:
        alias_score = 24 if any(
            alias in {term for group in QIANCHUAN_PRODUCT_ALIAS_GROUPS for term in group}
            for alias in matched_aliases
        ) else 12
        score += alias_score
        reasons.append("产品/别名明确")

    script_dates = _extract_qianchuan_dates(" ".join([script.title or "", script.script_content or ""]))
    material_dates = _extract_qianchuan_dates(material.material_name or "")
    if script_dates & material_dates:
        score += 6
        reasons.append("日期一致")

    type_hits = {
        term for term in QIANCHUAN_VIDEO_TYPE_TERMS
        if term in script_text and term in material_text
    }
    if type_hits:
        score += 4
        reasons.append("类型一致")

    category_key = (script.category or "").replace("烘焙", "").strip()
    if category_key and category_key not in QIANCHUAN_GENERIC_PRODUCT_TERMS and category_key in material_text:
        score += 2
        reasons.append("类目弱匹配")

    item = _qianchuan_material_to_dict(material)
    item.update({
        "script_id": script.id,
        "score": score,
        "matched_aliases": sorted(matched_aliases, key=lambda value: (-len(value), value)),
        "reasons": reasons,
    })
    return item


def _score_qianchuan_candidate(script: ViralScript, material: QianchuanMaterialPerformance) -> int:
    structured = _structured_qianchuan_candidate(script, material)
    if structured:
        return int(structured.get("score") or 0)
    script_tokens = _qianchuan_tokens(" ".join([
        script.title or "",
        script.category or "",
        script.video_type or "",
        (script.script_content or "")[:260],
    ]))
    material_tokens = _qianchuan_tokens(material.material_name or "")
    overlap = script_tokens & material_tokens
    score = sum(min(6, len(token)) for token in overlap)
    title = script.title or ""
    if script.category and script.category.replace("烘焙", "") in (material.material_name or ""):
        score += 4
    for token in script_tokens:
        if len(token) >= 4 and token in title and token in (material.material_name or "").lower():
            score += 4
    return score


def _candidate_materials(
    script: ViralScript,
    db: Session,
    bound_ids: set[str],
    limit: int = 12,
) -> list[dict]:
    latest_by_material: dict[str, QianchuanMaterialPerformance] = {}
    rows = db.query(QianchuanMaterialPerformance).order_by(
        QianchuanMaterialPerformance.transaction_amount.desc(),
        QianchuanMaterialPerformance.id.desc(),
    ).all()
    for row in rows:
        if row.material_id in bound_ids or row.material_id in latest_by_material:
            continue
        latest_by_material[row.material_id] = row

    scored = []
    for material in latest_by_material.values():
        item = _structured_qianchuan_candidate(script, material)
        if item:
            scored.append(item)
            continue
        score = _score_qianchuan_candidate(script, material)
        if score > 0:
            fallback = _qianchuan_material_to_dict(material)
            fallback["score"] = score
            fallback["matched_aliases"] = []
            fallback["reasons"] = ["弱文本匹配"]
            scored.append(fallback)
    scored.sort(key=lambda item: (item["score"], item["transaction_amount"]), reverse=True)
    return scored[:limit]


def _latest_qianchuan_materials(db: Session) -> list[QianchuanMaterialPerformance]:
    latest_by_material: dict[str, QianchuanMaterialPerformance] = {}
    rows = db.query(QianchuanMaterialPerformance).order_by(
        QianchuanMaterialPerformance.id.desc()
    ).all()
    for row in rows:
        if row.material_id not in latest_by_material:
            latest_by_material[row.material_id] = row
    return list(latest_by_material.values())


def _plan_qianchuan_auto_match(
    db: Session,
    min_score: int = QIANCHUAN_STRUCTURED_AUTO_BIND_SCORE,
    min_margin: int = QIANCHUAN_STRUCTURED_AUTO_BIND_MARGIN,
) -> dict:
    """Compatibility seam for the extracted Qianchuan matching planner."""
    from services.qianchuan_matching import plan_auto_match

    return plan_auto_match(
        db,
        viral_script_model=ViralScript,
        binding_model=QianchuanScriptBinding,
        latest_materials=_latest_qianchuan_materials,
        structured_candidate=_structured_qianchuan_candidate,
        min_score=min_score,
        min_margin=min_margin,
    )


def _apply_qianchuan_auto_match(db: Session, plan: dict) -> int:
    created = 0
    for item in plan.get("matches", []):
        script_id = int(item.get("script_id") or 0)
        material_id = str(item.get("material_id") or "").strip()
        if not script_id or not material_id:
            continue
        exists = db.query(QianchuanScriptBinding).filter(
            QianchuanScriptBinding.script_id == script_id,
            QianchuanScriptBinding.material_id == material_id,
        ).first()
        if exists:
            continue
        db.add(QianchuanScriptBinding(
            script_id=script_id,
            material_id=material_id,
            material_name=item.get("material_name") or "",
        ))
        created += 1
    if created:
        db.commit()
        for script_id in {int(item.get("script_id") or 0) for item in plan.get("matches", [])}:
            script = db.query(ViralScript).filter(ViralScript.id == script_id).first()
            if script:
                _sync_qianchuan_high_conversion(script, db)
    return created


def _auto_bind_qianchuan_materials(
    script: ViralScript,
    db: Session,
    min_score: int = QIANCHUAN_STRUCTURED_AUTO_BIND_SCORE,
) -> int:
    bindings = db.query(QianchuanScriptBinding.material_id).filter(
        QianchuanScriptBinding.script_id == script.id
    ).all()
    bound_ids = {material_id for (material_id,) in bindings}
    globally_bound_ids = {
        material_id for (material_id,) in db.query(QianchuanScriptBinding.material_id).filter(
            QianchuanScriptBinding.script_id != script.id
        ).all()
    }
    created = 0
    for item in _candidate_materials(script, db, bound_ids | globally_bound_ids, limit=50):
        if int(item.get("score") or 0) < min_score:
            continue
        if not item.get("matched_aliases"):
            continue
        material_id = str(item.get("material_id") or "").strip()
        if not material_id or material_id in bound_ids:
            continue
        db.add(QianchuanScriptBinding(
            script_id=script.id,
            material_id=material_id,
            material_name=item.get("material_name") or "",
        ))
        bound_ids.add(material_id)
        created += 1
    if created:
        db.commit()
        db.refresh(script)
    return created


def _qianchuan_performance_payload(script: ViralScript, db: Session) -> dict:
    bindings = db.query(QianchuanScriptBinding).filter(
        QianchuanScriptBinding.script_id == script.id
    ).order_by(QianchuanScriptBinding.id.desc()).all()
    bound_ids = {binding.material_id for binding in bindings}
    materials = []
    binding_items = []
    for binding in bindings:
        bound_materials = db.query(QianchuanMaterialPerformance).filter(
            QianchuanMaterialPerformance.material_id == binding.material_id
        ).order_by(QianchuanMaterialPerformance.id.asc()).all()
        materials.extend(bound_materials)
        binding_items.append(_qianchuan_binding_to_dict(binding, bound_materials))
    globally_bound_ids = {
        material_id for (material_id,) in db.query(QianchuanScriptBinding.material_id).filter(
            QianchuanScriptBinding.script_id != script.id
        ).all()
    }
    return {
        "script_id": script.id,
        "is_high_conversion": bool(script.is_high_conversion),
        "summary": _summarize_qianchuan_materials(materials),
        "bindings": binding_items,
        "candidates": _candidate_materials(script, db, bound_ids | globally_bound_ids),
    }


def _sync_qianchuan_high_conversion(script: ViralScript, db: Session) -> bool:
    payload = _qianchuan_performance_payload(script, db)
    changed = _apply_qianchuan_high_conversion_state(
        script,
        float(payload["summary"]["transaction_amount"] or 0),
    )
    if changed:
        db.commit()
        db.refresh(script)
    return changed


def _apply_qianchuan_high_conversion_state(script: ViralScript, transaction_amount: float) -> bool:
    should_mark = transaction_amount > 2000
    performance_data = dict(script.performance_data or {}) if isinstance(script.performance_data, dict) else {}
    was_auto_marked = bool(performance_data.get(QIANCHUAN_AUTO_HIGH_FLAG))
    changed = False
    if should_mark:
        if not script.is_high_conversion:
            script.is_high_conversion = 1
            changed = True
        if not was_auto_marked:
            performance_data[QIANCHUAN_AUTO_HIGH_FLAG] = True
            script.performance_data = performance_data
            changed = True
    elif was_auto_marked:
        if script.is_high_conversion:
            script.is_high_conversion = 0
            changed = True
        performance_data.pop(QIANCHUAN_AUTO_HIGH_FLAG, None)
        script.performance_data = performance_data
        changed = True
    return changed


def _sync_qianchuan_high_conversion_for_scripts(script_ids: set[int], db: Session) -> int:
    if not script_ids:
        return 0

    amount_by_script: dict[int, float] = {script_id: 0.0 for script_id in script_ids}
    rows = db.query(
        QianchuanScriptBinding.script_id,
        QianchuanMaterialPerformance.transaction_amount,
    ).join(
        QianchuanMaterialPerformance,
        QianchuanMaterialPerformance.material_id == QianchuanScriptBinding.material_id,
    ).filter(
        QianchuanScriptBinding.script_id.in_(script_ids)
    ).all()
    for script_id, amount in rows:
        amount_by_script[int(script_id)] = amount_by_script.get(int(script_id), 0.0) + float(amount or 0)

    changed = 0
    scripts = db.query(ViralScript).filter(ViralScript.id.in_(script_ids)).all()
    for script in scripts:
        if _apply_qianchuan_high_conversion_state(script, amount_by_script.get(int(script.id), 0.0)):
            changed += 1
    if changed:
        db.commit()
    return changed


def _qianchuan_auto_match_snapshot() -> dict:
    with _qianchuan_auto_match_lock:
        return dict(_qianchuan_auto_match_state)


def _execute_qianchuan_auto_match(
    payload: dict | None,
    db: Session,
    *,
    background_job_id: int | None = None,
):
    """Plan or apply full Qianchuan material bindings from imported performance rows."""
    payload = payload or {}
    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="参数格式错误")
    apply_requested = bool(payload.get("apply"))
    if payload.get("dry_run") is False:
        apply_requested = True
    mode = str(payload.get("mode") or ("apply" if apply_requested else "dry_run")).strip().lower()
    if mode not in {"dry_run", "apply"}:
        raise HTTPException(status_code=400, detail="mode 仅支持 dry_run 或 apply")
    try:
        min_score = int(payload.get("min_auto_score") or payload.get("min_score") or QIANCHUAN_STRUCTURED_AUTO_BIND_SCORE)
        min_margin = int(payload.get("min_margin") or QIANCHUAN_STRUCTURED_AUTO_BIND_MARGIN)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="min_score/min_margin 必须是数字")

    started_at = datetime.now().replace(microsecond=0).isoformat()
    with _qianchuan_auto_match_lock:
        if _qianchuan_auto_match_state["is_running"]:
            return ApiResponse(message="千川全量匹配正在运行", data=_qianchuan_auto_match_snapshot())
        _qianchuan_auto_match_state.update({
            "is_running": True,
            "started_at": started_at,
            "finished_at": None,
            "message": "匹配中",
        })

    from services.job_runs import finish_job, start_job
    job_id = background_job_id or start_job(
        "qianchuan_match",
        message="千川全量匹配中",
        details={"mode": mode, "min_score": min_score, "min_margin": min_margin},
        db=db,
    )

    try:
        plan = _plan_qianchuan_auto_match(db, min_score=min_score, min_margin=min_margin)
        created = _apply_qianchuan_auto_match(db, plan) if mode == "apply" else 0
        result = {
            "mode": mode,
            "min_score": min_score,
            "min_margin": min_margin,
            "script_count": db.query(ViralScript).count(),
            **plan,
            "created": created,
            "created_bindings": created,
            "would_create": 0 if mode == "apply" else plan.get("planned", 0),
        }
        finished_at = datetime.now().replace(microsecond=0).isoformat()
        with _qianchuan_auto_match_lock:
            _qianchuan_auto_match_state.update({
                "is_running": False,
                "finished_at": finished_at,
                "last_result": result,
                "message": "匹配完成",
                "mode": mode,
                "total_scripts": result.get("total_scripts"),
                "processed": result.get("processed"),
                "material_count": result.get("material_count"),
                "planned": result.get("planned"),
                "would_create": result.get("would_create"),
                "created": result.get("created"),
                "created_bindings": result.get("created_bindings"),
                "skipped_existing": result.get("skipped_existing"),
                "already_bound": result.get("already_bound"),
                "no_candidate": result.get("no_candidate"),
                "review_count": result.get("review_count"),
                "ambiguous_count": result.get("ambiguous_count"),
            })
        finish_job(job_id, status="succeeded", message="千川全量匹配完成", details=result, db=db)
        return ApiResponse(message="千川绑定自动匹配完成", data=result)
    except Exception as exc:
        _update = {
            "is_running": False,
            "finished_at": datetime.now().replace(microsecond=0).isoformat(),
            "message": "匹配失败",
        }
        with _qianchuan_auto_match_lock:
            _qianchuan_auto_match_state.update(_update)
        finish_job(job_id, status="failed", message="千川全量匹配失败", error_summary=str(exc), db=db)
        raise


def auto_match_qianchuan_bindings(payload: dict | None = None, db: Session = Depends(get_db)):
    """Plan or apply full Qianchuan material bindings from imported performance rows."""
    return _execute_qianchuan_auto_match(payload, db)


def qianchuan_auto_match_status(db: Session = Depends(get_db)):
    """Return the latest full Qianchuan auto-match status."""
    snapshot = _qianchuan_auto_match_snapshot()
    from services.job_runs import latest_job
    snapshot["job_run"] = latest_job("qianchuan_match", db=db)
    return ApiResponse(message="ok", data=snapshot)


def rematch_workbook_qianchuan_bindings(payload: dict | None = None, db: Session = Depends(get_db)):
    """Rebuild Qianchuan bindings for Excel-imported scripts using next-day material dates."""
    payload = payload or {}
    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="参数格式错误")
    apply_requested = bool(payload.get("apply"))
    if payload.get("dry_run") is False:
        apply_requested = True
    mode = str(payload.get("mode") or ("apply" if apply_requested else "dry_run")).strip().lower()
    if mode not in {"dry_run", "apply"}:
        raise HTTPException(status_code=400, detail="mode 仅支持 dry_run 或 apply")
    try:
        date_offset_days = int(payload.get("date_offset_days", 1))
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="date_offset_days 必须是数字")
    if date_offset_days < 0 or date_offset_days > 31:
        raise HTTPException(status_code=400, detail="date_offset_days 必须在 0 到 31 之间")
    workbook_sha256 = str(payload.get("workbook_sha256") or "").strip()

    plan = _plan_workbook_qianchuan_rematch(
        db,
        workbook_sha256=workbook_sha256,
        date_offset_days=date_offset_days,
    )
    cleared = 0
    created = 0
    if mode == "apply":
        cleared, created = _apply_workbook_qianchuan_rematch(db, plan)
    result = {
        "mode": mode,
        **plan,
        "cleared_bindings": cleared if mode == "apply" else plan.get("cleared_bindings", 0),
        "created_bindings": created,
        "would_create": 0 if mode == "apply" else plan.get("auto_bindings", 0),
    }
    return ApiResponse(message="Excel 脚本千川绑定重匹配完成", data=result)


async def import_qianchuan_performance(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Import Qianchuan material-performance rows and refresh bound scripts."""
    filename = file.filename or "qianchuan.xlsx"
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 文件")

    content = await read_upload_bytes(file, max_bytes=MAX_UPLOAD_SIZE)
    file_sha256 = hashlib.sha256(content).hexdigest()
    existing = db.query(QianchuanImportBatch).filter(
        QianchuanImportBatch.file_sha256 == file_sha256
    ).first()
    if existing:
        return ApiResponse(message="该千川数据表已导入，已跳过重复同步", data={
            "duplicate_file": True,
            "total_rows": existing.row_count,
            "imported": 0,
            "skipped": existing.row_count,
            "amount_field": existing.amount_field,
            "auto_bound": 0,
            "auto_marked_high": 0,
        })

    try:
        from services.bounded_executor import WorkQueueFull, run_blocking
        parsed = await run_blocking(parse_qianchuan_workbook, content, filename)
    except WorkQueueFull as exc:
        raise HTTPException(status_code=503, detail="文件解析任务繁忙，请稍后重试") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    batch = QianchuanImportBatch(
        filename=filename,
        file_sha256=file_sha256,
        row_count=parsed.row_count,
        imported_count=0,
        skipped_count=0,
        amount_field=parsed.amount_field,
    )
    db.add(batch)
    db.flush()

    imported = 0
    skipped = 0
    imported_material_ids: set[str] = set()
    seen_in_batch: set[str] = set()
    for row in parsed.rows:
        if row.material_id in seen_in_batch:
            skipped += 1
            continue
        seen_in_batch.add(row.material_id)
        db.add(QianchuanMaterialPerformance(
            batch_id=batch.id,
            material_id=row.material_id,
            material_name=row.material_name,
            material_evaluation=row.material_evaluation,
            material_duration=row.material_duration,
            material_created_time=row.material_created_time,
            material_source=row.material_source,
            tags=row.tags,
            amount_field=row.amount_field,
            transaction_amount=row.transaction_amount,
            order_count=row.order_count,
            user_pay_amount=row.user_pay_amount,
            roi=row.roi,
            impressions=row.impressions,
            ctr=row.ctr,
            spend=row.spend,
            clicks=row.clicks,
            cvr=row.cvr,
            play_3s_rate=row.play_3s_rate,
            play_10s_rate=row.play_10s_rate,
            avg_watch_seconds=row.avg_watch_seconds,
            completion_rate=row.completion_rate,
            plan_count=row.plan_count,
            product_count=row.product_count,
            raw_data=row.raw_data,
        ))
        imported += 1
        imported_material_ids.add(row.material_id)

    batch.imported_count = imported
    batch.skipped_count = skipped
    db.commit()

    auto_bound = 0
    auto_marked = 0
    if imported_material_ids:
        for script in db.query(ViralScript).all():
            auto_bound += _auto_bind_qianchuan_materials(script, db)
            if _sync_qianchuan_high_conversion(script, db):
                auto_marked += 1

    return ApiResponse(message=f"已导入 {imported} 条千川素材数据", data={
        "duplicate_file": False,
        "total_rows": parsed.row_count,
        "imported": imported,
        "skipped": skipped,
        "amount_field": parsed.amount_field,
        "auto_bound": auto_bound,
        "auto_marked_high": auto_marked,
    })


def get_viral_script_performance(script_id: int, db: Session = Depends(get_db)):
    script = db.query(ViralScript).filter(ViralScript.id == script_id).first()
    if not script:
        raise HTTPException(status_code=404, detail="爆款脚本不存在")
    return ApiResponse(message="ok", data=_qianchuan_performance_payload(script, db))


def bind_viral_script_performance(script_id: int, payload: dict, db: Session = Depends(get_db)):
    script = db.query(ViralScript).filter(ViralScript.id == script_id).first()
    if not script:
        raise HTTPException(status_code=404, detail="爆款脚本不存在")
    material_id = str((payload or {}).get("material_id") or "").strip()
    if not material_id:
        raise HTTPException(status_code=400, detail="请选择素材")
    material = db.query(QianchuanMaterialPerformance).filter(
        QianchuanMaterialPerformance.material_id == material_id
    ).order_by(QianchuanMaterialPerformance.id.desc()).first()
    if not material:
        raise HTTPException(status_code=404, detail="素材不存在")
    binding = db.query(QianchuanScriptBinding).filter(
        QianchuanScriptBinding.script_id == script.id,
        QianchuanScriptBinding.material_id == material_id,
    ).first()
    if not binding:
        binding = QianchuanScriptBinding(
            script_id=script.id,
            material_id=material.material_id,
            material_name=material.material_name,
        )
        db.add(binding)
        db.commit()
        db.refresh(binding)
    _sync_qianchuan_high_conversion(script, db)
    payload = _qianchuan_performance_payload(script, db)
    binding_payload = next((item for item in payload["bindings"] if item["id"] == binding.id), None)
    return ApiResponse(message="已绑定素材表现", data={
        "binding": binding_payload,
        "summary": payload["summary"],
        "is_high_conversion": bool(script.is_high_conversion),
    })


def unbind_viral_script_performance(script_id: int, binding_id: int, db: Session = Depends(get_db)):
    script = db.query(ViralScript).filter(ViralScript.id == script_id).first()
    if not script:
        raise HTTPException(status_code=404, detail="爆款脚本不存在")
    binding = db.query(QianchuanScriptBinding).filter(
        QianchuanScriptBinding.id == binding_id,
        QianchuanScriptBinding.script_id == script_id,
    ).first()
    if not binding:
        raise HTTPException(status_code=404, detail="绑定不存在")
    db.delete(binding)
    db.commit()
    _sync_qianchuan_high_conversion(script, db)
    return ApiResponse(message="已解绑素材表现", data=_qianchuan_performance_payload(script, db))


async def import_viral_workbook(
    request: Request,
    file: UploadFile = File(...),
):
    """Import Facai Excel scripts and cake reference images into the viral script library."""
    filename = (file.filename or "").replace("\\", "/").rsplit("/", 1)[-1]
    if not filename:
        raise HTTPException(status_code=400, detail="请选择 Excel 文件")
    suffix = Path(filename).suffix.lower()
    if suffix != ".xlsx":
        raise HTTPException(status_code=400, detail="仅支持 .xlsx Excel 文件")

    with _workbook_import_lock:
        if _workbook_import_state["is_running"]:
            return ApiResponse(message="Excel 脚本导入正在运行", data=_workbook_import_snapshot())

    content = await read_upload_bytes(file, max_bytes=SCRIPT_WORKBOOK_IMPORT_MAX_SIZE)
    try:
        from services.bounded_executor import WorkQueueFull, run_blocking
        from services.upload_validation import (
            LARGE_WORKBOOK_POLICY,
            UploadValidationError,
            validate_upload,
        )
        await run_blocking(validate_upload, filename, content, LARGE_WORKBOOK_POLICY)
    except WorkQueueFull as exc:
        raise HTTPException(status_code=503, detail="文件解析任务繁忙，请稍后重试") from exc
    except UploadValidationError as exc:
        raise HTTPException(status_code=exc.status_code, detail=str(exc)) from exc
    workbook_sha256 = hashlib.sha256(content).hexdigest()
    WORKBOOK_IMPORT_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    upload_path = WORKBOOK_IMPORT_UPLOAD_DIR / f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{workbook_sha256[:12]}{suffix}"
    upload_path.write_bytes(content)

    with _workbook_import_lock:
        if _workbook_import_state["is_running"]:
            upload_path.unlink(missing_ok=True)
            return ApiResponse(message="Excel 脚本导入正在运行", data=_workbook_import_snapshot())
        from services.job_runs import start_job
        client_id = request.headers.get("X-Facai-Client-Id")
        job_id = start_job(
            "workbook_import",
            message="Excel 脚本导入启动中",
            details={"filename": filename},
            owner_key=owner_key_for_request(request, client_id) if client_id else None,
            origin_path="/app/templates",
        )
        _workbook_import_state.update({
            "is_running": True,
            "filename": filename,
            "total": 0,
            "processed": 0,
            "created": 0,
            "updated": 0,
            "skipped": 0,
            "image_count": 0,
            "index_error_count": 0,
            "error_count": 0,
            "errors": [],
            "ids": [],
            "started_at": datetime.now().replace(microsecond=0).isoformat(),
            "finished_at": None,
            "message": "Excel 脚本导入启动中",
            "job_id": job_id,
        })

    from services.task_queue import enqueue_task, task_worker_status
    task_payload = {
        "workbook_path": str(upload_path),
        "filename": filename,
        "workbook_sha256": workbook_sha256,
        "job_id": job_id,
    }
    if task_worker_status()["alive"]:
        enqueue_task("workbook_import", task_payload, max_attempts=3, job_run_id=job_id)
    else:
        # Router-only test/dev apps do not run the main lifespan worker.
        threading.Thread(
            target=_run_workbook_import_task,
            args=(task_payload,),
            name="facai-script-workbook-import-fallback",
            daemon=True,
        ).start()
    return ApiResponse(message="Excel 脚本导入已启动", data=_workbook_import_snapshot())


def workbook_import_status():
    """Return the latest Excel script workbook import status."""
    return ApiResponse(message="ok", data=_workbook_import_snapshot())


def get_viral_script_cake_image(script_id: int, image_name: str, db: Session = Depends(get_db)):
    """Serve a cake reference image owned by a viral script."""
    script = db.query(ViralScript).filter(ViralScript.id == script_id).first()
    if not script:
        raise HTTPException(status_code=404, detail="爆款脚本不存在")
    images = _existing_cake_images(script)
    image_meta = next((item for item in images if isinstance(item, dict) and item.get("filename") == image_name), None)
    if not image_meta:
        raise HTTPException(status_code=404, detail="图片不存在")
    relative_path = image_meta.get("relative_path")
    if not relative_path:
        raise HTTPException(status_code=404, detail="图片不存在")

    base_dir = Path(VIRAL_SCRIPT_IMAGE_DIR).resolve()
    target = (base_dir / relative_path).resolve()
    if base_dir != target and base_dir not in target.parents:
        raise HTTPException(status_code=404, detail="图片不存在")
    if not target.is_file():
        raise HTTPException(status_code=404, detail="图片不存在")
    media_type = image_meta.get("content_type") or mimetypes.guess_type(str(target))[0] or "application/octet-stream"
    return FileResponse(str(target), media_type=media_type, filename=target.name)


def get_viral_script(script_id: int, db: Session = Depends(get_db)):
    """获取爆款脚本详情"""
    script = db.query(ViralScript).filter(ViralScript.id == script_id).first()
    if not script:
        raise HTTPException(status_code=404, detail="爆款脚本不存在")
    return script


def toggle_high_viral(script_id: int, db: Session = Depends(get_db)):
    """切换爆款脚本的高成交标记"""
    script = db.query(ViralScript).filter(ViralScript.id == script_id).first()
    if not script:
        raise HTTPException(status_code=404, detail="爆款脚本不存在")
    script.is_high_conversion = 1 if script.is_high_conversion == 0 else 0
    if isinstance(script.performance_data, dict) and script.performance_data.get(QIANCHUAN_AUTO_HIGH_FLAG):
        data = dict(script.performance_data)
        data.pop(QIANCHUAN_AUTO_HIGH_FLAG, None)
        script.performance_data = data
    from services.vector_sync import enqueue_vector_sync
    enqueue_vector_sync(db, "viral_script", script.id, "upsert")
    db.commit()
    index_sync_status = "synced" if _sync_viral_index(script, db) is True else "pending"
    return ApiResponse(
        message=f"已{'标记为' if script.is_high_conversion else '取消'}高成交",
        data={"is_high": bool(script.is_high_conversion), "index_sync_status": index_sync_status},
    )


def delete_viral_script(script_id: int, db: Session = Depends(get_db)):
    """删除爆款脚本"""
    script = db.query(ViralScript).filter(ViralScript.id == script_id).first()
    if not script:
        raise HTTPException(status_code=404, detail="爆款脚本不存在")
    from services.vector_sync import enqueue_vector_sync
    enqueue_vector_sync(db, "viral_script", script.id, "delete")
    db.delete(script)
    db.commit()
    index_sync_status = _delete_script_index(f"viral_{script.id}")
    return ApiResponse(message="已删除", data={"index_sync_status": index_sync_status})


async def upload_viral_script(
    title: str = Form(""),
    category: str = Form(""),
    video_type: str = Form(""),
    tags: str = Form(""),
    product_name: str = Form(""),
    content: str = Form(...),
    db: Session = Depends(get_db),
):
    """上传法采脚本，产品名匹配品类+AI分析"""
    script_text = format_script(content)
    if len(script_text) < 20:
        raise HTTPException(status_code=400, detail="脚本内容太短")

    # 通过产品名直接匹配数据库中产品获取品类
    auto_category = category or ""
    if product_name and not auto_category:
        from models import Product
        matched = db.query(Product).filter(Product.name == product_name).first()
        if matched:
            auto_category = matched.category
    if not auto_category:
        auto_category = "烘焙配件"

    ai_analysis = await analyze_script_ai(script_text, video_type)

    viral = ViralScript(
        category=auto_category,
        video_type=ai_analysis.get("video_type") or video_type or "机制类",
        title=title or (product_name + "脚本" if product_name else "未命名脚本"),
        script_content=script_text,
        tags=_combine_tags(_combine_tags(tags or "", product_name), ai_analysis.get("tags", "")),
        performance_data={
            "source": "手动上传",
            "product_name": product_name or "",
            "ai_structure": ai_analysis.get("structure", ""),
            "ai_viral_points": ai_analysis.get("viral_points", ""),
        },
    )
    db.add(viral)
    db.flush()
    from services.vector_sync import enqueue_vector_sync
    enqueue_vector_sync(db, "viral_script", viral.id, "upsert")
    db.commit()
    db.refresh(viral)
    index_sync_status = "synced" if _sync_viral_index(viral, db) is True else "pending"
    return ApiResponse(
        message=f"已上传并AI分析，归类为「{auto_category}」",
        data={"id": viral.id, "category": auto_category, "analysis": ai_analysis, "index_sync_status": index_sync_status},
    )


async def upload_viral_txt_batch(
    files: List[UploadFile] = File(...),
    category: str = Form("", max_length=100),
    video_type: str = Form("", max_length=100),
    tags: str = Form("", max_length=1000),
    product_name: str = Form("", max_length=200),
    db: Session = Depends(get_db),
):
    """Batch upload local .txt scripts into the viral script library."""
    result = {"total": len(files), "success": 0, "skipped": 0, "errors": [], "ids": []}
    if not files:
        raise HTTPException(status_code=400, detail="请选择 txt 文件")
    if len(files) > MAX_TXT_BATCH_FILES:
        raise HTTPException(
            status_code=413,
            detail=f"单次最多上传 {MAX_TXT_BATCH_FILES} 个 txt 文件",
        )

    auto_category = category or ""
    if product_name and not auto_category:
        from models import Product
        matched = db.query(Product).filter(Product.name == product_name).first()
        if matched:
            auto_category = matched.category
    if not auto_category:
        auto_category = "烘焙配件"

    for upload in files:
        filename = upload.filename or ""
        safe_name = filename.replace("\\", "/").rsplit("/", 1)[-1] or "未命名.txt"
        if not safe_name.lower().endswith(".txt"):
            result["skipped"] += 1
            result["errors"].append(f"{safe_name}: 仅支持 .txt 文件")
            continue

        try:
            content = await read_upload_bytes(upload, max_bytes=MAX_UPLOAD_SIZE)
            script_text = format_script(_decode_txt(content))
            if len(script_text) < 20:
                raise ValueError("脚本内容太短")

            ai_analysis = await analyze_script_ai(script_text, video_type)
            viral = ViralScript(
                category=auto_category,
                video_type=ai_analysis.get("video_type") or video_type or "机制类",
                title=_title_from_filename(safe_name),
                script_content=script_text,
                tags=_combine_tags(_combine_tags(tags or "", product_name), ai_analysis.get("tags", "")),
                performance_data={
                    "source": "批量TXT上传",
                    "product_name": product_name or "",
                    "filename": safe_name,
                    "ai_structure": ai_analysis.get("structure", ""),
                    "ai_viral_points": ai_analysis.get("viral_points", ""),
                },
            )
            db.add(viral)
            db.flush()
            from services.vector_sync import enqueue_vector_sync
            enqueue_vector_sync(db, "viral_script", viral.id, "upsert")
            db.commit()
            db.refresh(viral)
            _sync_viral_index(viral, db)
            result["success"] += 1
            result["ids"].append(viral.id)
        except HTTPException:
            raise
        except Exception as exc:
            db.rollback()
            result["skipped"] += 1
            result["errors"].append(f"{safe_name}: {exc}")

    return ApiResponse(message=f"已导入 {result['success']} 个 txt 脚本", data=result)


async def categorize_product_ai(name: str) -> str:
    if not ai_service.is_available: return ""
    try:
        r = await ai_service.chat([{"role":"system","content":"归类到：烘焙调色/烘焙装饰/烘焙调味/烘焙夹心/烘焙配件。只输出品类名。"},{"role":"user","content":name}], temperature=0.1, interface_key="viral_script_analyze")
        for c in ["烘焙调色","烘焙装饰","烘焙调味","烘焙夹心","烘焙配件"]:
            if c in r: return c
    except: pass
    return ""


# ========== 索引同步辅助 ==========

from models import ReferenceScript


def _sync_viral_index(viral, db: Session):
    from services.vector_sync import ensure_and_process_vector_sync

    status = ensure_and_process_vector_sync(db, "viral_script", viral.id, "upsert")
    return status == "succeeded"


def _sync_reference_index(ref, db: Session):
    from services.vector_sync import ensure_and_process_vector_sync

    status = ensure_and_process_vector_sync(db, "reference_script", ref.id, "upsert")
    return status == "succeeded"


def _delete_script_index(doc_id: str):
    from database import SessionLocal
    from services.vector_sync import ensure_and_process_vector_sync

    entity_type = "viral_script" if doc_id.startswith("viral_") else "reference_script"
    entity_id = int(doc_id.rsplit("_", 1)[-1])
    db = SessionLocal()
    try:
        status = ensure_and_process_vector_sync(db, entity_type, entity_id, "delete")
        return "synced" if status == "succeeded" else "pending"
    finally:
        db.close()


# Composite compatibility router. Route declarations live in focused domain
# modules, while existing callers continue importing routers.templates.router.
from routers import template_local_scan as _local_scan_routes
from routers import template_qianchuan as _qianchuan_routes
from routers import template_script_library as _script_library_routes
from routers import template_workbook_import as _workbook_routes

router.include_router(_workbook_routes.router)
router.include_router(_local_scan_routes.router)
router.include_router(_qianchuan_routes.router)
router.include_router(_script_library_routes.router)
