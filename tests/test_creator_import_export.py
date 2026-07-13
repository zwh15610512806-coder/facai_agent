import os
import tempfile
import unittest
import zipfile
from datetime import UTC, datetime, timedelta
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

import creator_models  # noqa: F401
import models  # noqa: F401
from database import Base
from models import Product
from creator_models import CreatorCollaboration, CreatorImportBatch
from routers import creators as creators_router
from services import creator_importer


def workbook_bytes(headers, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


class CreatorImportExportTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine(
            "sqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(bind=self.engine)
        self.Session = sessionmaker(bind=self.engine)
        self.db = self.Session()
        self.tmp = tempfile.TemporaryDirectory()
        self.original_import_dir = creator_importer.CREATOR_IMPORT_DIR
        creator_importer.CREATOR_IMPORT_DIR = Path(self.tmp.name)

        app = FastAPI()

        def override_db():
            yield self.db

        app.dependency_overrides[creators_router.get_db] = override_db
        app.include_router(creators_router.router, prefix="/api/creators")
        self.client = TestClient(app)

        self.product = Product(name="法采草莓果酱", category="果酱", price=59, status="active")
        self.db.add(self.product)
        self.db.commit()
        self.member = self.client.post("/api/creators/bd-members", json={"name": "小王"}).json()

    def tearDown(self):
        creator_importer.CREATOR_IMPORT_DIR = self.original_import_dir
        self.tmp.cleanup()
        self.db.close()
        Base.metadata.drop_all(bind=self.engine)
        self.engine.dispose()

    def _preview(self, kind, content, filename="import.xlsx", source_type="facai_template"):
        return self.client.post(
            "/api/creators/import/preview",
            data={"kind": kind, "source_type": source_type},
            files={
                "file": (
                    filename,
                    content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    def test_standard_templates_are_valid_xlsx_with_expected_columns(self):
        creators = self.client.get("/api/creators/import/templates/creators")
        collaborations = self.client.get("/api/creators/import/templates/collaborations")
        self.assertEqual(200, creators.status_code, creators.text)
        self.assertEqual(200, collaborations.status_code, collaborations.text)
        self.assertIn("no-store", creators.headers["cache-control"])

        creator_headers = [cell.value for cell in load_workbook(BytesIO(creators.content)).active[1]]
        collaboration_headers = [cell.value for cell in load_workbook(BytesIO(collaborations.content)).active[1]]
        self.assertIn("抖音号", creator_headers)
        self.assertIn("达人昵称", creator_headers)
        self.assertIn("合作编号", collaboration_headers)
        self.assertIn("实际支付金额（元）", collaboration_headers)

    def test_creator_preview_validate_commit_skips_bad_rows_and_blocks_same_committed_sha(self):
        content = workbook_bytes(
            ["达人昵称", "抖音号", "负责人", "粉丝数", "主营垂类", "联系人", "手机号"],
            [
                ["烘焙小麦", "@cake-wheat", "小王", 128000, "家庭烘焙|教程", "张麦", "13812345678"],
                ["缺身份达人", "", "小王", 1000, "烘焙", "李四", "13900000000"],
            ],
        )
        preview = self._preview("creators", content)
        self.assertEqual(200, preview.status_code, preview.text)
        preview_data = preview.json()
        self.assertEqual(2, preview_data["row_count"])
        self.assertEqual("nickname", preview_data["suggested_mapping"]["达人昵称"])
        self.assertEqual("douyin_handle", preview_data["suggested_mapping"]["抖音号"])

        token = preview_data["token"]
        validated = self.client.post(
            f"/api/creators/import/{token}/validate",
            json={"mapping": preview_data["suggested_mapping"]},
        )
        self.assertEqual(200, validated.status_code, validated.text)
        self.assertEqual(1, validated.json()["error_count"])
        self.assertEqual(1, validated.json()["imported_count"])

        committed = self.client.post(f"/api/creators/import/{token}/commit")
        self.assertEqual(200, committed.status_code, committed.text)
        self.assertEqual(1, committed.json()["imported_count"])
        self.assertEqual(1, committed.json()["skipped_count"])
        creators = self.client.get("/api/creators").json()
        self.assertEqual(["烘焙小麦"], [item["nickname"] for item in creators["items"]])

        errors = self.client.get(f"/api/creators/import/{token}/errors")
        self.assertEqual(200, errors.status_code)
        error_text = "|".join(
            str(cell.value or "")
            for row in load_workbook(BytesIO(errors.content)).active.iter_rows()
            for cell in row
        )
        self.assertIn("第3行", error_text)
        self.assertNotIn("13900000000", error_text)

        duplicate = self._preview("creators", content)
        self.assertEqual(409, duplicate.status_code)

        batch = self.db.query(CreatorImportBatch).filter(CreatorImportBatch.token == token).one()
        batch.created_at = datetime.now(UTC).replace(tzinfo=None) - timedelta(hours=25)
        self.db.commit()
        self.assertEqual(200, self.client.get(f"/api/creators/import/{token}/errors").status_code)
        self.db.refresh(batch)
        self.assertEqual("committed", batch.status)
        self.assertEqual(409, self._preview("creators", content).status_code)

    def test_manual_mapping_and_xls_rejection(self):
        content = workbook_bytes(["显示名", "账号"], [["映射达人", "mapped-id"]])
        preview = self._preview("creators", content, source_type="douyin_official")
        self.assertEqual(200, preview.status_code, preview.text)
        token = preview.json()["token"]
        validated = self.client.post(
            f"/api/creators/import/{token}/validate",
            json={"mapping": {"显示名": "nickname", "账号": "douyin_handle"}},
        )
        self.assertEqual(0, validated.json()["error_count"])
        self.assertEqual(200, self.client.post(f"/api/creators/import/{token}/commit").status_code)

        rejected = self._preview("creators", b"legacy", filename="legacy.xls")
        self.assertEqual(400, rejected.status_code)

    def test_collaboration_import_resolves_creator_products_and_actual_paid_cents(self):
        creator = self.client.post(
            "/api/creators",
            json={"nickname": "合作达人", "douyin_handle": "coop-creator"},
        ).json()
        content = workbook_bytes(
            [
                "合作编号",
                "达人抖音号",
                "合作形式",
                "合作日期",
                "合作状态",
                "实际支付金额（元）",
                "金额状态",
                "合作产品",
            ],
            [
                ["COOP-X1", "coop-creator", "直播", "2026-07-12", "已完成", 1234.56, "已确认", "法采草莓果酱"],
                ["COOP-X2", "coop-creator", "短视频", "2026-07-13", "已完成", 100, "已确认", "未知产品"],
            ],
        )
        preview = self._preview("collaborations", content)
        token = preview.json()["token"]
        validated = self.client.post(
            f"/api/creators/import/{token}/validate",
            json={"mapping": preview.json()["suggested_mapping"]},
        )
        self.assertEqual(1, validated.json()["error_count"])
        self.assertIn("不存在", validated.json()["errors"][0]["message"])
        committed = self.client.post(f"/api/creators/import/{token}/commit")
        self.assertEqual(1, committed.json()["imported_count"])
        metrics = self.client.get(f"/api/creators/{creator['id']}").json()["metrics"]
        self.assertEqual(123456, metrics["confirmed_paid_cents"])

    def test_collaboration_without_stable_id_uses_row_fingerprint_for_dedup(self):
        creator = self.client.post(
            "/api/creators",
            json={"nickname": "无编号合作达人", "douyin_handle": "fingerprint-coop"},
        ).json()
        headers = ["达人抖音号", "合作形式", "合作日期", "实际支付金额（元）"]
        row = ["fingerprint-coop", "直播", "2026-07-13", 88.8]

        first_preview = self._preview("collaborations", workbook_bytes(headers, [row])).json()
        first_validated = self.client.post(
            f"/api/creators/import/{first_preview['token']}/validate",
            json={"mapping": first_preview["suggested_mapping"]},
        )
        self.assertEqual(0, first_validated.json()["error_count"], first_validated.text)
        first_commit = self.client.post(
            f"/api/creators/import/{first_preview['token']}/commit"
        )
        self.assertEqual(1, first_commit.json()["imported_count"])

        saved = self.client.get(
            f"/api/creators/{creator['id']}/collaborations"
        ).json()
        self.assertEqual(1, len(saved))
        self.assertTrue(saved[0]["internal_code"].startswith("IMP-"))

        second_headers = headers + ["未映射备注列"]
        second_row = row + ["仅改变文件哈希"]
        second_preview = self._preview(
            "collaborations", workbook_bytes(second_headers, [second_row])
        ).json()
        second_validated = self.client.post(
            f"/api/creators/import/{second_preview['token']}/validate",
            json={"mapping": second_preview["suggested_mapping"]},
        )
        self.assertEqual(0, second_validated.json()["error_count"], second_validated.text)
        second_commit = self.client.post(
            f"/api/creators/import/{second_preview['token']}/commit"
        )
        self.assertEqual(0, second_commit.json()["imported_count"])
        self.assertEqual(1, second_commit.json()["updated_count"])
        self.assertEqual(
            1,
            len(self.client.get(f"/api/creators/{creator['id']}/collaborations").json()),
        )

    def test_filtered_exports_mask_creator_pii_and_include_sample_shipping_snapshot(self):
        creator = self.client.post(
            "/api/creators",
            json={
                "nickname": "导出达人",
                "douyin_handle": "export-creator",
                "contact_phone": "13812345678",
                "stage": "negotiating",
            },
        ).json()
        address = self.client.post(
            f"/api/creators/{creator['id']}/addresses",
            json={
                "recipient_name": "王达人",
                "phone": "13812345678",
                "province": "广东省",
                "city": "深圳市",
                "district": "南山区",
                "detail": "科技园 8 号",
            },
        ).json()
        self.client.post(
            f"/api/creators/{creator['id']}/sample-orders",
            json={
                "idempotency_key": "export-sample-order",
                "address_id": address["id"],
                "items": [{"product_id": self.product.id, "quantity": 1}],
            },
        )

        creator_export = self.client.get(
            "/api/creators/export", params={"entity": "creators", "stage": "negotiating"}
        )
        self.assertEqual(200, creator_export.status_code, creator_export.text)
        creator_sheet = load_workbook(BytesIO(creator_export.content)).active
        creator_text = "|".join(str(cell.value or "") for row in creator_sheet.iter_rows() for cell in row)
        self.assertIn("导出达人", creator_text)
        self.assertNotIn("13812345678", creator_text)
        self.assertNotIn("科技园", creator_text)

        sample_export = self.client.get(
            "/api/creators/export", params={"entity": "sample_orders", "creator_id": creator["id"]}
        )
        self.assertEqual(200, sample_export.status_code, sample_export.text)
        self.assertEqual("no-store", sample_export.headers["cache-control"])
        sample_text = "|".join(
            str(cell.value or "")
            for row in load_workbook(BytesIO(sample_export.content)).active.iter_rows()
            for cell in row
        )
        self.assertIn("13812345678", sample_text)
        self.assertIn("科技园 8 号", sample_text)
        self.assertIn("敏感信息", sample_text)

    def test_creator_identity_conflict_is_reported_and_skipped(self):
        self.client.post(
            "/api/creators",
            json={"nickname": "UID达人", "platform_uid": "uid-a", "douyin_handle": "handle-a"},
        )
        self.client.post(
            "/api/creators",
            json={"nickname": "账号达人", "platform_uid": "uid-b", "douyin_handle": "handle-b"},
        )
        content = workbook_bytes(
            ["达人昵称", "官方达人ID", "抖音号"],
            [["冲突达人", "uid-a", "handle-b"]],
        )
        preview = self._preview("creators", content)
        validated = self.client.post(
            f"/api/creators/import/{preview.json()['token']}/validate",
            json={"mapping": preview.json()["suggested_mapping"]},
        )
        self.assertEqual(1, validated.json()["error_count"])
        self.assertIn("身份冲突", validated.json()["errors"][0]["message"])

    def test_partial_import_updates_preserve_omitted_creator_and_collaboration_fields(self):
        creator = self.client.post(
            "/api/creators",
            json={
                "nickname": "保留字段达人",
                "douyin_handle": "preserve-import",
                "stage": "negotiating",
                "tags": ["高意向"],
                "contact_phone": "13812345678",
            },
        ).json()
        collaboration = self.client.post(
            f"/api/creators/{creator['id']}/collaborations",
            json={
                "source_type": "manual",
                "external_record_id": "EXT-PRESERVE-1",
                "internal_code": "COOP-PRESERVE-1",
                "collaboration_type": "live",
                "collaboration_date": "2026-07-12",
                "status": "completed",
                "actual_paid_cents": 123456,
                "amount_status": "confirmed",
                "notes": "原合作备注",
                "products": [{"product_id": self.product.id}],
            },
        ).json()

        creator_preview = self._preview(
            "creators",
            workbook_bytes(
                ["抖音号", "MCN机构"],
                [["preserve-import", "新机构"]],
            ),
        ).json()
        creator_validate = self.client.post(
            f"/api/creators/import/{creator_preview['token']}/validate",
            json={"mapping": creator_preview["suggested_mapping"]},
        )
        self.assertEqual(0, creator_validate.json()["error_count"], creator_validate.text)
        self.assertEqual(
            200,
            self.client.post(
                f"/api/creators/import/{creator_preview['token']}/commit"
            ).status_code,
        )
        saved_creator = self.client.get(f"/api/creators/{creator['id']}").json()
        self.assertEqual("negotiating", saved_creator["stage"])
        self.assertEqual(["高意向"], saved_creator["tags"])
        self.assertEqual("新机构", saved_creator["mcn_name"])

        collaboration_preview = self._preview(
            "collaborations",
            workbook_bytes(
                ["合作编号", "达人抖音号", "备注"],
                [["COOP-PRESERVE-1", "preserve-import", "更新后的备注"]],
            ),
        ).json()
        collaboration_validate = self.client.post(
            f"/api/creators/import/{collaboration_preview['token']}/validate",
            json={"mapping": collaboration_preview["suggested_mapping"]},
        )
        self.assertEqual(
            0,
            collaboration_validate.json()["error_count"],
            collaboration_validate.text,
        )
        self.assertEqual(
            200,
            self.client.post(
                f"/api/creators/import/{collaboration_preview['token']}/commit"
            ).status_code,
        )
        saved = self.client.get(
            f"/api/creators/{creator['id']}/collaborations"
        ).json()[0]
        self.assertEqual(collaboration["id"], saved["id"])
        self.assertEqual("completed", saved["status"])
        self.assertEqual("confirmed", saved["amount_status"])
        self.assertEqual(123456, saved["actual_paid_cents"])
        self.assertEqual("EXT-PRESERVE-1", saved["external_record_id"])
        self.assertEqual("更新后的备注", saved["notes"])
        self.assertEqual(["法采草莓果酱"], [p["product_name_snapshot"] for p in saved["products"]])

    def test_collaboration_external_and_internal_identity_conflict_never_overwrites(self):
        creator = self.client.post(
            "/api/creators", json={"nickname": "冲突合作达人", "douyin_handle": "conflict-coop"}
        ).json()
        for code, external_id, cents in (("CODE-A", "EXT-A", 100), ("CODE-B", "EXT-B", 200)):
            response = self.client.post(
                f"/api/creators/{creator['id']}/collaborations",
                json={
                    "source_type": "douyin_official",
                    "external_record_id": external_id,
                    "internal_code": code,
                    "collaboration_type": "live",
                    "collaboration_date": "2026-07-12",
                    "actual_paid_cents": cents,
                },
            )
            self.assertEqual(201, response.status_code, response.text)
        content = workbook_bytes(
            ["合作编号", "平台记录ID", "达人抖音号", "合作形式", "合作日期", "实际支付金额（元）"],
            [["CODE-B", "EXT-A", "conflict-coop", "直播", "2026-07-13", 9999]],
        )
        preview = self._preview("collaborations", content, source_type="douyin_official")
        validated = self.client.post(
            f"/api/creators/import/{preview.json()['token']}/validate",
            json={"mapping": preview.json()["suggested_mapping"]},
        )
        self.assertEqual(1, validated.json()["error_count"])
        committed = self.client.post(f"/api/creators/import/{preview.json()['token']}/commit")
        self.assertEqual(0, committed.json()["imported_count"])
        saved = {
            item.internal_code: item.actual_paid_cents
            for item in self.db.query(CreatorCollaboration).all()
        }
        self.assertEqual({"CODE-A": 100, "CODE-B": 200}, saved)

    def test_staged_sha_is_rechecked_and_expired_tokens_remove_temp_file(self):
        original = workbook_bytes(["达人昵称", "抖音号"], [["原达人", "original"]])
        preview = self._preview("creators", original)
        token = preview.json()["token"]
        self.client.post(
            f"/api/creators/import/{token}/validate",
            json={"mapping": preview.json()["suggested_mapping"]},
        )
        staged = creator_importer.CREATOR_IMPORT_DIR / f"{token}.xlsx"
        staged.write_bytes(workbook_bytes(["达人昵称", "抖音号"], [["篡改达人", "tampered"]]))
        changed = self.client.post(f"/api/creators/import/{token}/commit")
        self.assertEqual(409, changed.status_code)
        self.assertEqual(0, self.client.get("/api/creators").json()["total"])

        expiring = self._preview(
            "creators", workbook_bytes(["达人昵称", "抖音号"], [["过期达人", "expired"]])
        )
        expired_token = expiring.json()["token"]
        batch = self.db.query(CreatorImportBatch).filter(CreatorImportBatch.token == expired_token).one()
        batch.created_at = datetime.now(UTC).replace(tzinfo=None) - timedelta(hours=25)
        self.db.commit()
        expired_path = creator_importer.CREATOR_IMPORT_DIR / f"{expired_token}.xlsx"
        response = self.client.get(f"/api/creators/import/{expired_token}/errors")
        self.assertEqual(410, response.status_code)
        self.assertFalse(expired_path.exists())

    def test_xlsx_expansion_and_row_limits_are_enforced(self):
        output = BytesIO()
        with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("huge.txt", b"x" * 101)
        original_limit = creator_importer.MAX_XLSX_UNCOMPRESSED_BYTES
        creator_importer.MAX_XLSX_UNCOMPRESSED_BYTES = 100
        try:
            response = self._preview("creators", output.getvalue())
        finally:
            creator_importer.MAX_XLSX_UNCOMPRESSED_BYTES = original_limit
        self.assertEqual(400, response.status_code)
        self.assertIn("解压", response.json()["detail"])

        original_rows = creator_importer.MAX_IMPORT_ROWS
        creator_importer.MAX_IMPORT_ROWS = 1
        try:
            rows = self._preview(
                "creators",
                workbook_bytes(["达人昵称", "抖音号"], [["甲", "a"], ["乙", "b"]]),
            )
        finally:
            creator_importer.MAX_IMPORT_ROWS = original_rows
        self.assertEqual(400, rows.status_code)
        self.assertIn("行数", rows.json()["detail"])

    def test_import_rejects_non_finite_money_values_as_row_errors(self):
        self.client.post(
            "/api/creators", json={"nickname": "数值达人", "douyin_handle": "numeric-import"}
        )
        for index, value in enumerate(("NaN", "Infinity", "-Infinity"), start=1):
            with self.subTest(value=value):
                content = workbook_bytes(
                    ["合作编号", "达人抖音号", "合作日期", "实际支付金额（元）"],
                    [[f"NUM-{index}", "numeric-import", "2026-07-13", value]],
                )
                preview = self._preview("collaborations", content)
                validated = self.client.post(
                    f"/api/creators/import/{preview.json()['token']}/validate",
                    json={"mapping": preview.json()["suggested_mapping"]},
                )
                self.assertEqual(200, validated.status_code, validated.text)
                self.assertEqual(1, validated.json()["error_count"])
                self.assertIn("有限数字", validated.json()["errors"][0]["message"])

    def test_import_rejects_fractional_integer_fields(self):
        content = workbook_bytes(
            ["达人昵称", "抖音号", "粉丝数", "匹配度"],
            [
                ["小数粉丝", "fractional-followers", 1.5, 3],
                ["小数匹配度", "fractional-fit", 100, 2.5],
            ],
        )
        preview = self._preview("creators", content)
        validated = self.client.post(
            f"/api/creators/import/{preview.json()['token']}/validate",
            json={"mapping": preview.json()["suggested_mapping"]},
        )
        self.assertEqual(200, validated.status_code, validated.text)
        self.assertEqual(2, validated.json()["error_count"])
        self.assertTrue(all("必须是整数" in item["message"] for item in validated.json()["errors"]))

    def test_import_rejects_text_over_schema_limits(self):
        creator_content = workbook_bytes(
            ["达人昵称", "抖音号", "手机号"],
            [
                ["超长达人" * 51, "long-nickname", "13800000000"],
                ["超长手机", "long-phone", "1" * 51],
                ["超长账号", "h" * 201, "13800000001"],
            ],
        )
        creator_preview = self._preview("creators", creator_content)
        creator_validated = self.client.post(
            f"/api/creators/import/{creator_preview.json()['token']}/validate",
            json={"mapping": creator_preview.json()["suggested_mapping"]},
        )
        self.assertEqual(200, creator_validated.status_code, creator_validated.text)
        self.assertEqual(3, creator_validated.json()["error_count"])
        self.assertTrue(all("长度不能超过" in item["message"] for item in creator_validated.json()["errors"]))

        self.client.post(
            "/api/creators", json={"nickname": "长度合作达人", "douyin_handle": "length-coop"}
        )
        collaboration_content = workbook_bytes(
            ["合作编号", "平台记录ID", "达人抖音号", "合作日期", "备注"],
            [
                ["C" * 101, "EXT-1", "length-coop", "2026-07-13", "正常"],
                ["LONG-EXT", "E" * 201, "length-coop", "2026-07-13", "正常"],
                ["LONG-NOTE", "EXT-3", "length-coop", "2026-07-13", "N" * 20001],
            ],
        )
        collaboration_preview = self._preview("collaborations", collaboration_content)
        collaboration_validated = self.client.post(
            f"/api/creators/import/{collaboration_preview.json()['token']}/validate",
            json={"mapping": collaboration_preview.json()["suggested_mapping"]},
        )
        self.assertEqual(200, collaboration_validated.status_code, collaboration_validated.text)
        self.assertEqual(3, collaboration_validated.json()["error_count"])
        self.assertTrue(
            all("长度不能超过" in item["message"] for item in collaboration_validated.json()["errors"])
        )

    def test_import_rejects_partially_filled_shipping_address(self):
        content = workbook_bytes(
            ["达人昵称", "抖音号", "收件人", "收件电话", "省", "市", "详细地址"],
            [["地址不完整达人", "partial-address", "张三", "13812345678", "广东省", "深圳市", ""]],
        )
        preview = self._preview("creators", content)
        validated = self.client.post(
            f"/api/creators/import/{preview.json()['token']}/validate",
            json={"mapping": preview.json()["suggested_mapping"]},
        )

        self.assertEqual(200, validated.status_code, validated.text)
        self.assertEqual(1, validated.json()["error_count"])
        self.assertIn("地址", validated.json()["errors"][0]["message"])
        committed = self.client.post(f"/api/creators/import/{preview.json()['token']}/commit")
        self.assertEqual(200, committed.status_code, committed.text)
        self.assertEqual(0, self.client.get("/api/creators", params={"search": "地址不完整"}).json()["total"])

    def test_same_file_creator_identity_duplicates_are_reported(self):
        content = workbook_bytes(
            ["达人昵称", "官方达人ID", "抖音号"],
            [
                ["首行达人", "duplicate-uid-a", "same-file-handle"],
                ["重复达人", "duplicate-uid-b", "same-file-handle"],
            ],
        )
        preview = self._preview("creators", content)
        validated = self.client.post(
            f"/api/creators/import/{preview.json()['token']}/validate",
            json={"mapping": preview.json()["suggested_mapping"]},
        )
        self.assertEqual(200, validated.status_code, validated.text)
        self.assertEqual(1, validated.json()["imported_count"])
        self.assertEqual(1, validated.json()["error_count"])
        self.assertIn("与第2行达人身份重复", validated.json()["errors"][0]["message"])

    def test_same_existing_creator_split_identity_rows_are_reported(self):
        self.client.post(
            "/api/creators",
            json={
                "nickname": "已有双身份达人",
                "platform_uid": "existing-split-uid",
                "douyin_handle": "existing-split-handle",
            },
        )
        content = workbook_bytes(
            ["达人昵称", "官方达人ID", "抖音号"],
            [
                ["按UID更新", "existing-split-uid", ""],
                ["按抖音号更新", "", "existing-split-handle"],
            ],
        )
        preview = self._preview("creators", content)
        validated = self.client.post(
            f"/api/creators/import/{preview.json()['token']}/validate",
            json={"mapping": preview.json()["suggested_mapping"]},
        )

        self.assertEqual(200, validated.status_code, validated.text)
        self.assertEqual(1, validated.json()["imported_count"])
        self.assertEqual(1, validated.json()["error_count"])
        self.assertIn("与第2行达人身份重复", validated.json()["errors"][0]["message"])

    def test_same_file_collaboration_identifiers_are_reported(self):
        self.client.post(
            "/api/creators", json={"nickname": "重复合作达人", "douyin_handle": "duplicate-coop"}
        )
        content = workbook_bytes(
            ["合作编号", "平台记录ID", "达人抖音号", "合作日期"],
            [
                ["DUP-CODE", "DUP-EXT", "duplicate-coop", "2026-07-13"],
                ["DUP-CODE", "OTHER-EXT", "duplicate-coop", "2026-07-14"],
                ["OTHER-CODE", "DUP-EXT", "duplicate-coop", "2026-07-15"],
            ],
        )
        preview = self._preview("collaborations", content, source_type="douyin_official")
        validated = self.client.post(
            f"/api/creators/import/{preview.json()['token']}/validate",
            json={"mapping": preview.json()["suggested_mapping"]},
        )
        self.assertEqual(200, validated.status_code, validated.text)
        self.assertEqual(1, validated.json()["imported_count"])
        self.assertEqual(2, validated.json()["error_count"])
        self.assertTrue(
            all("与第2行合作稳定编号重复" in item["message"] for item in validated.json()["errors"])
        )

    def test_exports_escape_formula_like_user_strings(self):
        creator = self.client.post(
            "/api/creators", json={"nickname": "\x01=1+1", "douyin_handle": "formula-export"}
        ).json()
        collaboration = self.client.post(
            f"/api/creators/{creator['id']}/collaborations",
            json={
                "internal_code": "=2+2",
                "collaboration_type": "live",
                "collaboration_date": "2026-07-13",
                "notes": "@SUM(1,1)",
            },
        )
        self.assertEqual(201, collaboration.status_code, collaboration.text)
        address = self.client.post(
            f"/api/creators/{creator['id']}/addresses",
            json={
                "recipient_name": "=3+3",
                "phone": "+13812345678",
                "province": "广东省",
                "city": "深圳市",
                "detail": "=WEBSERVICE(\"https://example.invalid\")",
            },
        ).json()
        self.client.post(
            f"/api/creators/{creator['id']}/sample-orders",
            json={
                "idempotency_key": "formula-export-order",
                "address_id": address["id"],
                "items": [{"product_id": self.product.id}],
            },
        )

        expectations = {
            "creators": ("A2", "'=1+1"),
            "collaborations": ("A2", "'=2+2"),
            "sample_orders": ("D3", "'=3+3"),
        }
        for entity, (coordinate, expected) in expectations.items():
            with self.subTest(entity=entity):
                response = self.client.get(
                    "/api/creators/export", params={"entity": entity, "creator_id": creator["id"]}
                )
                self.assertEqual(200, response.status_code, response.text)
                cell = load_workbook(BytesIO(response.content), data_only=False).active[coordinate]
                self.assertEqual(expected, cell.value)
                self.assertEqual("s", cell.data_type)

    def test_exports_limit_each_entity_to_configured_data_rows(self):
        for index in range(2):
            creator = self.client.post(
                "/api/creators",
                json={"nickname": f"限量达人{index}", "douyin_handle": f"limited-export-{index}"},
            ).json()
            self.client.post(
                f"/api/creators/{creator['id']}/collaborations",
                json={
                    "internal_code": f"LIMIT-COOP-{index}",
                    "collaboration_type": "live",
                    "collaboration_date": "2026-07-13",
                },
            )
            address = self.client.post(
                f"/api/creators/{creator['id']}/addresses",
                json={
                    "recipient_name": f"收件人{index}",
                    "phone": f"1380000000{index}",
                    "province": "广东省",
                    "city": "深圳市",
                    "detail": f"地址{index}",
                },
            ).json()
            self.client.post(
                f"/api/creators/{creator['id']}/sample-orders",
                json={
                    "idempotency_key": f"limited-export-order-{index}",
                    "address_id": address["id"],
                    "items": [{"product_id": self.product.id}],
                },
            )

        with patch("services.creator_importer.MAX_EXPORT_ROWS", 1, create=True):
            for entity, header_rows in (("creators", 1), ("collaborations", 1), ("sample_orders", 2)):
                with self.subTest(entity=entity):
                    response = self.client.get("/api/creators/export", params={"entity": entity})
                    self.assertEqual(200, response.status_code, response.text)
                    sheet = load_workbook(BytesIO(response.content)).active
                    self.assertEqual(1, sheet.max_row - header_rows)

    def test_creator_id_export_filter_is_applied_before_row_limit(self):
        target = self.client.post(
            "/api/creators",
            json={"nickname": "较早目标达人", "douyin_handle": "early-export-target"},
        ).json()
        self.client.post(
            "/api/creators",
            json={"nickname": "较新其他达人", "douyin_handle": "newer-export-other"},
        )

        with patch("services.creator_importer.MAX_EXPORT_ROWS", 1):
            response = self.client.get(
                "/api/creators/export",
                params={"entity": "creators", "creator_id": target["id"]},
            )

        self.assertEqual(200, response.status_code, response.text)
        text = "|".join(
            str(cell.value or "")
            for row in load_workbook(BytesIO(response.content)).active.iter_rows()
            for cell in row
        )
        self.assertIn("较早目标达人", text)
        self.assertNotIn("较新其他达人", text)

    def test_all_exports_use_bounded_query_count(self):
        for index in range(10):
            creator = self.client.post(
                "/api/creators",
                json={"nickname": f"批量导出达人{index}", "douyin_handle": f"bulk-export-{index}"},
            ).json()
            self.client.put(
                f"/api/creators/{creator['id']}/portrait",
                json={"follower_count": 10_000 + index, "primary_categories": ["烘焙"]},
            )
            self.client.post(
                f"/api/creators/{creator['id']}/collaborations",
                json={
                    "internal_code": f"QUERY-COOP-{index}",
                    "collaboration_type": "live",
                    "collaboration_date": "2026-07-13",
                    "products": [{"product_id": self.product.id}],
                },
            )
            address = self.client.post(
                f"/api/creators/{creator['id']}/addresses",
                json={
                    "recipient_name": f"查询收件人{index}",
                    "phone": f"138000001{index:02d}",
                    "province": "广东省",
                    "city": "深圳市",
                    "detail": f"查询地址{index}",
                },
            ).json()
            self.client.post(
                f"/api/creators/{creator['id']}/sample-orders",
                json={
                    "idempotency_key": f"query-count-order-{index}",
                    "address_id": address["id"],
                    "items": [{"product_id": self.product.id}],
                },
            )

        for entity in ("creators", "collaborations", "sample_orders"):
            with self.subTest(entity=entity):
                self.db.expire_all()
                statements = []

                def record_statement(_conn, _cursor, statement, _params, _context, _many):
                    statements.append(statement)

                event.listen(self.engine, "before_cursor_execute", record_statement)
                try:
                    response = self.client.get(
                        "/api/creators/export", params={"entity": entity}
                    )
                finally:
                    event.remove(self.engine, "before_cursor_execute", record_statement)

                self.assertEqual(200, response.status_code, response.text)
                self.assertLessEqual(len(statements), 3, statements)

    def test_archived_creator_import_is_reported_instead_of_updated(self):
        creator = self.client.post(
            "/api/creators", json={"nickname": "归档达人", "douyin_handle": "archived-import"}
        ).json()
        self.assertEqual(200, self.client.delete(f"/api/creators/{creator['id']}").status_code)
        content = workbook_bytes(
            ["达人昵称", "抖音号"], [["重新导入达人", "archived-import"]]
        )
        preview = self._preview("creators", content)
        validated = self.client.post(
            f"/api/creators/import/{preview.json()['token']}/validate",
            json={"mapping": preview.json()["suggested_mapping"]},
        )
        self.assertEqual(200, validated.status_code, validated.text)
        self.assertEqual(0, validated.json()["imported_count"])
        self.assertEqual(1, validated.json()["error_count"])
        self.assertIn("已归档", validated.json()["errors"][0]["message"])

    def test_owner_errors_are_generic_and_duplicate_product_names_are_rejected(self):
        phone = "13999998888"
        content = workbook_bytes(["达人昵称", "抖音号", "负责人"], [["隐私达人", "private-owner", phone]])
        preview = self._preview("creators", content)
        validated = self.client.post(
            f"/api/creators/import/{preview.json()['token']}/validate",
            json={"mapping": preview.json()["suggested_mapping"]},
        )
        self.assertEqual(1, validated.json()["error_count"])
        self.assertNotIn(phone, validated.text)

        self.db.add(Product(name="法采草莓果酱", category="重复", price=10, status="active"))
        self.db.commit()
        creator = self.client.post(
            "/api/creators", json={"nickname": "产品歧义达人", "douyin_handle": "ambiguous-product"}
        ).json()
        collaboration = workbook_bytes(
            ["合作编号", "达人抖音号", "合作形式", "合作日期", "合作产品"],
            [["AMB-1", "ambiguous-product", "直播", "2026-07-12", "法采草莓果酱"]],
        )
        preview = self._preview("collaborations", collaboration)
        validated = self.client.post(
            f"/api/creators/import/{preview.json()['token']}/validate",
            json={"mapping": preview.json()["suggested_mapping"]},
        )
        self.assertEqual(1, validated.json()["error_count"])
        self.assertIn("不唯一", validated.json()["errors"][0]["message"])

    def test_sample_export_honors_owner_and_search_before_returning_full_pii(self):
        other_member = self.client.post("/api/creators/bd-members", json={"name": "小李"}).json()
        phones = []
        for index, member in enumerate((self.member, other_member), start=1):
            phone = f"1380000000{index}"
            phones.append(phone)
            creator = self.client.post(
                "/api/creators",
                json={
                    "nickname": f"导出筛选{index}",
                    "douyin_handle": f"export-filter-{index}",
                    "owner_id": member["id"],
                },
            ).json()
            address = self.client.post(
                f"/api/creators/{creator['id']}/addresses",
                json={"recipient_name": "王", "phone": phone, "province": "粤", "city": "深", "detail": f"地址{index}"},
            ).json()
            self.client.post(
                f"/api/creators/{creator['id']}/sample-orders",
                json={
                    "idempotency_key": f"filter-order-{index}",
                    "address_id": address["id"],
                    "items": [{"product_id": self.product.id}],
                },
            )
        response = self.client.get(
            "/api/creators/export",
            params={"entity": "sample_orders", "owner_id": self.member["id"], "search": "筛选1"},
        )
        text = "|".join(
            str(cell.value or "")
            for row in load_workbook(BytesIO(response.content)).active.iter_rows()
            for cell in row
        )
        self.assertIn(phones[0], text)
        self.assertNotIn(phones[1], text)

    def test_all_export_entities_honor_category_and_follower_filters(self):
        creators = []
        for index, (category, followers) in enumerate((("家庭烘焙", 180_000), ("数码", 880_000)), start=1):
            creator = self.client.post(
                "/api/creators",
                json={"nickname": f"筛选达人{index}", "douyin_handle": f"tier-export-{index}"},
            ).json()
            self.client.put(
                f"/api/creators/{creator['id']}/portrait",
                json={"primary_categories": [category], "follower_count": followers},
            )
            self.client.post(
                f"/api/creators/{creator['id']}/collaborations",
                json={
                    "internal_code": f"FILTER-COOP-{index}",
                    "collaboration_type": "live",
                    "collaboration_date": "2026-07-13",
                },
            )
            address = self.client.post(
                f"/api/creators/{creator['id']}/addresses",
                json={
                    "recipient_name": f"收件人{index}",
                    "phone": f"1380000000{index}",
                    "province": "广东省",
                    "city": "深圳市",
                    "detail": f"FILTER-ADDRESS-{index}",
                },
            ).json()
            self.client.post(
                f"/api/creators/{creator['id']}/sample-orders",
                json={
                    "idempotency_key": f"filter-category-order-{index}",
                    "address_id": address["id"],
                    "items": [{"product_id": self.product.id}],
                },
            )
            creators.append(creator)

        params = {"category": "家庭烘焙", "follower_tier": "100k_500k"}
        for entity in ("creators", "collaborations", "sample_orders"):
            response = self.client.get(
                "/api/creators/export", params={"entity": entity, **params}
            )
            self.assertEqual(200, response.status_code, response.text)
            text = "|".join(
                str(cell.value or "")
                for row in load_workbook(BytesIO(response.content)).active.iter_rows()
                for cell in row
            )
            self.assertIn(creators[0]["nickname"], text, entity)
            self.assertNotIn(creators[1]["nickname"], text, entity)

    def test_cleanup_lock_after_commit_does_not_turn_success_into_500(self):
        content = workbook_bytes(["达人昵称", "抖音号"], [["锁文件达人", "locked-file"]])
        preview = self._preview("creators", content)
        token = preview.json()["token"]
        staged_path = creator_importer.CREATOR_IMPORT_DIR / f"{token}.xlsx"
        self.client.post(
            f"/api/creators/import/{token}/validate",
            json={"mapping": preview.json()["suggested_mapping"]},
        )
        with patch("services.creator_importer.Path.unlink", side_effect=PermissionError("locked")):
            response = self.client.post(f"/api/creators/import/{token}/commit")
        self.assertEqual(200, response.status_code, response.text)
        self.assertTrue(staged_path.exists())
        self.assertEqual(1, self.client.get("/api/creators", params={"search": "锁文件"}).json()["total"])

        retry = self._preview(
            "creators",
            workbook_bytes(["达人昵称", "抖音号"], [["触发清理达人", "cleanup-retry-trigger"]]),
        )
        self.assertEqual(200, retry.status_code, retry.text)
        self.assertFalse(staged_path.exists())

    def test_preview_commit_failure_removes_untracked_staged_file(self):
        content = workbook_bytes(["达人昵称", "抖音号"], [["提交失败达人", "failed-preview"]])
        with patch.object(self.db, "commit", side_effect=RuntimeError("database unavailable")):
            with self.assertRaises(RuntimeError):
                self._preview("creators", content)
        self.db.rollback()
        self.assertEqual([], list(creator_importer.CREATOR_IMPORT_DIR.glob("*.xlsx")))

    def test_stale_untracked_staged_file_is_retried_on_next_preview(self):
        orphan = creator_importer.CREATOR_IMPORT_DIR / "orphan-without-batch.xlsx"
        orphan.write_bytes(b"private workbook bytes")
        stale_time = datetime.now(UTC).timestamp() - creator_importer.ORPHAN_CLEANUP_GRACE_SECONDS - 1
        os.utime(orphan, (stale_time, stale_time))

        response = self._preview(
            "creators",
            workbook_bytes(["达人昵称", "抖音号"], [["清理触发达人", "orphan-cleanup"]]),
        )

        self.assertEqual(200, response.status_code, response.text)
        self.assertFalse(orphan.exists())


if __name__ == "__main__":
    unittest.main()
