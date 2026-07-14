import os
import tempfile
import unittest
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch
from uuid import uuid4

from openpyxl import load_workbook
from pydantic import ValidationError

from integrations.exports import (
    ExportColumn,
    _iter_export_items,
    escape_spreadsheet_text,
    export_job_view,
    scan_orphan_exports,
    write_export_file,
)
from integrations.reporting import PageResult
from integrations.types import ExportStatus, ResourceType
from integrations.schemas import ExportCreateRequest


UTC = timezone.utc


class ExportSafetyTests(unittest.TestCase):
    def test_export_request_is_strict_and_filter_keys_are_closed(self):
        request = ExportCreateRequest.model_validate(
            {
                "resource_type": "products",
                "format": "csv",
                "filters": {
                    "provider": "doudian",
                    "date_from": "2026-01-01",
                    "link_status": "unlinked",
                },
            }
        )
        self.assertEqual(request.resource_type.value, "products")
        self.assertEqual(request.filters.link_status, "unlinked")
        with self.assertRaises(ValidationError):
            ExportCreateRequest.model_validate(
                {
                    "resource_type": "orders",
                    "format": "pdf",
                    "filters": {"buyer_phone": "13800138000"},
                }
            )
        with self.assertRaises(ValidationError):
            ExportCreateRequest.model_validate(
                {
                    "resource_type": "products",
                    "format": "csv",
                    "filters": {"link_status": "all"},
                }
            )
        with self.assertRaises(ValidationError):
            ExportCreateRequest.model_validate(
                {
                    "resource_type": "orders",
                    "format": "csv",
                    "filters": {"link_status": "linked"},
                }
            )
        with self.assertRaises(ValidationError):
            ExportCreateRequest.model_validate(
                {
                    "resource_type": "refunds",
                    "format": "csv",
                    "filters": {"status": "paid"},
                }
            )
        for sensitive_search in (
            "13800138000",
            "11010520000101001X",
            "Bearer secret-token",
            "client-secret=abc",
            "refresh token=abc",
            "app_secret=abc",
        ):
            with self.subTest(search=sensitive_search), self.assertRaises(
                ValidationError
            ):
                ExportCreateRequest.model_validate(
                    {
                        "resource_type": "orders",
                        "format": "csv",
                        "filters": {"search": sensitive_search},
                    }
                )
        with self.assertRaises(ValidationError):
            ExportCreateRequest.model_validate(
                {
                    "resource_type": "orders",
                    "format": "csv",
                    "filters": {
                        "date_from": "2025-01-01",
                        "date_to": "2026-01-02",
                    },
                }
            )

    def test_one_escape_seam_blocks_all_spreadsheet_formula_prefixes(self):
        dangerous = (
            "=SUM(1,1)",
            "+cmd|' /C calc'!A0",
            "-1+1",
            "@SUM(A1:A2)",
            "  =HYPERLINK(\"https://evil.invalid\")",
            "\t=cmd",
            "\r+cmd",
            "\n@cmd",
        )
        for value in dangerous:
            with self.subTest(value=repr(value)):
                escaped = escape_spreadsheet_text(value)
                self.assertTrue(escaped.startswith("'"))
                self.assertEqual(escaped[1:], value)

        self.assertEqual(escape_spreadsheet_text("00123"), "00123")
        self.assertEqual(escape_spreadsheet_text("ordinary title"), "ordinary title")

    def test_csv_and_xlsx_are_atomic_explicit_and_formula_free(self):
        columns = (
            ExportColumn("external_id", "平台 ID", "text"),
            ExportColumn("title", "标题", "text"),
            ExportColumn("amount", "金额", "decimal"),
            ExportColumn("count", "数量", "integer"),
            ExportColumn("updated_at", "更新时间", "datetime"),
        )
        rows = (
            {
                "external_id": "00123",
                "title": " =SUM(1,1)",
                "amount": Decimal("12.30"),
                "count": 2,
                "updated_at": datetime(2026, 7, 13, 1, 2, tzinfo=UTC),
                "ignored_secret": "must-not-export",
            },
        )
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            csv_artifact = write_export_file(
                archive_dir=root,
                public_id=str(uuid4()),
                export_format="csv",
                columns=columns,
                rows=rows,
            )
            xlsx_artifact = write_export_file(
                archive_dir=root,
                public_id=str(uuid4()),
                export_format="xlsx",
                columns=columns,
                rows=rows,
            )

            csv_bytes = (root / csv_artifact.relative_path).read_bytes()
            self.assertTrue(csv_bytes.startswith(b"\xef\xbb\xbf"))
            csv_text = csv_bytes.decode("utf-8-sig")
            self.assertIn("00123", csv_text)
            self.assertIn("' =SUM(1,1)", csv_text)
            self.assertNotIn("must-not-export", csv_text)

            workbook = load_workbook(
                root / xlsx_artifact.relative_path,
                read_only=False,
                data_only=False,
            )
            sheet = workbook.active
            self.assertEqual(sheet["A2"].value, "00123")
            self.assertEqual(sheet["A2"].data_type, "s")
            self.assertEqual(sheet["B2"].value, "' =SUM(1,1)")
            self.assertEqual(sheet["B2"].data_type, "s")
            self.assertTrue(
                all(cell.data_type != "f" for row in sheet.iter_rows() for cell in row)
            )
            self.assertEqual(csv_artifact.row_count, 1)
            self.assertEqual(xlsx_artifact.row_count, 1)
            self.assertFalse(list((root / "exports").glob("*.tmp")))

    def test_public_id_and_columns_fail_closed(self):
        with tempfile.TemporaryDirectory() as directory:
            with self.assertRaises(ValueError):
                write_export_file(
                    archive_dir=directory,
                    public_id="../escape",
                    export_format="csv",
                    columns=(ExportColumn("id", "ID", "text"),),
                    rows=(),
                )
            with self.assertRaises(ValueError):
                ExportColumn("secret", "Secret", "unknown")

    def test_expired_view_does_not_mark_ready_job_before_file_cleanup(self):
        now = datetime(2026, 7, 13, 1, 0, tzinfo=UTC)
        job = SimpleNamespace(
            public_id=str(uuid4()),
            requester_session_digest="a" * 64,
            resource_type=ResourceType.ORDERS,
            format="csv",
            status=ExportStatus.READY,
            row_count=1,
            created_at=now - timedelta(hours=25),
            expires_at=now - timedelta(hours=1),
            error_code=None,
            error_summary=None,
        )

        view = export_job_view(job, now=now)

        self.assertEqual(view["status"], ExportStatus.EXPIRED.value)
        self.assertIsNone(view["download_url"])
        self.assertIs(job.status, ExportStatus.READY)

    def test_export_rows_are_fetched_and_yielded_one_page_at_a_time(self):
        now = datetime(2026, 7, 13, 1, 0, tzinfo=UTC)
        job = SimpleNamespace(
            resource_type=ResourceType.AD_ENTITIES,
            filters={"date_from": "2026-07-13", "date_to": "2026-07-13"},
        )
        first_items = [
            {"connection_id": 1, "external_entity_id": f"entity-{index}"}
            for index in range(200)
        ]
        second_items = [
            {"connection_id": 1, "external_entity_id": "entity-200"}
        ]
        pages = (
            PageResult(first_items, 201, 1, 200, 2),
            PageResult(second_items, 201, 2, 200, 2),
        )
        with patch("integrations.exports.list_ad_entities", side_effect=pages) as query:
            rows = _iter_export_items(object(), export_job=job, now=now)
            self.assertEqual(query.call_count, 0)
            first = next(rows)
            self.assertEqual(first["external_entity_id"], "entity-0")
            self.assertEqual(query.call_count, 1)
            remaining = list(rows)

        self.assertEqual(len(remaining), 200)
        self.assertEqual(query.call_count, 2)

    def test_orphan_scan_removes_old_final_and_temp_but_preserves_young_and_unknown(self):
        now = datetime(2026, 7, 13, 3, 0, tzinfo=UTC)
        with tempfile.TemporaryDirectory() as directory:
            export_dir = Path(directory) / "exports"
            export_dir.mkdir()
            old_id = str(uuid4())
            young_id = str(uuid4())
            old_final = export_dir / f"{old_id}.csv"
            old_temp = export_dir / f".{old_id}.abc12345.csv.tmp"
            young_temp = export_dir / f".{young_id}.abc12345.xlsx.tmp"
            unknown = export_dir / ".not-an-export.tmp"
            for path in (old_final, old_temp, young_temp, unknown):
                path.write_text("sensitive-row", encoding="utf-8")
            old_timestamp = (now - timedelta(hours=2)).timestamp()
            os.utime(old_final, (old_timestamp, old_timestamp))
            os.utime(old_temp, (old_timestamp, old_timestamp))
            os.utime(unknown, (old_timestamp, old_timestamp))

            result = scan_orphan_exports(
                archive_dir=directory,
                known_relative_paths=(),
                now=now,
            )

            self.assertEqual(result.failure_count, 0)
            self.assertFalse(old_final.exists())
            self.assertFalse(old_temp.exists())
            self.assertTrue(young_temp.exists())
            self.assertTrue(unknown.exists())


if __name__ == "__main__":
    unittest.main()
