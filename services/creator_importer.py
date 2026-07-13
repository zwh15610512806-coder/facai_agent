"""Excel templates, staged imports, error reports, and exports for creators."""
from __future__ import annotations

import hashlib
import json
import re
import uuid
import zipfile
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
from pathlib import Path
from typing import Any

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from sqlalchemy.orm import Session, joinedload, selectinload
from sqlalchemy import exists, func, or_, select
from sqlalchemy.exc import IntegrityError

from config import UPLOAD_DIR
from creator_models import (
    BdMember,
    Creator,
    CreatorAddress,
    CreatorCollaboration,
    CreatorCollaborationProduct,
    CreatorImportBatch,
    CreatorPortrait,
    CreatorSampleOrder,
    normalize_douyin_handle,
    normalize_platform_uid,
)
from creator_schemas import (
    AmountStatus,
    CollaborationStatus,
    CollaborationType,
    CreatorStage,
)
from models import Product
from services import creator_service


CREATOR_IMPORT_DIR = Path(UPLOAD_DIR) / "creator_imports"
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
IMPORT_TOKEN_TTL_HOURS = 24
MAX_XLSX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
MAX_XLSX_ENTRIES = 2_000
MAX_IMPORT_ROWS = 20_000
MAX_IMPORT_COLUMNS = 200
MAX_CELL_CHARACTERS = 100_000
MAX_EXPORT_ROWS = 50_000
ORPHAN_CLEANUP_GRACE_SECONDS = 300

CREATOR_TEMPLATE = [
    "平台",
    "官方达人ID",
    "抖音号",
    "达人昵称",
    "主页链接",
    "MCN机构",
    "负责人",
    "当前阶段",
    "标签",
    "粉丝数",
    "主营垂类",
    "内容形式",
    "受众画像",
    "地区",
    "内容风格",
    "合作偏好",
    "价格带",
    "匹配度",
    "风险备注",
    "联系人",
    "手机号",
    "微信号",
    "收件人",
    "收件电话",
    "省",
    "市",
    "区",
    "详细地址",
]
COLLABORATION_TEMPLATE = [
    "合作编号",
    "平台记录ID",
    "达人官方ID",
    "达人抖音号",
    "合作形式",
    "合作日期",
    "合作状态",
    "实际支付金额（元）",
    "金额状态",
    "负责人",
    "合作产品",
    "备注",
]

CREATOR_ALIASES = {
    "平台": "platform",
    "官方达人id": "platform_uid",
    "达人id": "platform_uid",
    "抖音号": "douyin_handle",
    "达人昵称": "nickname",
    "昵称": "nickname",
    "主页链接": "homepage_url",
    "mcn机构": "mcn_name",
    "mcn": "mcn_name",
    "负责人": "owner_name",
    "bd负责人": "owner_name",
    "当前阶段": "stage",
    "标签": "tags",
    "粉丝数": "follower_count",
    "主营垂类": "primary_categories",
    "内容形式": "content_formats",
    "受众画像": "audience_profile",
    "地区": "regions",
    "内容风格": "style_tags",
    "合作偏好": "cooperation_preferences",
    "价格带": "price_range",
    "匹配度": "fit_score",
    "风险备注": "risk_notes",
    "联系人": "contact_name",
    "手机号": "contact_phone",
    "微信号": "wechat_id",
    "收件人": "recipient_name",
    "收件电话": "recipient_phone",
    "省": "province",
    "市": "city",
    "区": "district",
    "详细地址": "address_detail",
}
COLLABORATION_ALIASES = {
    "合作编号": "internal_code",
    "平台记录id": "external_record_id",
    "达人官方id": "creator_platform_uid",
    "达人抖音号": "creator_douyin_handle",
    "合作形式": "collaboration_type",
    "合作日期": "collaboration_date",
    "合作状态": "status",
    "实际支付金额（元）": "actual_paid_yuan",
    "实际支付金额": "actual_paid_yuan",
    "支付金额": "actual_paid_yuan",
    "金额状态": "amount_status",
    "负责人": "owner_name",
    "合作产品": "product_names",
    "产品": "product_names",
    "备注": "notes",
}
CREATOR_CANONICAL_FIELDS = set(CREATOR_ALIASES.values())
COLLABORATION_CANONICAL_FIELDS = set(COLLABORATION_ALIASES.values())

STAGE_VALUES = {item.value: item.value for item in CreatorStage}
STAGE_VALUES.update(
    {
        "待联系": "lead",
        "已建联": "contacted",
        "洽谈中": "negotiating",
        "已寄样": "sampled",
        "已排期": "scheduled",
        "合作中": "cooperating",
        "已完成": "completed",
        "暂停": "paused",
    }
)
COLLABORATION_TYPE_VALUES = {item.value: item.value for item in CollaborationType}
COLLABORATION_TYPE_VALUES.update({"短视频": "short_video", "直播": "live", "图文": "graphic", "其他": "other"})
COLLABORATION_STATUS_VALUES = {item.value: item.value for item in CollaborationStatus}
COLLABORATION_STATUS_VALUES.update({"待合作": "planned", "进行中": "in_progress", "已完成": "completed", "已取消": "cancelled"})
AMOUNT_STATUS_VALUES = {item.value: item.value for item in AmountStatus}
AMOUNT_STATUS_VALUES.update({"待确认": "pending", "已确认": "confirmed"})


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).lower()


def _text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _limited_text(value: Any, label: str, maximum: int) -> str | None:
    text = _text(value)
    if text is not None and len(text) > maximum:
        raise ValueError(f"{label}长度不能超过{maximum}个字符")
    return text


def _split(value: Any) -> list[str]:
    text = _text(value)
    if not text:
        return []
    return [item.strip() for item in re.split(r"[|｜、,，;；]", text) if item.strip()]


def _provided_fields(row: dict[str, Any]) -> set[str]:
    return {
        field
        for field, value in row.items()
        if value is not None and str(value).strip() != ""
    }


def _import_row_fingerprint(row: dict[str, Any]) -> str:
    normalized = {}
    for field, value in row.items():
        if value is None or str(value).strip() == "":
            continue
        if isinstance(value, (date, datetime)):
            normalized[field] = value.isoformat()
        elif isinstance(value, (float, Decimal)):
            normalized[field] = str(Decimal(str(value)).normalize())
        else:
            normalized[field] = str(value).strip()
    payload = json.dumps(normalized, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _parse_int(value: Any, label: str, *, minimum: int | None = None, maximum: int | None = None) -> int | None:
    if value in (None, ""):
        return None
    try:
        number = Decimal(str(value).replace(",", ""))
        if not number.is_finite() or number != number.to_integral_value():
            raise ValueError
        result = int(number)
    except (InvalidOperation, OverflowError, ValueError):
        raise ValueError(f"{label}必须是整数")
    if minimum is not None and result < minimum:
        raise ValueError(f"{label}不能小于{minimum}")
    if maximum is not None and result > maximum:
        raise ValueError(f"{label}不能大于{maximum}")
    return result


def _parse_money_cents(value: Any) -> int:
    if value in (None, ""):
        return 0
    try:
        money = Decimal(str(value).replace(",", ""))
    except InvalidOperation:
        raise ValueError("实际支付金额必须是数字")
    if not money.is_finite():
        raise ValueError("实际支付金额必须是有限数字")
    if money < 0:
        raise ValueError("实际支付金额不能为负数")
    try:
        return int((money * 100).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    except (InvalidOperation, OverflowError, ValueError):
        raise ValueError("实际支付金额必须是有效数字")


def _parse_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    if not text:
        raise ValueError("合作日期不能为空")
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError("合作日期格式应为 YYYY-MM-DD")


def _mapping_for(kind: str, headers: list[str]) -> dict[str, str]:
    aliases = CREATOR_ALIASES if kind == "creators" else COLLABORATION_ALIASES
    result = {}
    for header in headers:
        canonical = aliases.get(_normalize_header(header))
        if canonical:
            result[header] = canonical
    return result


def _safe_unlink(path: Path) -> None:
    try:
        path.unlink()
    except OSError:
        pass


def _retry_terminal_staged_file_cleanup(db: Session) -> None:
    """Retry deletion of terminal-batch uploads that were temporarily locked."""
    if not CREATOR_IMPORT_DIR.exists():
        return
    staged_paths = []
    for path in CREATOR_IMPORT_DIR.glob("*.xlsx"):
        staged_paths.append(path)
        if len(staged_paths) >= 500:
            break
    if not staged_paths:
        return
    tokens = [path.stem for path in staged_paths]
    batch_statuses = {
        token: status
        for token, status in db.query(
            CreatorImportBatch.token, CreatorImportBatch.status
        ).filter(CreatorImportBatch.token.in_(tokens))
    }
    orphan_cutoff = datetime.now(UTC).timestamp() - ORPHAN_CLEANUP_GRACE_SECONDS
    for path in staged_paths:
        status = batch_statuses.get(path.stem)
        is_terminal = status in {"committed", "duplicate", "expired"}
        try:
            is_stale_orphan = status is None and path.stat().st_mtime < orphan_cutoff
        except OSError:
            is_stale_orphan = False
        if is_terminal or is_stale_orphan:
            _safe_unlink(path)


def _validate_xlsx_container(path_or_bytes: Path | bytes) -> None:
    source = BytesIO(path_or_bytes) if isinstance(path_or_bytes, bytes) else path_or_bytes
    try:
        with zipfile.ZipFile(source) as archive:
            infos = archive.infolist()
            if len(infos) > MAX_XLSX_ENTRIES:
                raise HTTPException(status_code=400, detail="Excel 压缩包文件项过多")
            expanded = sum(max(0, item.file_size) for item in infos)
            if expanded > MAX_XLSX_UNCOMPRESSED_BYTES:
                raise HTTPException(status_code=400, detail="Excel 解压后体积过大")
    except HTTPException:
        raise
    except (zipfile.BadZipFile, OSError) as exc:
        raise HTTPException(status_code=400, detail="Excel 文件不是有效的 .xlsx 压缩包") from exc


def _read_workbook(path_or_bytes: Path | bytes) -> tuple[list[str], list[tuple[int, dict[str, Any]]]]:
    _validate_xlsx_container(path_or_bytes)
    source = BytesIO(path_or_bytes) if isinstance(path_or_bytes, bytes) else path_or_bytes
    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Excel 文件无法读取") from exc
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header_values = next(rows, None)
        if not header_values:
            raise HTTPException(status_code=400, detail="Excel 文件没有表头")
        headers = [str(value or "").strip() for value in header_values]
        if not any(headers):
            raise HTTPException(status_code=400, detail="Excel 文件没有有效表头")
        if len(headers) > MAX_IMPORT_COLUMNS:
            raise HTTPException(status_code=400, detail="Excel 列数超过限制")
        nonblank_headers = [header for header in headers if header]
        if len(nonblank_headers) != len(set(nonblank_headers)):
            raise HTTPException(status_code=400, detail="Excel 表头不能重复")
        parsed = []
        for row_number, values in enumerate(rows, start=2):
            if not any(value not in (None, "") for value in values):
                continue
            if len(parsed) >= MAX_IMPORT_ROWS:
                raise HTTPException(status_code=400, detail="Excel 数据行数超过限制")
            for value in values:
                if isinstance(value, str) and len(value) > MAX_CELL_CHARACTERS:
                    raise HTTPException(status_code=400, detail="Excel 单元格内容过长")
            parsed.append((row_number, {headers[index]: values[index] if index < len(values) else None for index in range(len(headers))}))
        return headers, parsed
    finally:
        workbook.close()


def build_template(kind: str) -> bytes:
    if kind not in {"creators", "collaborations"}:
        raise HTTPException(status_code=404, detail="导入模板不存在")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "达人资料" if kind == "creators" else "合作记录"
    sheet.append(CREATOR_TEMPLATE if kind == "creators" else COLLABORATION_TEMPLATE)
    sheet.freeze_panes = "A2"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def preview_upload(
    db: Session,
    *,
    kind: str,
    source_type: str,
    filename: str,
    content: bytes,
) -> dict:
    if kind not in {"creators", "collaborations"}:
        raise HTTPException(status_code=400, detail="kind 仅支持 creators 或 collaborations")
    _cleanup_expired_batches(db)
    sha256 = hashlib.sha256(content).hexdigest()
    duplicate = db.query(CreatorImportBatch).filter(
        CreatorImportBatch.kind == kind,
        CreatorImportBatch.file_sha256 == sha256,
        CreatorImportBatch.status == "committed",
    ).first()
    if duplicate:
        raise HTTPException(status_code=409, detail="该文件已经导入")
    headers, rows = _read_workbook(content)
    token = uuid.uuid4().hex
    CREATOR_IMPORT_DIR.mkdir(parents=True, exist_ok=True)
    path = CREATOR_IMPORT_DIR / f"{token}.xlsx"
    path.write_bytes(content)
    batch = CreatorImportBatch(
        token=token,
        kind=kind,
        source_type=(source_type or "facai_template")[:50],
        filename=Path(filename).name[:500],
        file_sha256=sha256,
        status="previewed",
        row_count=len(rows),
    )
    db.add(batch)
    try:
        db.commit()
    except Exception:
        db.rollback()
        _safe_unlink(path)
        raise
    return {
        "token": token,
        "kind": kind,
        "source_type": batch.source_type,
        "headers": headers,
        "suggested_mapping": _mapping_for(kind, headers),
        "sample_rows": [row for _, row in rows[:5]],
        "row_count": len(rows),
    }


def _batch(db: Session, token: str) -> CreatorImportBatch:
    _retry_terminal_staged_file_cleanup(db)
    batch = db.query(CreatorImportBatch).filter(CreatorImportBatch.token == token).first()
    if not batch:
        raise HTTPException(status_code=404, detail="导入任务不存在")
    if batch.status == "expired" or (
        batch.status in {"previewed", "validated"}
        and batch.created_at
        and batch.created_at < datetime.now(UTC).replace(tzinfo=None) - timedelta(hours=IMPORT_TOKEN_TTL_HOURS)
    ):
        batch.status = "expired"
        db.commit()
        _safe_unlink(CREATOR_IMPORT_DIR / f"{batch.token}.xlsx")
        raise HTTPException(status_code=410, detail="导入任务已过期")
    return batch


def _cleanup_expired_batches(db: Session) -> None:
    _retry_terminal_staged_file_cleanup(db)
    cutoff = datetime.now(UTC).replace(tzinfo=None) - timedelta(hours=IMPORT_TOKEN_TTL_HOURS)
    expired = db.query(CreatorImportBatch).filter(
        CreatorImportBatch.created_at < cutoff,
        CreatorImportBatch.status.in_(["previewed", "validated", "expired"]),
    ).all()
    if not expired:
        return
    for batch in expired:
        batch.status = "expired"
        _safe_unlink(CREATOR_IMPORT_DIR / f"{batch.token}.xlsx")
    db.commit()


def _verify_staged_file(batch: CreatorImportBatch) -> Path:
    path = CREATOR_IMPORT_DIR / f"{batch.token}.xlsx"
    if not path.exists():
        raise HTTPException(status_code=410, detail="导入临时文件已失效")
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    if digest != batch.file_sha256:
        raise HTTPException(status_code=409, detail="导入文件内容已变化，请重新上传")
    return path


def _mapped_rows(batch: CreatorImportBatch, mapping: dict[str, str]):
    path = _verify_staged_file(batch)
    _, rows = _read_workbook(path)
    for row_number, source in rows:
        yield row_number, {
            canonical: source.get(header)
            for header, canonical in mapping.items()
            if header in source and canonical
        }


def _owner_id(db: Session, name: Any) -> int | None:
    owner_name = _limited_text(name, "负责人", 100)
    if not owner_name:
        return None
    owner = db.query(BdMember).filter(BdMember.name == owner_name, BdMember.active.is_(True)).first()
    if not owner:
        raise ValueError("负责人不存在或已停用")
    return owner.id


def _creator_by_identity(
    db: Session,
    platform: str,
    platform_uid: Any,
    douyin_handle: Any,
) -> Creator | None:
    normalized_uid = normalize_platform_uid(_text(platform_uid))
    handle = normalize_douyin_handle(_text(douyin_handle))
    normalized_handle = handle.casefold() if handle else None
    uid_creator = None
    handle_creator = None
    if normalized_uid:
        uid_creator = db.query(Creator).filter(
            Creator.platform == platform,
            Creator.platform_uid_normalized == normalized_uid,
        ).first()
    if normalized_handle:
        handle_creator = db.query(Creator).filter(
            Creator.platform == platform,
            Creator.douyin_handle_normalized == normalized_handle,
        ).first()
    if uid_creator and handle_creator and uid_creator.id != handle_creator.id:
        raise ValueError("达人身份冲突，请人工核对官方 ID 与抖音号")
    creator = uid_creator or handle_creator
    if creator and creator.archived_at is not None:
        raise ValueError("达人已归档，请先恢复后再导入")
    return creator


def _normalize_creator_row(db: Session, row: dict[str, Any]) -> dict:
    provided = _provided_fields(row)
    nickname = _limited_text(row.get("nickname"), "达人昵称", 200)
    platform = (_limited_text(row.get("platform"), "平台", 30) or "douyin").lower()
    platform_uid = _limited_text(row.get("platform_uid"), "官方达人 ID", 200)
    handle = normalize_douyin_handle(_limited_text(row.get("douyin_handle"), "抖音号", 200))
    if not platform_uid and not handle:
        raise ValueError("官方达人 ID 和抖音号至少填写一项")
    existing = _creator_by_identity(db, platform, platform_uid, handle)
    if not nickname and not existing:
        raise ValueError("新增达人时昵称不能为空")
    nickname = nickname or existing.nickname
    stage_text = _text(row.get("stage")) or (existing.stage if existing else "lead")
    stage = STAGE_VALUES.get(stage_text)
    if not stage:
        raise ValueError("达人阶段无效")
    audience = _text(row.get("audience_profile"))
    try:
        audience_profile = json.loads(audience) if audience and audience.startswith("{") else ({"description": audience} if audience else {})
    except json.JSONDecodeError:
        audience_profile = {"description": audience}
    address = {
        "recipient_name": _limited_text(row.get("recipient_name"), "收件人", 100),
        "phone": _limited_text(row.get("recipient_phone"), "收件电话", 50),
        "province": _limited_text(row.get("province"), "省", 100),
        "city": _limited_text(row.get("city"), "市", 100),
        "district": _limited_text(row.get("district"), "区", 100),
        "detail": _limited_text(row.get("address_detail"), "详细地址", 1000),
    }
    address_fields = {
        "recipient_name",
        "recipient_phone",
        "province",
        "city",
        "district",
        "address_detail",
    }
    if provided & address_fields:
        required_address = {
            "recipient_name": "收件人",
            "phone": "收件电话",
            "province": "省",
            "city": "市",
            "detail": "详细地址",
        }
        missing = [label for field, label in required_address.items() if not address[field]]
        if missing:
            raise ValueError("填写寄样地址时缺少：" + "、".join(missing))
    return {
        "_provided_fields": provided,
        "_existing_creator_id": existing.id if existing else None,
        "platform": platform,
        "platform_uid": platform_uid,
        "douyin_handle": handle,
        "nickname": nickname,
        "homepage_url": _limited_text(row.get("homepage_url"), "主页链接", 1000),
        "mcn_name": _limited_text(row.get("mcn_name"), "MCN 机构", 200),
        "owner_id": _owner_id(db, row.get("owner_name")) if "owner_name" in provided else None,
        "stage": stage,
        "tags": _split(row.get("tags")),
        "contact_name": _limited_text(row.get("contact_name"), "联系人", 100),
        "contact_phone": _limited_text(row.get("contact_phone"), "手机号", 50),
        "wechat_id": _limited_text(row.get("wechat_id"), "微信号", 100),
        "portrait": {
            "follower_count": _parse_int(row.get("follower_count"), "粉丝数", minimum=0),
            "primary_categories": _split(row.get("primary_categories")),
            "content_formats": _split(row.get("content_formats")),
            "audience_profile": audience_profile,
            "regions": _split(row.get("regions")),
            "style_tags": _split(row.get("style_tags")),
            "cooperation_preferences": _split(row.get("cooperation_preferences")),
            "price_range": _limited_text(row.get("price_range"), "价格带", 200),
            "fit_score": _parse_int(row.get("fit_score"), "匹配度", minimum=1, maximum=5),
            "risk_notes": _limited_text(row.get("risk_notes"), "风险备注", 10000),
        },
        "address": address,
    }


def _normalize_collaboration_row(
    db: Session, row: dict[str, Any], source_type: str
) -> dict:
    provided = _provided_fields(row)
    internal_code = _limited_text(row.get("internal_code"), "合作编号", 100)
    external_record_id = _limited_text(row.get("external_record_id"), "平台记录 ID", 200)
    if not internal_code and not external_record_id:
        source_hash = hashlib.sha256(source_type.encode("utf-8")).hexdigest()[:8]
        internal_code = f"IMP-{source_hash}-{_import_row_fingerprint(row)[:24]}"
    by_external = None
    if external_record_id:
        by_external = db.query(CreatorCollaboration).filter(
            CreatorCollaboration.source_type == source_type,
            CreatorCollaboration.external_record_id == external_record_id,
        ).first()
    by_code = None
    if internal_code:
        by_code = db.query(CreatorCollaboration).filter(
            CreatorCollaboration.internal_code == internal_code
        ).first()
    if by_external and by_code and by_external.id != by_code.id:
        raise ValueError("合作身份冲突，请人工核对平台记录 ID 与合作编号")
    existing = by_external or by_code
    if not internal_code:
        internal_code = existing.internal_code if existing else (
            f"{source_type}-{hashlib.sha256(external_record_id.encode('utf-8')).hexdigest()[:20]}"
        )

    creator_uid = _limited_text(row.get("creator_platform_uid"), "达人官方 ID", 200)
    creator_handle = _limited_text(row.get("creator_douyin_handle"), "达人抖音号", 200)
    creator = None
    if creator_uid or creator_handle:
        creator = _creator_by_identity(db, "douyin", creator_uid, creator_handle)
        if not creator:
            raise ValueError("未找到对应达人")
    elif existing:
        creator = existing.creator
    if not creator:
        raise ValueError("新增合作时必须提供达人官方 ID 或抖音号")
    if creator.archived_at is not None:
        raise ValueError("达人已归档，请先恢复后再导入")
    if existing and existing.creator_id != creator.id:
        raise ValueError("合作记录与达人身份不一致")

    type_text = _text(row.get("collaboration_type")) or (
        existing.collaboration_type if existing else "other"
    )
    status_text = _text(row.get("status")) or (existing.status if existing else "planned")
    amount_text = _text(row.get("amount_status")) or (
        existing.amount_status if existing else "pending"
    )
    collaboration_type = COLLABORATION_TYPE_VALUES.get(type_text)
    status = COLLABORATION_STATUS_VALUES.get(status_text)
    amount_status = AMOUNT_STATUS_VALUES.get(amount_text)
    if not collaboration_type:
        raise ValueError("合作形式无效")
    if not status:
        raise ValueError("合作状态无效")
    if not amount_status:
        raise ValueError("金额状态无效")
    if existing and status not in creator_service.COLLABORATION_TRANSITIONS[existing.status]:
        raise ValueError("合作状态不能这样流转")
    products = None
    if "product_names" in provided:
        product_names = _split(row.get("product_names"))
        if len(product_names) != len(set(product_names)):
            raise ValueError("同一合作不能重复填写相同产品")
        products = []
        for name in product_names:
            matches = db.query(Product).filter(Product.name == name).all()
            if not matches:
                raise ValueError(f"产品“{name}”不存在")
            if len(matches) > 1:
                raise ValueError(f"产品名称“{name}”不唯一，请先整理产品库")
            product = matches[0]
            products.append({"product_id": product.id, "product_name_snapshot": product.name})
    collaboration_date = (
        _parse_date(row.get("collaboration_date"))
        if "collaboration_date" in provided
        else (existing.collaboration_date if existing else None)
    )
    if collaboration_date is None:
        raise ValueError("新增合作时合作日期不能为空")
    return {
        "_provided_fields": provided,
        "creator_id": creator.id,
        "owner_id": _owner_id(db, row.get("owner_name")) if "owner_name" in provided else (existing.owner_id if existing else None),
        "source_type": existing.source_type if existing else source_type,
        "external_record_id": external_record_id or (existing.external_record_id if existing else None),
        "internal_code": internal_code,
        "collaboration_type": collaboration_type,
        "collaboration_date": collaboration_date,
        "status": status,
        "actual_paid_cents": (
            _parse_money_cents(row.get("actual_paid_yuan"))
            if "actual_paid_yuan" in provided
            else (existing.actual_paid_cents if existing else 0)
        ),
        "amount_status": amount_status,
        "notes": (
            _limited_text(row.get("notes"), "备注", 20000)
            if "notes" in provided
            else (existing.notes if existing else None)
        ),
        "products": products,
    }


def _stable_import_keys(kind: str, values: dict) -> list[tuple[str, ...]]:
    if kind == "creators":
        platform = values["platform"]
        keys = []
        if values.get("_existing_creator_id") is not None:
            keys.append(("creator_id", str(values["_existing_creator_id"])))
        normalized_uid = normalize_platform_uid(values.get("platform_uid"))
        if normalized_uid:
            keys.append(("creator_uid", platform, normalized_uid))
        handle = normalize_douyin_handle(values.get("douyin_handle"))
        if handle:
            keys.append(("creator_handle", platform, handle.casefold()))
        return keys

    keys = [("collaboration_code", values["internal_code"])]
    if values.get("external_record_id"):
        keys.append(
            (
                "collaboration_external",
                values["source_type"],
                values["external_record_id"],
            )
        )
    return keys


def _validate_rows(db: Session, batch: CreatorImportBatch, mapping: dict[str, str]):
    valid = []
    errors = []
    seen_keys: dict[tuple[str, ...], int] = {}
    for row_number, row in _mapped_rows(batch, mapping):
        try:
            normalized = (
                _normalize_creator_row(db, row)
                if batch.kind == "creators"
                else _normalize_collaboration_row(db, row, batch.source_type)
            )
            stable_keys = _stable_import_keys(batch.kind, normalized)
            duplicate_row = next(
                (seen_keys[key] for key in stable_keys if key in seen_keys),
                None,
            )
            if duplicate_row is not None:
                label = "达人身份" if batch.kind == "creators" else "合作稳定编号"
                raise ValueError(f"与第{duplicate_row}行{label}重复")
            for key in stable_keys:
                seen_keys[key] = row_number
            valid.append((row_number, normalized))
        except ValueError as exc:
            errors.append({"row": row_number, "message": str(exc)[:500]})
    return valid, errors


def _result(batch: CreatorImportBatch, *, imported_count: int | None = None, skipped_count: int | None = None) -> dict:
    return {
        "token": batch.token,
        "status": batch.status,
        "row_count": batch.row_count,
        "imported_count": batch.imported_count if imported_count is None else imported_count,
        "updated_count": batch.updated_count,
        "skipped_count": batch.skipped_count if skipped_count is None else skipped_count,
        "error_count": batch.error_count,
        "errors": batch.errors or [],
    }


def _validate_mapping(batch: CreatorImportBatch, mapping: dict[str, str]) -> dict[str, str]:
    headers, _ = _read_workbook(_verify_staged_file(batch))
    allowed = CREATOR_CANONICAL_FIELDS if batch.kind == "creators" else COLLABORATION_CANONICAL_FIELDS
    normalized = {str(header): str(field) for header, field in mapping.items() if field}
    if any(header not in headers for header in normalized):
        raise HTTPException(status_code=422, detail="字段映射包含不存在的表头")
    if any(field not in allowed for field in normalized.values()):
        raise HTTPException(status_code=422, detail="字段映射包含不支持的目标字段")
    if len(normalized.values()) != len(set(normalized.values())):
        raise HTTPException(status_code=422, detail="多个表头不能映射到同一目标字段")
    return normalized


def validate_import(db: Session, token: str, mapping: dict[str, str]) -> dict:
    batch = _batch(db, token)
    if batch.status == "committed":
        raise HTTPException(status_code=409, detail="导入任务已经提交")
    mapping = _validate_mapping(batch, mapping)
    valid, errors = _validate_rows(db, batch, mapping)
    batch.mapping = mapping
    batch.errors = errors
    batch.error_count = len(errors)
    batch.status = "validated"
    db.commit()
    return _result(batch, imported_count=len(valid), skipped_count=len(errors))


def _upsert_creator(db: Session, values: dict) -> bool:
    creator = _creator_by_identity(
        db,
        values["platform"],
        values["platform_uid"],
        values["douyin_handle"],
    )
    created = creator is None
    provided = values.get("_provided_fields", set())
    if creator is None:
        handle = normalize_douyin_handle(values["douyin_handle"])
        creator = Creator(
            platform=values["platform"],
            platform_uid=values["platform_uid"],
            platform_uid_normalized=normalize_platform_uid(values["platform_uid"]),
            douyin_handle=handle,
            douyin_handle_normalized=handle.casefold() if handle else None,
            nickname=values["nickname"],
        )
        db.add(creator)
        db.flush()
    for field in (
        "nickname",
        "homepage_url",
        "mcn_name",
        "owner_id",
        "stage",
        "tags",
        "contact_name",
        "contact_phone",
        "wechat_id",
    ):
        value = values.get(field)
        if (created or field in provided) and value not in (None, ""):
            setattr(creator, field, value)
    if values.get("platform_uid") and (created or "platform_uid" in provided):
        creator.platform_uid = values["platform_uid"]
        creator.platform_uid_normalized = normalize_platform_uid(values["platform_uid"])
    if values.get("douyin_handle") and (created or "douyin_handle" in provided):
        creator.douyin_handle = values["douyin_handle"]
        creator.douyin_handle_normalized = values["douyin_handle"].casefold()
    portrait_values = values["portrait"]
    portrait = creator.portrait or CreatorPortrait(creator=creator)
    for field, value in portrait_values.items():
        if (created or field in provided) and value not in (None, "", [], {}):
            setattr(portrait, field, value)
    db.add(portrait)
    address = values["address"]
    if all(address.get(field) for field in ("recipient_name", "phone", "province", "city", "detail")):
        default = next((item for item in creator.addresses if item.is_default), None)
        target = default or CreatorAddress(creator=creator, is_default=True)
        for field, value in address.items():
            setattr(target, field, value)
        db.add(target)
    return created


def _upsert_collaboration(db: Session, values: dict) -> bool:
    by_external = None
    if values["external_record_id"]:
        by_external = db.query(CreatorCollaboration).filter(
            CreatorCollaboration.source_type == values["source_type"],
            CreatorCollaboration.external_record_id == values["external_record_id"],
        ).first()
    by_code = db.query(CreatorCollaboration).filter(
        CreatorCollaboration.internal_code == values["internal_code"]
    ).first()
    if by_external and by_code and by_external.id != by_code.id:
        raise HTTPException(status_code=409, detail="合作身份冲突，请重新验证导入文件")
    item = by_external or by_code
    created = item is None
    provided = values.get("_provided_fields", set())
    if item is None:
        item = CreatorCollaboration(
            creator_id=values["creator_id"],
            internal_code=values["internal_code"],
        )
        db.add(item)
    field_sources = {
        "owner_id": "owner_name",
        "external_record_id": "external_record_id",
        "collaboration_type": "collaboration_type",
        "collaboration_date": "collaboration_date",
        "status": "status",
        "actual_paid_cents": "actual_paid_yuan",
        "amount_status": "amount_status",
        "notes": "notes",
    }
    if created or {"creator_platform_uid", "creator_douyin_handle"} & provided:
        item.creator_id = values["creator_id"]
    if created:
        item.source_type = values["source_type"]
    for field, source_field in field_sources.items():
        if created or source_field in provided:
            setattr(item, field, values[field])
    if values["products"] is not None:
        item.products.clear()
        if item.id is not None:
            db.flush()
        for product in values["products"]:
            item.products.append(CreatorCollaborationProduct(**product))
    return created


def commit_import(db: Session, token: str) -> dict:
    batch = _batch(db, token)
    if batch.status != "validated":
        raise HTTPException(status_code=409, detail="请先完成导入验证")
    duplicate = db.query(CreatorImportBatch).filter(
        CreatorImportBatch.id != batch.id,
        CreatorImportBatch.kind == batch.kind,
        CreatorImportBatch.file_sha256 == batch.file_sha256,
        CreatorImportBatch.status == "committed",
    ).first()
    if duplicate:
        raise HTTPException(status_code=409, detail="该文件已经导入")
    _verify_staged_file(batch)
    valid, errors = _validate_rows(db, batch, batch.mapping or {})
    imported = 0
    updated = 0
    try:
        for _, row in valid:
            created = _upsert_creator(db, row) if batch.kind == "creators" else _upsert_collaboration(db, row)
            if created:
                imported += 1
            else:
                updated += 1
        batch.errors = errors
        batch.error_count = len(errors)
        batch.imported_count = imported
        batch.updated_count = updated
        batch.skipped_count = len(errors)
        batch.status = "committed"
        batch.committed_at = datetime.now(UTC).replace(tzinfo=None)
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        duplicate = db.query(CreatorImportBatch).filter(
            CreatorImportBatch.kind == batch.kind,
            CreatorImportBatch.file_sha256 == batch.file_sha256,
            CreatorImportBatch.status == "committed",
        ).first()
        if duplicate:
            raise HTTPException(status_code=409, detail="该文件已经导入") from exc
        raise HTTPException(status_code=409, detail="导入数据与现有记录冲突") from exc
    except Exception:
        db.rollback()
        raise
    _safe_unlink(CREATOR_IMPORT_DIR / f"{batch.token}.xlsx")
    return _result(batch)


def error_report(db: Session, token: str) -> bytes:
    batch = _batch(db, token)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "导入错误"
    sheet.append(["行号", "错误说明"])
    for error in batch.errors or []:
        sheet.append([f"第{error.get('row')}行", error.get("message")])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _filter_export_creator_query(
    query,
    *,
    creator_id: int | None,
    stage: str | None,
    owner_id: int | None,
    search: str | None,
    category: str | None,
    follower_tier: str | None,
):
    query = query.filter(Creator.archived_at.is_(None))
    if creator_id:
        query = query.filter(Creator.id == creator_id)
    if stage:
        query = query.filter(Creator.stage == stage)
    if owner_id:
        query = query.filter(Creator.owner_id == owner_id)
    if search:
        term = f"%{search.strip()}%"
        query = query.filter(
            or_(
                Creator.nickname.ilike(term),
                Creator.douyin_handle.ilike(term),
                Creator.platform_uid.ilike(term),
                Creator.mcn_name.ilike(term),
            )
        )
    if category or follower_tier:
        query = query.join(
            CreatorPortrait, CreatorPortrait.creator_id == Creator.id
        )
    if category:
        category_values = func.json_each(
            CreatorPortrait.primary_categories
        ).table_valued("key", "value").alias("export_creator_categories")
        query = query.filter(
            exists(
                select(1)
                .select_from(category_values)
                .where(category_values.c.value == category)
            )
        )
    if follower_tier:
        count = CreatorPortrait.follower_count
        ranges = {
            "under_10k": count < 10_000,
            "10k_100k": (count >= 10_000) & (count < 100_000),
            "100k_500k": (count >= 100_000) & (count < 500_000),
            "500k_1m": (count >= 500_000) & (count < 1_000_000),
            "1m_plus": count >= 1_000_000,
        }
        condition = ranges.get(follower_tier)
        if condition is None:
            raise HTTPException(status_code=422, detail="粉丝量级无效")
        query = query.filter(condition)
    return query


def _excel_safe_export_value(value: Any) -> Any:
    if isinstance(value, str):
        value = ILLEGAL_CHARACTERS_RE.sub("", value)
        if value.lstrip().startswith(("=", "+", "-", "@")):
            return "'" + value
    return value


def _append_excel_row(sheet, values: list[Any]) -> None:
    sheet.append([_excel_safe_export_value(value) for value in values])


def _creator_export_rows(
    db: Session,
    *,
    creator_id: int | None,
    stage: str | None,
    owner_id: int | None,
    search: str | None,
    category: str | None,
    follower_tier: str | None,
):
    filtered_ids = _filter_export_creator_query(
        db.query(Creator.id.label("creator_id")),
        creator_id=creator_id,
        stage=stage,
        owner_id=owner_id,
        search=search,
        category=category,
        follower_tier=follower_tier,
    ).subquery()
    paid_totals = (
        db.query(
            CreatorCollaboration.creator_id.label("creator_id"),
            func.sum(CreatorCollaboration.actual_paid_cents).label("confirmed_paid_cents"),
        )
        .filter(
            CreatorCollaboration.amount_status == "confirmed",
            CreatorCollaboration.status != "cancelled",
        )
        .group_by(CreatorCollaboration.creator_id)
        .subquery()
    )
    return (
        db.query(
            Creator.nickname.label("nickname"),
            Creator.platform.label("platform"),
            Creator.platform_uid.label("platform_uid"),
            Creator.douyin_handle.label("douyin_handle"),
            BdMember.name.label("owner_name"),
            Creator.stage.label("stage"),
            CreatorPortrait.follower_count.label("follower_count"),
            CreatorPortrait.primary_categories.label("primary_categories"),
            Creator.tags.label("tags"),
            Creator.contact_phone.label("contact_phone"),
            func.coalesce(paid_totals.c.confirmed_paid_cents, 0).label(
                "confirmed_paid_cents"
            ),
        )
        .join(filtered_ids, filtered_ids.c.creator_id == Creator.id)
        .outerjoin(BdMember, BdMember.id == Creator.owner_id)
        .outerjoin(CreatorPortrait, CreatorPortrait.creator_id == Creator.id)
        .outerjoin(paid_totals, paid_totals.c.creator_id == Creator.id)
        .order_by(Creator.updated_at.desc(), Creator.id.desc())
        .limit(MAX_EXPORT_ROWS)
        .all()
    )


def export_workbook(
    db: Session,
    *,
    entity: str,
    creator_id: int | None = None,
    stage: str | None = None,
    owner_id: int | None = None,
    search: str | None = None,
    category: str | None = None,
    follower_tier: str | None = None,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    if entity == "creators":
        sheet.title = "达人资料"
        _append_excel_row(sheet, ["达人昵称", "平台", "官方达人ID", "抖音号", "负责人", "阶段", "粉丝数", "主营垂类", "标签", "联系电话（脱敏）", "累计实付（元）"])
        creator_rows = _creator_export_rows(
            db,
            creator_id=creator_id,
            stage=stage,
            owner_id=owner_id,
            search=search,
            category=category,
            follower_tier=follower_tier,
        )
        for item in creator_rows:
            _append_excel_row(sheet, [
                item.nickname, item.platform, item.platform_uid, item.douyin_handle,
                item.owner_name, item.stage, item.follower_count, "|".join(item.primary_categories or []),
                "|".join(item.tags or []), creator_service.mask_phone(item.contact_phone), item.confirmed_paid_cents / 100,
            ])
    elif entity == "collaborations":
        sheet.title = "合作记录"
        _append_excel_row(sheet, ["合作编号", "达人昵称", "合作形式", "合作日期", "状态", "实际支付金额（元）", "金额状态", "合作产品", "备注"])
        query = _filter_export_creator_query(
            db.query(CreatorCollaboration).join(Creator),
            creator_id=creator_id,
            stage=stage,
            owner_id=owner_id,
            search=search,
            category=category,
            follower_tier=follower_tier,
        )
        rows = (
            query.options(
                joinedload(CreatorCollaboration.creator),
                selectinload(CreatorCollaboration.products),
            )
            .order_by(CreatorCollaboration.collaboration_date.desc())
            .limit(MAX_EXPORT_ROWS)
            .all()
        )
        for item in rows:
            _append_excel_row(sheet, [
                item.internal_code, item.creator.nickname, item.collaboration_type, item.collaboration_date,
                item.status, item.actual_paid_cents / 100, item.amount_status,
                "|".join(product.product_name_snapshot for product in item.products), item.notes,
            ])
    elif entity == "sample_orders":
        sheet.title = "寄样履约"
        _append_excel_row(sheet, ["敏感信息：仅用于寄样履约，请妥善保管"])
        _append_excel_row(sheet, ["寄样单ID", "达人昵称", "状态", "收件人", "电话", "省", "市", "区", "详细地址", "产品", "快递公司", "运单号", "创建时间"])
        query = _filter_export_creator_query(
            db.query(CreatorSampleOrder).join(Creator),
            creator_id=creator_id,
            stage=stage,
            owner_id=owner_id,
            search=search,
            category=category,
            follower_tier=follower_tier,
        )
        rows = (
            query.options(
                joinedload(CreatorSampleOrder.creator),
                selectinload(CreatorSampleOrder.items),
            )
            .order_by(CreatorSampleOrder.created_at.desc())
            .limit(MAX_EXPORT_ROWS)
            .all()
        )
        for order in rows:
            _append_excel_row(sheet, [
                order.id, order.creator.nickname, order.status, order.recipient_name_snapshot,
                order.phone_snapshot, order.province_snapshot, order.city_snapshot, order.district_snapshot,
                order.address_detail_snapshot,
                "|".join(f"{item.product_name_snapshot}×{item.quantity}" for item in order.items),
                order.shipping_company, order.tracking_number, order.created_at,
            ])
    else:
        raise HTTPException(status_code=400, detail="entity 仅支持 creators、collaborations、sample_orders")
    sheet.freeze_panes = "A2" if entity != "sample_orders" else "A3"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
