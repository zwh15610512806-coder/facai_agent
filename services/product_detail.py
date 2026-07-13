"""Build product detail data from local product materials."""
from __future__ import annotations

from functools import lru_cache
from pathlib import Path
import re
from typing import Any

import import_materials


HIDDEN_SELLING_POINT_TYPE = "__hidden__"


USELESS_POINT_TYPES = {
    "资料标题",
    "产品手卡",
    "售价",
    "产品价格",
}

USELESS_CONTENT_MARKERS = (
    "打印预览尺寸",
    "代表产品：",
    "产品——手卡",
    "产品--手卡",
)


PROFILE_SECTION_DEFINITIONS = [
    ("product_info", "产品信息"),
    ("product_price", "产品价格"),
    ("product_usage", "产品用途"),
    ("usage_scenarios", "使用场景"),
    ("main_selling_points", "主要卖点"),
]

PROFILE_INFO_TYPES = {
    "产品名称",
    "规格",
    "保质期",
    "保存方式",
    "储存方式",
    "状态",
    "命名对照",
    "标准命名",
    "口味",
    "用户人群",
    "适用人群",
    "颜色体系",
    "SKU规格",
}

PROFILE_USAGE_TYPES = {
    "用途",
    "用途简述",
    "用途场景",
    "使用",
    "使用方法",
    "制作方法",
    "用量",
    "推荐用量",
    "场景用量",
    "FAQ：用途",
    "FAQ：怎么制作？",
    "FAQ：怎么使用？",
}

PROFILE_SCENARIO_TYPES = {
    "场景",
    "主要场景",
    "使用场景",
    "经营场景",
    "门店方案",
    "解决方案",
    "具体行动及结果",
}

PROFILE_SELLING_TYPES = {
    "一句话卖点",
    "产品亮点",
    "产品卖点",
    "卖点",
    "卖点排序",
    "产品价值",
    "核心亮点",
    "核心优势",
    "痛点切入",
    "差异化卖点",
    "差异化卖点 （重点）",
    "优势",
    "特质",
    "对比",
    "客户价值",
    "销售表达",
}

PROFILE_WEAK_INFO_LABELS = {
    "产品名称",
    "品类",
    "品牌",
    "产品描述",
    "资料来源",
    "命名对照",
    "标准命名",
    "规格",
    "保质期",
    "保存方式",
    "储存方式",
    "状态",
    "产品售价",
    "原价",
}

PROFILE_CATEGORY_FALLBACKS = {
    "烘焙夹心": {
        "product_usage": [
            ("用途简述", "适合用于蛋糕夹心、甜品夹层、面包夹馅或口感颗粒搭配。"),
            ("推荐搭配", "可搭配奶油、果酱、茶酱、慕斯、蛋糕胚等基础出品，增强夹层风味和咀嚼感。"),
            ("使用方法", "按门店出品需求加入夹层、撒入奶油夹心或做表面点缀，建议先小量试做确认口感比例。"),
        ],
        "usage_scenarios": [
            ("使用场景", "适合门店做夹心蛋糕上新、甜品夹层组合、增加产品口感层次时使用。"),
            ("门店方案", "适合搭配口味蛋糕、生日蛋糕、甜品杯和面包夹馅，作为快速增加产品差异化的夹心素材。"),
            ("经营场景", "适合新品测试、节日款组合、门店菜单升级和需要提升口感记忆点的产品线。"),
        ],
        "main_selling_points": [
            ("主要卖点", "帮助蛋糕和甜品增加夹心口感与层次，适合门店快速做差异化产品。"),
            ("核心亮点", "夹心颗粒类产品使用灵活，能在不大改原有配方的情况下增加口感变化。"),
            ("产品价值", "适合做风味补充和口感升级，帮助门店把基础蛋糕做出更清晰的卖点表达。"),
        ],
    },
    "烘焙装饰": {
        "product_usage": [
            ("用途简述", "适合用于蛋糕装饰、表面点缀、造型出单和成品陈列。"),
            ("推荐搭配", "可搭配奶油蛋糕、翻糖造型、主题插件、节日装饰和门店陈列款。"),
            ("使用方法", "根据蛋糕主题和色系选择装饰位置，控制装饰密度，避免遮挡主体造型。"),
        ],
        "usage_scenarios": [
            ("使用场景", "适合生日蛋糕、节日主题蛋糕、门店陈列款和社交平台出片款。"),
            ("门店方案", "适合快速提升成品视觉完整度，帮助门店做主题化、套系化的蛋糕出样。"),
            ("经营场景", "适合高频定制款、节日活动款、橱窗展示款和需要提高照片传播效果的产品。"),
        ],
        "main_selling_points": [
            ("主要卖点", "提升蛋糕外观完成度和出片效果，降低门店装饰搭配成本。"),
            ("核心亮点", "可直接服务成品视觉表达，让普通蛋糕更容易形成主题感和仪式感。"),
            ("产品价值", "帮助门店提升出样效率和视觉溢价空间，适合批量化做陈列款。"),
        ],
    },
    "烘焙调味": {
        "product_usage": [
            ("用途简述", "适合用于奶油、慕斯、淋面、饮品或蛋糕风味调配。"),
            ("推荐搭配", "可搭配淡奶油、慕斯、巴斯克、蛋糕卷、饮品和甜品杯等常见门店出品。"),
            ("使用方法", "按目标风味少量多次加入并试味，兼顾颜色、甜度和整体风味平衡。"),
        ],
        "usage_scenarios": [
            ("使用场景", "适合门店上新风味蛋糕、饮品甜品搭配、季节限定口味开发。"),
            ("门店方案", "适合做口味蛋糕、下午茶甜品、饮品联名款和小红书风味趋势款。"),
            ("经营场景", "适合新品菜单更新、爆款口味复刻、节日限定和门店提高复购的风味升级。"),
        ],
        "main_selling_points": [
            ("主要卖点", "帮助门店快速建立稳定风味，适合做口味差异化和复购款。"),
            ("核心亮点", "减少从零熬制或调配风味的时间，让门店更快完成新品测试。"),
            ("产品价值", "能把基础奶油、蛋糕和饮品做出口味记忆点，适合持续扩展菜单。"),
        ],
    },
    "烘焙调色": {
        "product_usage": [
            ("用途简述", "适合用于奶油、蛋糕胚、淋面、翻糖等烘焙调色场景。"),
            ("推荐搭配", "可搭配奶油调色、蛋糕胚上色、翻糖造型、淋面调色和主题甜品色彩还原。"),
            ("使用方法", "建议少量多次添加，先做小样确认色相和深浅，再批量用于正式出品。"),
        ],
        "usage_scenarios": [
            ("使用场景", "适合网红色系还原、主题蛋糕配色、节日款和定制蛋糕调色。"),
            ("门店方案", "适合门店做主题色系蛋糕、节日限定色、儿童款和社交平台热门颜色复刻。"),
            ("经营场景", "适合接定制单、做视觉爆款、统一门店色系和提升成品照片识别度。"),
        ],
        "main_selling_points": [
            ("主要卖点", "帮助门店更稳定地完成颜色表达，提升主题蛋糕和定制款出品效率。"),
            ("核心亮点", "调色路径清晰，适合高频做颜色还原和批量出单。"),
            ("产品价值", "减少反复试色成本，让门店更容易把图片色、主题色转化为稳定成品。"),
        ],
    },
    "烘焙配件": {
        "product_usage": [
            ("用途简述", "适合用于蛋糕打包、门店出单、展示陈列和随餐配套。"),
            ("推荐搭配", "可搭配生日蛋糕、甜品礼盒、外带配送、门店陈列和活动套餐。"),
            ("使用方法", "按产品规格和出单场景搭配使用，保证包装、展示和交付体验完整。"),
        ],
        "usage_scenarios": [
            ("使用场景", "适合门店外带配送、生日蛋糕交付、节日礼盒和仪式感套餐。"),
            ("门店方案", "适合补齐蛋糕交付链路，让消费者从取餐、分享、食用到拍照都有完整体验。"),
            ("经营场景", "适合门店提升客单体验、节日套餐包装、批量出单和外卖配送场景。"),
        ],
        "main_selling_points": [
            ("主要卖点", "补齐交付和使用体验，提升门店成品包装的完整度。"),
            ("核心亮点", "配件类产品能直接影响消费者收到产品后的仪式感和便利度。"),
            ("产品价值", "适合门店做标准化交付，减少临时搭配成本，提升成品专业度。"),
        ],
    },
}

PROFILE_PRODUCT_FALLBACKS = {
    "开心果碎": {
        "product_usage": [
            ("用途简述", "适合用于蛋糕夹心、甜品夹层、面包夹馅和开心果风味口感颗粒搭配。"),
            ("推荐搭配", "可搭配奶油、慕斯、巴斯克、蛋糕胚、果酱或茶酱，增加坚果风味和颗粒口感。"),
            ("使用方法", "可撒入夹层、拌入奶油夹心或用于表面点缀，建议根据口感需求控制添加量。"),
        ],
        "usage_scenarios": [
            ("使用场景", "适合开心果风味蛋糕、夹心蛋糕上新、甜品杯夹层和面包夹馅。"),
            ("门店方案", "适合作为开心果风味产品的口感补充，配合开心果酱、奶油或蛋糕胚形成更完整的风味表达。"),
            ("经营场景", "适合门店做风味升级、新品试做、节日限定款和需要提升口感记忆点的产品。"),
        ],
        "main_selling_points": [
            ("主要卖点", "为蛋糕和甜品增加开心果风味颗粒感，让夹心层次更清晰。"),
            ("核心亮点", "作为夹心颗粒使用灵活，可用于夹层、拌料和表面点缀，帮助门店快速做风味差异化。"),
            ("产品价值", "适合把基础蛋糕升级成带坚果风味和咀嚼记忆点的门店新品。"),
        ],
    },
    "巧克力脆馅": {
        "product_usage": [
            ("用途简述", "适合用于蛋糕夹心、甜品夹层和巧克力风味口感颗粒搭配。"),
            ("推荐搭配", "可搭配奶油、慕斯、蛋糕胚、果酱或巧克力风味甜品，增加夹心层次。"),
            ("使用方法", "可撒入夹层或拌入夹心体系，建议根据成品甜度和脆感需求调整添加量。"),
        ],
        "usage_scenarios": [
            ("使用场景", "适合巧克力风味蛋糕、夹心蛋糕上新、甜品夹层组合和提升口感层次。"),
            ("门店方案", "适合门店开发巧克力风味夹心款、儿童款、生日蛋糕和甜品杯组合。"),
            ("经营场景", "适合需要增加脆感记忆点、丰富夹心层次和做口味差异化的产品线。"),
        ],
        "main_selling_points": [
            ("主要卖点", "为蛋糕和甜品增加巧克力脆感夹心元素，帮助门店做出更有记忆点的夹心口感。"),
            ("核心亮点", "脆感夹心能强化咀嚼体验，适合与奶油、蛋糕胚和甜品夹层组合。"),
            ("产品价值", "帮助基础蛋糕快速升级成带口感层次的夹心款，便于门店做新品表达。"),
        ],
    },
}

PROFILE_FALLBACK_PRIORITY_BASES = {
    "product_usage": 9000,
    "usage_scenarios": 9100,
    "main_selling_points": 9200,
}

PROFILE_MIN_DETAIL_ITEMS = {
    "product_usage": 3,
    "usage_scenarios": 3,
    "main_selling_points": 3,
}


PRODUCT_CARD_POINT_LABELS = {
    "一句话卖点",
    "产品名称",
    "规格",
    "售价",
    "一件85折",
    "五件75折",
    "展会一件69折",
    "口味",
    "用户人群",
    "适用人群",
    "保质期",
    "保存方式",
    "储存方式",
    "用途",
    "用途简述",
    "使用场景",
    "用途场景",
    "应用场景",
    "场景用量",
    "推荐用量",
    "使用方法",
    "制作方法",
    "用量",
    "产品亮点",
    "产品卖点",
    "卖点",
    "卖点排序",
    "产品价值",
    "核心优势",
    "痛点切入",
    "差异化卖点",
    "差异化卖点 （重点）",
    "与拉线膏区别",
    "推荐搭配",
    "对比",
}

PRODUCT_CARD_SKIP_SHEETS = {
    "常见问题",
    "解决方案",
    "WpsReserved_CellImgList",
}

PRODUCT_CARD_SHEET_ALIASES = {
    "果泥": "夹心果泥",
    "果馅（多肉）": "多肉果酱",
    "芋泥": "夹心芋泥",
    "白色翻糖": "白色翻糖膏",
    "彩色翻糖": "彩色翻糖膏",
    "翻糖片": "翻糖压片",
    "拉线膏": "彩色拉线膏",
    "红丝绒液": "红丝绒",
    "薄荷糖浆": "调味糖浆",
    "脆皮酱": "巧克力脆皮酱",
    "2元盒装": "盒装刀叉",
    "2.5元盒装": "盒装刀叉",
    "5元盒装刀叉": "盒装刀叉",
    "1.1浆纸盘": "盒装刀叉",
}

PRODUCT_CARD_SOLUTION_PRODUCTS = {
    "夹心类": [
        "夹心珠",
        "夹心脆",
        "奶冻粉",
        "Q弹奶冻粉",
        "Q 弹奶冻粉",
        "晶冻粉",
        "慕斯粉（液）",
        "布蕾粉",
        "夹心芋泥",
        "夹心果泥",
    ],
    "口味蛋糕": ["调味果酱", "茶酱", "开心果酱", "夹心果泥", "多肉果酱", "焦糖酱"],
    "造型蛋糕": ["白色翻糖膏", "彩色翻糖膏", "翻糖压片", "手绘膏", "彩色拉线膏", "拉线膏"],
    "网红颜色还原蛋糕": ["水性色素", "油性色素", "果蔬色素", "浅柔色素", "浅系色素", "色粉盘", "红丝绒"],
    "餐盘仪式感": ["盒装刀叉", "刀叉", "纸盘", "1.1浆纸盘", "5元盒装刀叉", "2元盒装", "2.5元盒装"],
}

RISKY_2026_PHRASES = (
    "待核验",
    "全国销量第一",
    "全网销量领先",
    "全网销量 TOP1",
    "全网TOP1",
    "TOP1",
    "安全健康",
    "安全放心",
    "没有任何影响",
    "无任何影响",
    "0脂肪酸",
    "无反式脂肪酸",
    "没有反式脂肪酸",
    "无添加剂",
    "儿童也可以放心",
    "宝妈也能放心",
    "健康+好吃",
    "更健康",
    "可提高10%",
    "轻松溢价",
    "卖高价",
)


def build_material_product_detail(
    product_name: str,
    root: Path | str | None = None,
) -> dict[str, Any]:
    root_path = Path(root) if root is not None else Path(__file__).resolve().parents[1]
    try:
        paths = import_materials.get_material_paths(root_path)
    except (OSError, FileNotFoundError):
        return {
            "source_name": _base_name(product_name),
            "manual_source": "",
            "knowledge_sources": [],
            "selling_points": [],
            "sku_prices": [],
        }
    products = _load_material_products(str(root_path.resolve()))
    source = _find_material_product(product_name, products)
    product_names = [product_name]
    if source:
        product_names.append(source.name)

    selling_points = []
    if source:
        selling_points = [
            {
                "point_type": point.get("point_type", "卖点"),
                "content": point.get("content", ""),
                "priority": point.get("priority", index),
            }
            for index, point in enumerate(source.selling_points, start=1)
        ]
    knowledge_points, knowledge_sources = _find_2026_knowledge_points(
        str(root_path.resolve()),
        product_names,
    )
    selling_points = _clean_merge_points(selling_points, knowledge_points)
    manual_source = _manual_source_name(str(paths.product_manual_md))
    sources = [manual_source, *knowledge_sources]

    return {
        "source_name": source.name if source else _base_name(product_name),
        "manual_source": manual_source,
        "knowledge_sources": sources,
        "selling_points": selling_points,
        "sku_prices": build_sku_prices(product_name, root_path),
    }


def build_sku_prices(
    product_name: str,
    root: Path | str | None = None,
) -> list[dict[str, Any]]:
    root_path = Path(root) if root is not None else Path(__file__).resolve().parents[1]
    paths = import_materials.get_material_paths(root_path)

    rule_name = _resolve_price_rule_name(product_name)
    rule = import_materials.PRICE_RULES.get(rule_name)
    if not rule:
        return []

    override = rule.get("override")
    has_price_filters = any(
        rule.get(key) for key in ("products", "specs", "exclude_specs", "knife_price")
    )
    if isinstance(override, (int, float)) and override > 0 and not has_price_filters:
        return [_override_sku(rule_name, float(override))]

    if rule.get("knife_price"):
        rows = (
            import_materials._parse_knife_price_rows(
                paths.knife_price_xlsx,
                str(rule["knife_price"]),
            )
            if paths.knife_price_xlsx is not None else []
        )
    elif paths.price_system_xlsx is not None:
        rows = import_materials._parse_price_system_rows(paths.price_system_xlsx)
    else:
        rows = []

    matched_rows = [
        row for row in rows
        if import_materials._row_matches_price_rule(row, rule)
    ]

    skus = [_sku_from_price_row(row) for row in matched_rows]
    if skus:
        return skus

    if isinstance(override, (int, float)) and override > 0:
        return [_override_sku(rule_name, float(override))]
    return []


@lru_cache(maxsize=8)
def _manual_source_name(path_text: str) -> str:
    path = Path(path_text)
    try:
        text = import_materials._read_text(path)
    except OSError:
        return path.name
    for line in text.splitlines()[:20]:
        match = re.match(r"^#\s+(.+?)\s*$", line)
        if match and "产品手册" in match.group(1):
            return f"{match.group(1)}.md"
    return path.name


def build_product_detail_payload(product: Any, root: Path | str | None = None) -> dict[str, Any]:
    material = build_material_product_detail(product.name, root)
    db_points = []
    hidden_priorities: set[int] = set()
    for point in sorted(product.selling_points, key=lambda item: item.priority):
        if point.point_type == HIDDEN_SELLING_POINT_TYPE:
            try:
                hidden_priorities.add(int(point.priority))
            except (TypeError, ValueError):
                pass
            continue
        if _is_useless_selling_point(point.point_type, point.content):
            continue
        db_points.append({
            "id": point.id,
            "product_id": point.product_id,
            "point_type": point.point_type,
            "content": _clean_selling_point_content(point.content),
            "priority": point.priority,
        })
    material_points = material["selling_points"]
    if (db_points or hidden_priorities) and material_points:
        selling_points = list(material_points)
        for priority in hidden_priorities:
            index = priority - 1
            if 0 <= index < len(selling_points):
                selling_points[index] = None
        for offset, point in enumerate(db_points):
            try:
                index = max(int(point.get("priority") or (offset + 1)) - 1, 0)
            except (TypeError, ValueError):
                index = offset
            if index + 1 in hidden_priorities:
                continue
            if index < len(selling_points):
                selling_points[index] = point
            else:
                selling_points.append(point)
        selling_points = [point for point in selling_points if point is not None]
    else:
        selling_points = db_points or material_points
    selling_points = _clean_merge_points(selling_points, [])

    detail = {
        "id": product.id,
        "name": product.name,
        "category": product.category,
        "price": product.price,
        "original_price": product.original_price,
        "commission_rate": product.commission_rate,
        "brand": product.brand,
        "description": product.description,
        "info_file": product.info_file,
        "pending_fields": product.pending_fields or [],
        "status": product.status,
        "source_name": material["source_name"],
        "manual_source": material["manual_source"],
        "knowledge_sources": material["knowledge_sources"],
        "selling_points": selling_points,
        "sku_prices": material["sku_prices"],
        "hidden_selling_point_priorities": sorted(hidden_priorities),
    }
    detail["profile_sections"] = build_profile_sections(detail)
    return detail


def build_profile_sections(detail: dict[str, Any]) -> list[dict[str, Any]]:
    """Build the five fixed product-detail sections shown in the product workspace."""
    points = detail.get("selling_points") or []
    point_buckets = _bucket_profile_points(points)
    sources = _profile_sources(detail)

    info_items = [
        _profile_item("产品名称", detail.get("name")),
        _profile_item("品类", detail.get("category")),
    ]
    if detail.get("brand"):
        info_items.append(_profile_item("品牌", detail.get("brand")))
    if detail.get("description"):
        info_items.append(_profile_item("产品描述", detail.get("description")))
    info_items.extend(point_buckets["product_info"])
    if sources:
        info_items.append({
            "label": "资料来源",
            "content": "、".join(sources),
            "values": sources,
            "source": "",
        })
    info_items = _dedupe_profile_items(info_items)

    price_skus = [_profile_sku_price(sku) for sku in detail.get("sku_prices") or []]
    price_items = []
    if detail.get("price") is not None:
        item = _profile_item("产品售价", f"¥{_profile_price(detail.get('price'))}")
        if item:
            price_items.append({**item, "field": "price", "editable": True})
    if detail.get("original_price") is not None:
        item = _profile_item("原价", f"¥{_profile_price(detail.get('original_price'))}")
        if item:
            price_items.append({**item, "field": "original_price", "editable": True})
    price_items.extend(point_buckets["product_price"])

    sections_by_id = {
        "product_info": {"items": info_items, "sku_prices": []},
        "product_price": {"items": _dedupe_profile_items(price_items), "sku_prices": price_skus},
        "product_usage": {"items": _dedupe_profile_items(point_buckets["product_usage"]), "sku_prices": []},
        "usage_scenarios": {"items": _dedupe_profile_items(point_buckets["usage_scenarios"]), "sku_prices": []},
        "main_selling_points": {"items": _dedupe_profile_items(point_buckets["main_selling_points"]), "sku_prices": []},
    }
    _fill_empty_profile_sections(detail, sections_by_id)

    sections = []
    for section_id, title in PROFILE_SECTION_DEFINITIONS:
        section = sections_by_id[section_id]
        sections.append({
            "id": section_id,
            "title": title,
            "items": section["items"],
            "sku_prices": section["sku_prices"],
        })
    return sections


def _fill_empty_profile_sections(
    detail: dict[str, Any],
    sections_by_id: dict[str, dict[str, list[dict[str, Any]]]],
) -> None:
    for section_id in ("product_usage", "usage_scenarios"):
        sections_by_id[section_id]["items"] = _with_minimum_profile_items(
            detail,
            section_id,
            sections_by_id[section_id]["items"],
        )

    selling_items = sections_by_id["main_selling_points"]["items"]
    if not _has_substantive_profile_items(selling_items):
        selling_items = []
    sections_by_id["main_selling_points"]["items"] = _with_minimum_profile_items(
        detail,
        "main_selling_points",
        selling_items,
    )


def _with_minimum_profile_items(
    detail: dict[str, Any],
    section_id: str,
    items: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    min_count = PROFILE_MIN_DETAIL_ITEMS.get(section_id, 0)
    current = _dedupe_profile_items(items)
    if len(current) >= min_count:
        return current
    return _dedupe_profile_items([
        *current,
        *_visible_fallback_profile_items(detail, section_id),
    ])


def _visible_fallback_profile_items(detail: dict[str, Any], section_id: str) -> list[dict[str, Any]]:
    hidden_priorities = {
        int(priority)
        for priority in detail.get("hidden_selling_point_priorities", [])
        if str(priority).isdigit()
    }
    return [
        item for item in _fallback_profile_items(detail, section_id)
        if int(item.get("priority") or 0) not in hidden_priorities
    ]


def _fallback_profile_items(detail: dict[str, Any], section_id: str) -> list[dict[str, Any]]:
    product_name = str(detail.get("name") or "")
    product_specs = PROFILE_PRODUCT_FALLBACKS.get(product_name, {}).get(section_id)
    base_priority = PROFILE_FALLBACK_PRIORITY_BASES.get(section_id, 9900)
    if product_specs:
        return [
            _fallback_profile_item(label, content, base_priority + index)
            for index, (label, content) in enumerate(product_specs, start=1)
        ]

    category_key = _profile_category_key(detail)
    category_specs = PROFILE_CATEGORY_FALLBACKS.get(category_key, {}).get(section_id, [])
    return [
        _fallback_profile_item(label, content, base_priority + index)
        for index, (label, content) in enumerate(category_specs, start=1)
    ]


def _fallback_profile_item(label: str, content: str, priority: int) -> dict[str, Any] | None:
    item = _profile_item(label, content)
    if not item:
        return None
    item.update({
        "priority": priority,
        "generated": True,
        "editable": True,
    })
    return item


def _profile_category_key(detail: dict[str, Any]) -> str:
    category = str(detail.get("category") or "")
    name = str(detail.get("name") or "")
    if category in PROFILE_CATEGORY_FALLBACKS:
        return category
    if "夹心" in category or "夹心" in name or any(part in name for part in ("脆馅", "脆珠", "薄脆")):
        return "烘焙夹心"
    if "配件" in category or any(part in name for part in ("刀叉", "纸盘", "盒装")):
        return "烘焙配件"
    if "调色" in category or any(part in name for part in ("色素", "色粉", "红丝绒")):
        return "烘焙调色"
    if "装饰" in category or any(part in name for part in ("翻糖", "拉线", "糖珠", "插件")):
        return "烘焙装饰"
    if "调味" in category or any(part in name for part in ("果酱", "茶酱", "糖浆", "香草", "焦糖", "开心果酱")):
        return "烘焙调味"
    return category


def _has_substantive_profile_items(items: list[dict[str, Any]]) -> bool:
    return any(not _is_weak_profile_item(item) for item in items)


def _is_weak_profile_item(item: dict[str, Any]) -> bool:
    label = import_materials._clean_markdown(item.get("label", ""))
    return label in PROFILE_WEAK_INFO_LABELS or label in PROFILE_INFO_TYPES


def _bucket_profile_points(points: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    buckets = {
        "product_info": [],
        "product_price": [],
        "product_usage": [],
        "usage_scenarios": [],
        "main_selling_points": [],
    }
    for point in points:
        point_type = import_materials._clean_markdown(point.get("point_type", "卖点"))
        content = _clean_selling_point_content(point.get("content", ""))
        if not point_type or not content or _is_useless_selling_point(point_type, content):
            continue
        item = _point_to_profile_item({**point, "point_type": point_type, "content": content})
        section_id = _profile_section_for_point(point_type)
        if section_id:
            buckets[section_id].append(item)
        elif point_type.startswith("FAQ"):
            buckets["main_selling_points"].append(item)
    return buckets


def _profile_section_for_point(point_type: str) -> str:
    if point_type in PROFILE_INFO_TYPES:
        return "product_info"
    if point_type in {"售价", "产品价格", "SKU售价", "一件85折", "五件75折", "展会一件69折"} or "价格" in point_type or "折" in point_type:
        return "product_price"
    if point_type in PROFILE_USAGE_TYPES or "用法" in point_type or "制作" in point_type or "用量" in point_type:
        return "product_usage"
    if point_type in PROFILE_SCENARIO_TYPES or "场景" in point_type or "方案" in point_type:
        return "usage_scenarios"
    if point_type in PROFILE_SELLING_TYPES or "卖点" in point_type or "亮点" in point_type or "优势" in point_type or "价值" in point_type:
        return "main_selling_points"
    return "main_selling_points"


def _point_to_profile_item(point: dict[str, Any]) -> dict[str, Any]:
    return {
        "label": import_materials._clean_markdown(point.get("point_type", "卖点")),
        "content": _clean_selling_point_content(point.get("content", "")),
        "source": point.get("source", ""),
        "point_id": point.get("id"),
        "product_id": point.get("product_id"),
        "priority": point.get("priority"),
        "editable": bool(point.get("id")),
    }


def _profile_item(label: str, content: Any) -> dict[str, Any] | None:
    text = _clean_selling_point_content(str(content or ""))
    if not text:
        return None
    return {"label": label, "content": text, "source": ""}


def _dedupe_profile_items(items: list[dict[str, Any] | None]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for item in items:
        if not item:
            continue
        label = import_materials._clean_markdown(item.get("label", ""))
        content = _clean_selling_point_content(item.get("content", ""))
        if not label or not content or _is_useless_selling_point(label, content):
            continue
        key = (_selling_point_type_key(label), _selling_point_content_key(content))
        if key in seen:
            continue
        seen.add(key)
        out.append({**item, "label": label, "content": content})
    return out


def _profile_sources(detail: dict[str, Any]) -> list[str]:
    sources = []
    if detail.get("manual_source"):
        sources.append(detail["manual_source"])
    sources.extend(detail.get("knowledge_sources") or [])
    seen: set[str] = set()
    out: list[str] = []
    for source in sources:
        source_text = import_materials._clean_markdown(str(source or ""))
        if not source_text or source_text in seen:
            continue
        seen.add(source_text)
        out.append(source_text)
    return out


def _profile_sku_price(sku: dict[str, Any]) -> dict[str, Any]:
    return {
        "product": sku.get("product", ""),
        "spec": sku.get("spec", ""),
        "name": " · ".join(part for part in [sku.get("product"), sku.get("spec")] if part) or "默认规格",
        "price": _profile_price(sku.get("price")),
        "daily_price": _profile_price(sku.get("daily_price")),
        "count": sku.get("count"),
        "category": sku.get("category", ""),
        "activity_prices": [_profile_activity_price(activity) for activity in sku.get("activity_prices", [])],
    }


def _profile_activity_price(activity: dict[str, Any]) -> dict[str, Any]:
    price = activity.get("final_price")
    if price is None:
        price = activity.get("activity_price")
    if price is None:
        price = activity.get("tag_price")
    meta = []
    if activity.get("activity_price") is not None and _profile_price(activity.get("activity_price")) != _profile_price(price):
        meta.append(f"活动价 ¥{_profile_price(activity.get('activity_price'))}")
    if activity.get("unit_price") is not None and _profile_price(activity.get("unit_price")) != _profile_price(price):
        meta.append(f"均价 ¥{_profile_price(activity.get('unit_price'))}")
    if activity.get("discount"):
        meta.append(str(activity["discount"]))
    if activity.get("coupon") and activity.get("coupon") != "0":
        meta.append(f"券 {activity['coupon']}")
    if activity.get("single_activity"):
        meta.append(str(activity["single_activity"]))
    return {
        "mechanism": activity.get("mechanism") or "活动价",
        "price": _profile_price(price),
        "meta": " / ".join(meta),
    }


def _profile_price(value: Any) -> str:
    if value is None or value == "":
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    if number.is_integer():
        return str(int(number))
    return f"{number:.2f}".rstrip("0").rstrip(".")


@lru_cache(maxsize=4)
def _load_material_products(root: str) -> tuple[import_materials.ProductInput, ...]:
    root_path = Path(root)
    paths = import_materials.get_material_paths(root_path)
    products = import_materials.merge_product_inputs(
        import_materials.parse_product_knowledge(paths.product_knowledge_md)
        + import_materials.parse_product_manual(paths.product_manual_md)
        + import_materials.parse_2026_product_knowledge(paths.product_2026_dir)
    )
    return tuple(products)


@lru_cache(maxsize=4)
def _load_2026_knowledge(root: str) -> tuple[dict[str, Any], ...]:
    knowledge_dir = Path(root) / "资料" / "2026产品知识库"
    if not knowledge_dir.exists():
        return ()

    entries: list[dict[str, Any]] = []
    index_path = knowledge_dir / "00_产品知识总索引.md"
    if index_path.exists():
        entries.extend(_parse_2026_index(index_path))

    solutions_path = knowledge_dir / "01_五大门店解决方案.md"
    if solutions_path.exists():
        entries.extend(_parse_2026_solutions(solutions_path))

    overview_path = knowledge_dir / "02_核心产品卖点速览.md"
    if overview_path.exists():
        entries.extend(_parse_2026_overview(overview_path))

    faq_path = knowledge_dir / "04_产品常见问题精选.md"
    if faq_path.exists():
        entries.extend(_parse_2026_faq_markdown(faq_path))

    shallow_color_path = knowledge_dir / "06_浅柔色素产品档案.md"
    if shallow_color_path.exists():
        entries.extend(_parse_shallow_color_archive(shallow_color_path))

    product_card_path = knowledge_dir / "【法采】2026年产品手卡.xlsx"
    if product_card_path.exists():
        entries.extend(_parse_product_card_workbook(product_card_path))

    shallow_color_xlsx_path = knowledge_dir / "【法采浅柔色素】产品一页纸.xlsx"
    if shallow_color_xlsx_path.exists():
        entries.extend(_parse_shallow_color_workbook(shallow_color_xlsx_path))

    return tuple(entries)


@lru_cache(maxsize=4)
def _load_2026_aliases(root: str) -> dict[str, set[str]]:
    knowledge_dir = Path(root) / "资料" / "2026产品知识库"
    aliases = _static_2026_aliases()
    naming_path = knowledge_dir / "05_产品命名主数据与旧称对照.md"
    if naming_path.exists():
        _merge_alias_map(aliases, _parse_2026_alias_markdown(naming_path))
    naming_xlsx_path = knowledge_dir / "产品命名规则.xlsx"
    if naming_xlsx_path.exists():
        _merge_alias_map(aliases, _parse_product_naming_workbook(naming_xlsx_path))
    return {name: set(values) for name, values in aliases.items()}


def _find_2026_knowledge_points(
    root: str,
    product_names: list[str],
) -> tuple[list[dict[str, Any]], list[str]]:
    entries = _load_2026_knowledge(root)
    if not entries:
        return [], []

    aliases = _load_2026_aliases(root)
    candidate_keys = _candidate_keys(product_names, aliases)
    points: list[dict[str, Any]] = []
    sources: list[str] = []
    for entry in entries:
        if not _entry_matches_product(entry["name"], candidate_keys):
            continue
        points.extend(entry["points"])
        if entry["source"] not in sources:
            sources.append(entry["source"])
    if points:
        alias_names = _alias_names_for_products(product_names, aliases)
        if alias_names:
            _append_2026_point(points, "命名对照", "旧资料名/别名：" + "、".join(alias_names[:8]), "05_产品命名主数据与旧称对照.md")
        for alias_source in _alias_sources(root, product_names, aliases):
            if alias_source not in sources:
                sources.append(alias_source)
    return _clean_merge_points([], points), sources


def _parse_2026_index(path: Path) -> list[dict[str, Any]]:
    text = import_materials._read_text(path)
    entries: list[dict[str, Any]] = []
    for title, block in _split_level3_sections(text):
        if not re.match(r"^[A-Z]\.", title):
            continue
        products: list[str] = []
        values: list[str] = []
        collecting_values = False
        for raw_line in block.splitlines():
            line = raw_line.strip()
            if not line:
                continue
            if line.startswith("核心客户价值"):
                collecting_values = True
                continue
            if not line.startswith("- "):
                continue
            item = import_materials._clean_markdown(line[2:])
            if collecting_values:
                values.append(item)
            else:
                products.extend(_split_product_list(item))
        if not products or not values:
            continue
        scenario = _clean_solution_title(title)
        content = f"{scenario}：" + "；".join(values)
        for product in products:
            points: list[dict[str, Any]] = []
            _append_2026_point(points, "经营场景", content, path.name)
            if points:
                entries.append({"name": product, "points": points, "source": path.name})
    return entries


def _parse_2026_solutions(path: Path) -> list[dict[str, Any]]:
    text = import_materials._read_text(path)
    entries: list[dict[str, Any]] = []
    for title, block in _split_level2_sections(text):
        if not re.match(r"^\d+\.", title):
            continue
        products: list[str] = []
        for line in _extract_heading_bullets(block, "### 代表产品"):
            products.extend(_split_product_list(line))
        if not products:
            continue

        solution_title = _clean_solution_title(title)
        solution_items = _extract_heading_bullets(block, "### 解决方案")
        value_text = _extract_heading_text(block, "### 门店价值")
        for product in products:
            points: list[dict[str, Any]] = []
            if value_text:
                _append_2026_point(points, "门店方案", f"{solution_title}：{value_text}", path.name)
            if solution_items:
                _append_2026_point(points, "解决方案", f"{solution_title}：" + "；".join(solution_items), path.name)
            if points:
                entries.append({"name": product, "points": points, "source": path.name})
    return entries


def _parse_2026_faq_markdown(path: Path) -> list[dict[str, Any]]:
    text = import_materials._read_text(path)
    entries: list[dict[str, Any]] = []
    for title, block in _split_level2_sections(text):
        if title.startswith("产品常见问题"):
            continue
        points: list[dict[str, Any]] = []
        for raw_line in block.splitlines():
            line = raw_line.strip()
            if not line.startswith("- "):
                continue
            body = import_materials._clean_markdown(line[2:])
            match = re.match(r"(.+?)[：:]\s*(.+)$", body)
            if match:
                point_type = f"FAQ：{match.group(1).strip()}"
                content = match.group(2).strip()
            else:
                point_type = "FAQ"
                content = body
            _append_2026_point(points, point_type, content, path.name)
        if points:
            entries.append({"name": title, "points": points, "source": path.name})
    return entries


def _parse_2026_overview(path: Path) -> list[dict[str, Any]]:
    text = import_materials._read_text(path)
    entries: list[dict[str, Any]] = []
    for title, block in _split_level2_sections(text):
        if title == "卖点使用规则":
            continue
        points = []
        for line in block.splitlines():
            line = line.strip()
            if not line.startswith("- "):
                continue
            body = import_materials._clean_markdown(line[2:])
            match = re.match(r"(.+?)[：:]\s*(.+)$", body)
            if not match:
                continue
            point_type = match.group(1).strip()
            content = match.group(2).strip()
            if "待核验" in point_type or "待核验" in content:
                continue
            points.append({
                "point_type": point_type,
                "content": content,
                "priority": len(points) + 1,
                "source": path.name,
            })
        if points:
            entries.append({"name": title, "points": points, "source": path.name})
    return entries


def _parse_shallow_color_archive(path: Path) -> list[dict[str, Any]]:
    text = import_materials._read_text(path)
    points: list[dict[str, Any]] = []

    sales_block = _extract_markdown_block(text, "### 可优先使用")
    for line in sales_block.splitlines():
        line = line.strip()
        if not line.startswith("- "):
            continue
        _append_2026_point(points, "销售表达", line[2:], path.name)

    value_block = _extract_markdown_block(text, "## 四、产品能力与客户价值")
    for line in value_block.splitlines():
        line = line.strip()
        if not line.startswith("|") or "---" in line or "门店痛点" in line:
            continue
        cells = [import_materials._clean_markdown(cell) for cell in line.strip("|").split("|")]
        if len(cells) < 3:
            continue
        content = f"{cells[0]}；{cells[1]}；{cells[2]}"
        _append_2026_point(points, "客户价值", content, path.name)

    if not points:
        return []
    return [{"name": "浅柔色素", "points": points, "source": path.name}]


def _parse_product_card_workbook(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    entries: list[dict[str, Any]] = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "常见问题" in workbook.sheetnames:
            entries.extend(_parse_product_card_faq_sheet(workbook["常见问题"], path.name))
        if "解决方案" in workbook.sheetnames:
            entries.extend(_parse_product_card_solution_sheet(workbook["解决方案"], path.name))

        for sheet_name in workbook.sheetnames:
            if sheet_name in PRODUCT_CARD_SKIP_SHEETS:
                continue
            worksheet = workbook[sheet_name]
            points = _parse_product_card_sheet(worksheet, path.name)
            if points:
                entries.append({
                    "name": _standardize_product_card_name(sheet_name),
                    "points": points,
                    "source": path.name,
                })
    finally:
        workbook.close()
    return entries


def _parse_product_card_sheet(worksheet: Any, source: str) -> list[dict[str, Any]]:
    points: list[dict[str, Any]] = []
    active_label = ""
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        values = [_cell_text(value) for value in row]
        if not any(values):
            continue

        if row_index <= 2:
            title = _join_content(values)
            if "手卡" in title or "一页纸" in title:
                _append_2026_point(points, "资料标题", title, source)

        label_index, label = _find_row_label(values)
        if label:
            active_label = label
            content = _join_content(values[label_index + 1:])
            if not content:
                content = _join_content(value for index, value in enumerate(values) if index != label_index)
            _append_2026_point(points, label, content, source)
            continue

        if active_label and _row_looks_like_continuation(values):
            _append_2026_point(points, active_label, _join_content(values), source)
    return points


def _parse_product_card_faq_sheet(worksheet: Any, source: str) -> list[dict[str, Any]]:
    by_product: dict[str, list[dict[str, Any]]] = {}
    current_product = ""
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        values = [_cell_text(value) for value in row]
        if len(values) < 5:
            continue
        product = values[2]
        question = values[3]
        answer = values[4]
        if product:
            current_product = product
        if not current_product or not question or not answer:
            continue
        points = by_product.setdefault(_standardize_product_card_name(current_product), [])
        _append_2026_point(points, f"FAQ：{question}", answer, source)

    return [
        {"name": product, "points": points, "source": source}
        for product, points in by_product.items()
        if points
    ]


def _parse_product_card_solution_sheet(worksheet: Any, source: str) -> list[dict[str, Any]]:
    rows = [
        [_cell_text(value) for value in row]
        for row in worksheet.iter_rows(min_row=1, max_row=8, values_only=True)
    ]
    scenario_row = _find_sheet_row(rows, "场景")
    solution_row = _find_sheet_row(rows, "法采提供解决方案")
    action_row = _find_sheet_row(rows, "具体行动及结果")
    value_row = _find_sheet_row(rows, "提供价值")
    if not scenario_row:
        return []

    entries_by_product: dict[str, list[dict[str, Any]]] = {}
    for column_index, raw_scenario in enumerate(scenario_row):
        scenario = _strip_leading_number(raw_scenario)
        if not scenario or scenario == "场景":
            continue
        products = _solution_products_for_scenario(scenario)
        if not products:
            continue
        for product in products:
            points = entries_by_product.setdefault(product, [])
            if column_index < len(solution_row):
                _append_2026_point(points, "解决方案", solution_row[column_index], source)
            if column_index < len(action_row):
                _append_2026_point(points, "具体行动及结果", action_row[column_index], source)
            if column_index < len(value_row):
                _append_2026_point(points, "门店价值", f"{scenario}：{value_row[column_index]}", source)

    return [
        {"name": product, "points": points, "source": source}
        for product, points in entries_by_product.items()
        if points
    ]


def _parse_shallow_color_workbook(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    points: list[dict[str, Any]] = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "产品一页纸" in workbook.sheetnames:
            _append_shallow_color_product_sheet(points, workbook["产品一页纸"], path.name)
        if "浅系色素具体颜色以及名称" in workbook.sheetnames:
            _append_shallow_color_names(points, workbook["浅系色素具体颜色以及名称"], path.name)
    finally:
        workbook.close()

    if not points:
        return []
    return [{"name": "浅柔色素", "points": points, "source": path.name}]


def _append_shallow_color_product_sheet(
    points: list[dict[str, Any]],
    worksheet: Any,
    source: str,
) -> None:
    pending_skus: list[str] = []
    for row in worksheet.iter_rows(min_row=1, max_row=24, values_only=True):
        values = [_cell_text(value) for value in row]
        if not any(values):
            continue
        label_index, label = _find_shallow_color_label(values)
        if not label:
            continue
        content_values = [value for value in values[label_index + 1:] if value]
        if label.lower() == "sku":
            pending_skus = content_values
            _append_2026_point(points, "SKU规格", "、".join(pending_skus), source)
            continue
        if "售价" in label and pending_skus:
            prices = content_values
            pairs = [
                f"{sku}：{price}"
                for sku, price in zip(pending_skus, prices)
                if sku and price
            ]
            _append_2026_point(points, "SKU售价", "；".join(pairs), source)
            pending_skus = []
            continue
        if label in {"用途简述", "储存方式", "保质期"}:
            _append_2026_point(points, label, _join_content(content_values), source)


def _append_shallow_color_names(
    points: list[dict[str, Any]],
    worksheet: Any,
    source: str,
) -> None:
    color_names: list[str] = []
    for row in worksheet.iter_rows(min_row=3, max_row=40, values_only=True):
        values = [_cell_text(value) for value in row]
        if len(values) >= 4 and values[3]:
            color_names.append(values[3])
    if color_names:
        _append_2026_point(points, "颜色体系", "浅柔色素颜色：" + "、".join(color_names[:18]), source)


def _append_2026_point(
    points: list[dict[str, Any]],
    point_type: str,
    content: str,
    source: str,
) -> None:
    cleaned = import_materials._clean_markdown(content)
    if not cleaned or _is_risky_2026_content(cleaned) or _is_useless_selling_point(point_type, cleaned):
        return
    if len(cleaned) > 520:
        cleaned = cleaned[:517].rstrip() + "..."
    points.append({
        "point_type": point_type,
        "content": cleaned,
        "priority": len(points) + 1,
        "source": source,
    })


def _split_level2_sections(text: str) -> list[tuple[str, str]]:
    matches = list(re.finditer(r"^##\s+(.+?)\s*$", text, flags=re.MULTILINE))
    sections: list[tuple[str, str]] = []
    for index, match in enumerate(matches):
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        sections.append((import_materials._clean_markdown(match.group(1)), text[start:end].strip()))
    return sections


def _extract_markdown_block(text: str, heading: str) -> str:
    start = text.find(heading)
    if start < 0:
        return ""
    next_heading = re.search(r"^#{2,3}\s+", text[start + len(heading):], flags=re.MULTILINE)
    if not next_heading:
        return text[start + len(heading):]
    end = start + len(heading) + next_heading.start()
    return text[start + len(heading):end]


def _split_level3_sections(text: str) -> list[tuple[str, str]]:
    matches = list(re.finditer(r"^###\s+(.+?)\s*$", text, flags=re.MULTILINE))
    sections: list[tuple[str, str]] = []
    for index, match in enumerate(matches):
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        sections.append((import_materials._clean_markdown(match.group(1)), text[start:end].strip()))
    return sections


def _extract_heading_bullets(block: str, heading: str) -> list[str]:
    heading_block = _extract_markdown_block(block, heading)
    bullets: list[str] = []
    for line in heading_block.splitlines():
        line = line.strip()
        if line.startswith("- "):
            bullets.append(import_materials._clean_markdown(line[2:]))
    return bullets


def _extract_heading_text(block: str, heading: str) -> str:
    heading_block = _extract_markdown_block(block, heading)
    lines = []
    for line in heading_block.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("- "):
            line = line[2:]
        lines.append(import_materials._clean_markdown(line))
    return "；".join(lines)


def _split_product_list(value: str) -> list[str]:
    cleaned = import_materials._clean_markdown(value)
    if "：" in cleaned:
        prefix, suffix = cleaned.split("：", 1)
        if "SKU" in prefix or "产品" in prefix:
            cleaned = suffix
    cleaned = (
        cleaned.replace("及其", "、")
        .replace("等", "")
        .replace(" SKU", "")
        .replace("SKU", "")
        .replace("其他包装产品", "包装产品")
    )
    parts = re.split(r"[、,，/]+", cleaned)
    return [part.strip(" 。；;") for part in parts if part.strip(" 。；;")]


def _clean_solution_title(value: str) -> str:
    return re.sub(r"^[A-Z]\.\s*|^\d+\.\s*", "", value).strip()


def _strip_leading_number(value: str) -> str:
    return re.sub(r"^\d+[、.]\s*", "", import_materials._clean_markdown(value)).strip()


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and abs(value - int(value)) < 0.0001:
        text = str(int(value))
    else:
        text = str(value)
    text = text.replace("\n", " ").strip()
    if text.startswith("=DISPIMG"):
        return ""
    return import_materials._clean_markdown(text)


def _join_content(values: Any) -> str:
    return "；".join(value for value in values if value)


def _find_row_label(values: list[str]) -> tuple[int, str]:
    for index, value in enumerate(values):
        if not value:
            continue
        normalized = _label_key(value)
        for label in PRODUCT_CARD_POINT_LABELS:
            if normalized == _label_key(label):
                return index, label
    return -1, ""


def _find_sheet_row(rows: list[list[str]], label: str) -> list[str]:
    label_key = _label_key(label)
    for row in rows:
        if any(_label_key(value) == label_key for value in row if value):
            return row
    return []


def _find_shallow_color_label(values: list[str]) -> tuple[int, str]:
    labels = {"sku", "售价（到手价）", "用途简述", "保质期", "储存方式"}
    for index, value in enumerate(values):
        if not value:
            continue
        normalized = _label_key(value)
        for label in labels:
            if normalized == _label_key(label):
                return index, label
    return -1, ""


def _label_key(value: str) -> str:
    return re.sub(r"[\s（）()：:]+", "", value).lower()


def _row_looks_like_continuation(values: list[str]) -> bool:
    content = _join_content(values)
    if not content or len(content) < 4:
        return False
    if _is_risky_2026_content(content):
        return False
    return any(
        marker in content
        for marker in ("【", "：", "1.", "1、", "适合", "用于", "比例", "搭配", "场景", "口感", "规格")
    )


def _standardize_product_card_name(sheet_name: str) -> str:
    return PRODUCT_CARD_SHEET_ALIASES.get(sheet_name, sheet_name)


def _solution_products_for_scenario(scenario: str) -> list[str]:
    for key, products in PRODUCT_CARD_SOLUTION_PRODUCTS.items():
        if key in scenario:
            return products
    return []


def _parse_2026_alias_markdown(path: Path) -> dict[str, set[str]]:
    aliases: dict[str, set[str]] = {}
    for raw_line in import_materials._read_text(path).splitlines():
        line = raw_line.strip()
        if not line.startswith("|") or "---" in line:
            continue
        cells = [import_materials._clean_markdown(cell) for cell in line.strip("|").split("|")]
        if not cells or cells[0] in {"旧资料名称", "名称", "系列"}:
            continue
        if len(cells) >= 3 and ("统一" in cells[2] or "确认" in cells[2] or "更名" in cells[2]):
            standard = cells[1]
            for old_name in _split_alias_names(cells[0]):
                _add_alias(aliases, standard, old_name)
        if len(cells) >= 4 and cells[1] in {"SKU", "品类统称"}:
            parent = cells[2].split("/")[0].strip()
            if parent and parent != "色素":
                _add_alias(aliases, parent, cells[0])
    return aliases


def _parse_product_naming_workbook(path: Path) -> dict[str, set[str]]:
    from openpyxl import load_workbook

    aliases: dict[str, set[str]] = {}
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            rows = [[_cell_text(value) for value in row] for row in worksheet.iter_rows(values_only=True)]
            for row_index, values in enumerate(rows):
                if "标准产品名称" in values and "旧资料名称" in values:
                    old_index = values.index("旧资料名称")
                    standard_index = values.index("标准产品名称")
                    for data_values in rows[row_index + 1:]:
                        if len(data_values) <= max(old_index, standard_index):
                            continue
                        standard = data_values[standard_index]
                        for old_name in _split_alias_names(data_values[old_index]):
                            _add_alias(aliases, standard, old_name)
                        continue
                    break
    finally:
        workbook.close()
    return aliases


def _split_alias_names(value: str) -> list[str]:
    cleaned = import_materials._clean_markdown(value)
    cleaned = re.sub(r"（.*?）", "", cleaned)
    return [part.strip() for part in re.split(r"[、,，/]+", cleaned) if part.strip()]


def _static_2026_aliases() -> dict[str, set[str]]:
    aliases: dict[str, set[str]] = {}
    pairs = [
        ("慕斯粉（液）", "慕斯粉"),
        ("彩色拉线膏", "拉线膏"),
        ("彩色拉线膏", "手绘拉线膏"),
        ("彩色拉线膏", "防晕染彩色拉线膏"),
        ("白色翻糖膏", "白色翻糖"),
        ("彩色翻糖膏", "彩色翻糖"),
        ("翻糖压片", "翻糖片"),
        ("翻糖压片", "彩色翻糖片"),
        ("巧克力脆皮酱", "脆皮酱"),
        ("红丝绒", "红丝绒液"),
        ("红丝绒", "红丝绒香精"),
        ("调味糖浆", "薄荷糖浆"),
        ("浅柔色素", "浅系色素"),
        ("浅柔色素", "浅色色素"),
        ("夹心果泥", "果泥"),
        ("夹心果泥", "果泥果酱"),
        ("夹心果泥", "免熬果泥果酱"),
        ("多肉果酱", "果馅（多肉）"),
        ("多肉果酱", "果馅（多肉果酱）"),
        ("夹心芋泥", "芋泥"),
        ("夹心芋泥", "免调味芋泥"),
        ("夹心脆", "巧克力夹心脆"),
        ("奶冻粉", "Q 弹奶冻粉"),
        ("奶冻粉", "Q弹奶冻粉"),
        ("奶冻粉", "晶冻粉"),
        ("水性色素", "胶状色素-小"),
        ("水性色素", "胶状色素-大"),
        ("油性色素", "油性色素-小"),
        ("油性色素", "油性色素-大"),
        ("盒装刀叉", "刀叉"),
        ("盒装刀叉", "2元盒装"),
        ("盒装刀叉", "2.5元盒装"),
        ("盒装刀叉", "5元盒装刀叉"),
        ("盒装刀叉", "1.1浆纸盘"),
        ("零卡糖", "0卡糖粉"),
    ]
    for standard, alias in pairs:
        _add_alias(aliases, standard, alias)
    return aliases


def _merge_alias_map(target: dict[str, set[str]], extra: dict[str, set[str]]) -> None:
    for name, aliases in extra.items():
        for alias in aliases:
            _add_alias(target, name, alias)


def _add_alias(aliases: dict[str, set[str]], standard: str, alias: str) -> None:
    standard = import_materials._clean_markdown(standard)
    alias = import_materials._clean_markdown(alias)
    if not standard or not alias:
        return
    aliases.setdefault(standard, set()).add(alias)
    aliases.setdefault(alias, set()).add(standard)


def _alias_sources(root: str, product_names: list[str], aliases: dict[str, set[str]]) -> list[str]:
    knowledge_dir = Path(root) / "资料" / "2026产品知识库"
    naming_path = knowledge_dir / "05_产品命名主数据与旧称对照.md"
    if not naming_path.exists():
        return []
    for name in product_names:
        names = {name, _base_name(name)}
        for candidate in names:
            related = aliases.get(candidate, set())
            if any(_name_key(alias) != _name_key(candidate) for alias in related):
                return [naming_path.name]
    return []


def _alias_names_for_products(product_names: list[str], aliases: dict[str, set[str]]) -> list[str]:
    names: list[str] = []
    seen: set[str] = set()
    for name in product_names:
        root_keys = {_name_key(name), _name_key(_base_name(name))}
        candidates = {name, _base_name(name)}
        for candidate in list(candidates):
            candidates.update(aliases.get(candidate, set()))
        for candidate in list(candidates):
            candidates.update(aliases.get(candidate, set()))
        for candidate in sorted(candidates, key=lambda value: (_name_key(value), value)):
            candidate_key = _name_key(candidate)
            if candidate_key in root_keys:
                continue
            for alias in sorted(aliases.get(candidate, set())):
                alias_key = _name_key(alias)
                if not alias_key or alias_key in root_keys or alias_key in seen:
                    continue
                names.append(alias)
                seen.add(alias_key)
            if candidate_key and candidate_key not in seen:
                names.append(candidate)
                seen.add(candidate_key)
    return names


def _is_risky_2026_content(content: str) -> bool:
    compact = re.sub(r"\s+", "", content).lower()
    for phrase in RISKY_2026_PHRASES:
        if re.sub(r"\s+", "", phrase).lower() in compact:
            return True
    return False


def _candidate_keys(product_names: list[str], aliases: dict[str, set[str]] | None = None) -> set[str]:
    keys: set[str] = set()
    for name in product_names:
        for candidate in _candidate_names(name, aliases):
            key = _name_key(candidate)
            if key:
                keys.add(key)
    return keys


def _candidate_names(product_name: str, aliases: dict[str, set[str]] | None = None) -> set[str]:
    base = _base_name(product_name)
    candidates = {product_name, base}
    alias_map = _static_2026_aliases()
    if aliases:
        _merge_alias_map(alias_map, aliases)
    for name in list(candidates):
        candidates.update(alias_map.get(name, set()))
        candidates.update(alias_map.get(_base_name(name), set()))
    return candidates


def _entry_matches_product(entry_name: str, candidate_keys: set[str]) -> bool:
    entry_key = _name_key(_base_name(entry_name))
    raw_entry_key = _name_key(entry_name)
    if entry_key in candidate_keys or raw_entry_key in candidate_keys:
        return True
    for key in candidate_keys:
        if key and key in raw_entry_key:
            return True
    return False


def _merge_points(
    primary: list[dict[str, Any]],
    extra: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    merged: list[dict[str, Any]] = []
    by_type: dict[str, dict[str, Any]] = {}
    seen: set[str] = set()
    for point in [*primary, *extra]:
        point_type = point.get("point_type", "卖点")
        content = point.get("content", "")
        key = (import_materials._clean_markdown(point_type), import_materials._clean_markdown(content))
        if not key[1] or key in seen:
            continue
        merged.append({
            **point,
            "point_type": point_type,
            "content": content,
            "priority": len(merged) + 1,
        })
        seen.add(key)
    return merged


def _clean_merge_points(
    primary: list[dict[str, Any]],
    extra: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    merged: list[dict[str, Any]] = []
    by_type: dict[str, dict[str, Any]] = {}
    seen: set[str] = set()
    for point in [*primary, *extra]:
        point_type = import_materials._clean_markdown(point.get("point_type", "卖点"))
        content = _clean_selling_point_content(point.get("content", ""))
        if not content or _is_useless_selling_point(point_type, content):
            continue
        content_key = _selling_point_content_key(content)
        if content_key in seen:
            continue
        type_key = _selling_point_type_key(point_type)
        if type_key in by_type:
            existing = by_type[type_key]
            combined = _merge_selling_point_content(existing["content"], content)
            if combined != existing["content"]:
                existing["content"] = combined
            seen.add(content_key)
            continue

        item = {
            **point,
            "point_type": point_type,
            "content": content,
            "priority": len(merged) + 1,
        }
        merged.append(item)
        by_type[type_key] = item
        seen.add(content_key)
    for index, point in enumerate(merged, start=1):
        point["priority"] = index
    return merged


def _is_useless_selling_point(point_type: str, content: str) -> bool:
    clean_type = import_materials._clean_markdown(point_type)
    clean_content = import_materials._clean_markdown(content)
    if clean_type in USELESS_POINT_TYPES:
        return True
    if any(marker in clean_content for marker in USELESS_CONTENT_MARKERS):
        return True
    if "手卡" in clean_content and len(clean_content) <= 80:
        return True
    if "一页纸" in clean_content and len(clean_content) <= 80:
        return True
    return False


def _clean_selling_point_content(content: str) -> str:
    cleaned = import_materials._clean_markdown(content)
    cleaned = re.sub(r"\s*\?{2,}\s*", " ", cleaned)
    return re.sub(r"\s+", " ", cleaned).strip()


def _selling_point_type_key(point_type: str) -> str:
    return re.sub(r"[\s：:，,。；;、（）()]+", "", point_type).lower()


def _selling_point_content_key(content: str) -> str:
    cleaned = import_materials._clean_markdown(content)
    return re.sub(r"[\s：:，,。；;、（）()【】\[\]\"'“”‘’]+", "", cleaned).lower()


def _merge_selling_point_content(existing: str, incoming: str) -> str:
    existing_key = _selling_point_content_key(existing)
    incoming_key = _selling_point_content_key(incoming)
    if not incoming_key or incoming_key in existing_key:
        return existing
    if existing_key and existing_key in incoming_key:
        return incoming
    combined = existing.rstrip("；;。") + "；" + incoming
    if len(combined) > 900:
        combined = combined[:897].rstrip("；;，,。 ") + "..."
    return combined


def _find_material_product(
    product_name: str,
    products: tuple[import_materials.ProductInput, ...],
) -> import_materials.ProductInput | None:
    requested = _name_key(product_name)

    for product in products:
        if _name_key(product.name) == requested:
            return product

    requested_base = _name_key(_base_name(product_name))
    for product in products:
        if _name_key(product.name) == requested_base:
            return product

    aliases = _aliases_by_product()
    for product in products:
        keys = {_name_key(product.name), *{_name_key(alias) for alias in aliases.get(product.name, [])}}
        if requested in keys or requested_base in keys:
            return product

    for product in products:
        product_key = _name_key(product.name)
        if product_key and (product_key in requested or requested in product_key):
            return product
    return None


def _resolve_price_rule_name(product_name: str) -> str:
    requested = _name_key(product_name)
    requested_base = _name_key(_base_name(product_name))
    candidates = list(import_materials.PRICE_RULES)

    for name in candidates:
        if _name_key(name) in {requested, requested_base}:
            return name

    aliases = _aliases_by_product()
    for name in candidates:
        keys = {_name_key(name), *{_name_key(alias) for alias in aliases.get(name, [])}}
        if requested in keys or requested_base in keys:
            return name

    for name in candidates:
        key = _name_key(name)
        if key and (key in requested or requested_base in key):
            return name
    return _base_name(product_name)


def _aliases_by_product() -> dict[str, list[str]]:
    aliases = dict(import_materials.MANUAL_ALIASES)
    aliases.setdefault("水性色素", []).append("水性色素（胶状）")
    aliases.setdefault("翻糖膏", []).append("防潮翻糖膏")
    return aliases


def _sku_from_price_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "product": row.get("product", ""),
        "spec": row.get("spec", ""),
        "price": _clean_price(row.get("tag_price")),
        "daily_price": _clean_price(row.get("daily_price")),
        "activity_prices": _clean_activity_prices(row.get("activity_prices", [])),
        "count": _clean_count(row.get("count")),
        "category": row.get("category", ""),
    }


def _override_sku(rule_name: str, override: float) -> dict[str, Any]:
    return {
        "product": rule_name,
        "spec": "默认售价",
        "price": import_materials._round_price(override),
        "daily_price": None,
        "activity_prices": [],
        "count": None,
        "category": "",
    }


def _clean_activity_prices(activity_prices: Any) -> list[dict[str, Any]]:
    if not isinstance(activity_prices, list):
        return []
    cleaned: list[dict[str, Any]] = []
    for activity in activity_prices:
        if not isinstance(activity, dict):
            continue
        item = {
            "mechanism": activity.get("mechanism", ""),
            "tag_price": _clean_price(activity.get("tag_price")),
            "discount": _clean_activity_text(activity.get("discount", "")),
            "coupon": _clean_activity_text(activity.get("coupon", "")),
            "activity_price": _clean_price(activity.get("activity_price")),
            "final_price": _clean_price(activity.get("final_price")),
            "unit_price": _clean_price(activity.get("unit_price")),
            "single_activity": _clean_activity_text(activity.get("single_activity", "")),
            "count": _clean_count(activity.get("count")),
        }
        if item["tag_price"] is None and item["activity_price"] is None and item["final_price"] is None:
            continue
        cleaned.append(item)
    return cleaned


def _clean_activity_text(value: Any) -> str:
    if value is None:
        return ""
    return import_materials._clean_markdown(str(value))


def _clean_price(value: Any) -> float | None:
    if not isinstance(value, (int, float)) or value <= 0:
        return None
    return import_materials._round_price(float(value))


def _clean_count(value: Any) -> int | float | None:
    if not isinstance(value, (int, float)):
        return None
    if abs(value - int(value)) < 0.0001:
        return int(value)
    return value


def _base_name(value: str) -> str:
    name = re.sub(r"[（(].*?[）)]", "", value or "")
    return name.strip()


def _name_key(value: str) -> str:
    return re.sub(r"[\s（）()、/\\-]+", "", value or "").lower()
